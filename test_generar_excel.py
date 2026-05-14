import pandas as pd
import unicodedata
import os
from glob import glob
from openpyxl import Workbook

def _norm(s):
    return unicodedata.normalize('NFKD', str(s).strip().lower()).encode('ascii', 'ignore').decode('ascii')

def _normalize_column_name(name):
    if pd.isna(name):
        return ''
    normalized = unicodedata.normalize('NFKD', str(name))
    normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
    return normalized.lower().replace(' ', '').replace('_', '').replace('-', '')

def _normalize_value(value):
    normalized = unicodedata.normalize('NFKD', str(value))
    normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
    normalized = normalized.lower()
    normalized = ''.join(c for c in normalized if c.isalnum())
    return normalized

# Archivos de prueba
archivos = [
    'data/raw/FORMATOS RA CICLO UNO RC/FormatoRA_AdmonPub_VNAL.xlsx',
    'data/raw/FORMATOS RA CICLO UNO RC/FormatoRA_NegIntl_PMED.xlsx',
    'data/raw/FORMATOS RA CICLO UNO RC/FormatoRA_ContPub_HMED.xlsx',
    'data/raw/FORMATOS RA CICLO UNO RC/FormatoRA_ComunicacionSocialPeriodismo_PBOG.xlsx'
]

wb = Workbook()

for filepath in archivos:
    filename = os.path.basename(filepath).replace('.xlsx', '')
    print(f"Procesando: {filename}")
    
    # Crear hoja
    ws = wb.create_sheet(title=filename[:31])
    
    # Leer datos
    df = pd.read_excel(filepath, sheet_name='Paso 5 Estrategias micro', header=1, engine='openpyxl')
    df.columns = [_normalize_column_name(c) for c in df.columns]
    
    asig_col = None
    for c in df.columns:
        if 'nombreasignaturaomodulo' in c:
            asig_col = c
            break
    
    # Obtener todas las asignaturas
    if asig_col:
        all_vals = df[asig_col].dropna()
        
        # Clasificar: asignaturas vs filas de totales
        asignaturas = []
        filas_totales = []
        
        for v in all_vals:
            v_str = str(v).strip()
            try:
                float(v_str)
                filas_totales.append(v)
            except:
                asignaturas.append(v)
        
        # Normalizar
        asigs_normalizadas = pd.Series(asignaturas).apply(_normalize_value)
        asigs_calc = asigs_normalizadas.nunique()
        
        # Leer total oficial
        raw = pd.read_excel(filepath, sheet_name='Paso 5 Estrategias micro', header=None, engine='openpyxl')
        asigs_oficial = 0
        
        for r in range(raw.shape[0]):
            for c in range(raw.shape[1]):
                cell = raw.iloc[r, c]
                if pd.notna(cell):
                    cn = _norm(str(cell))
                    if asigs_oficial == 0 and ('total modulos' in cn or 'total asignaturas del programa' in cn or ('total materias' in cn and 'asignatura' in cn)):
                        next_col = c + 1
                        if next_col < raw.shape[1]:
                            raw_val = raw.iloc[r, next_col]
                            try:
                                asigs_oficial = int(float(raw_val)) if isinstance(raw_val, (int, float)) else int(float(str(raw_val).strip()))
                            except:
                                asigs_oficial = 0
        
        # Escribir información
        ws['A1'] = 'Archivo'
        ws['B1'] = filename
        ws['A2'] = 'Asignaturas Calculadas'
        ws['B2'] = asigs_calc
        ws['A3'] = 'Asignaturas Oficial'
        ws['B3'] = asigs_oficial
        ws['A4'] = 'Diferencia'
        ws['B4'] = asigs_calc - asigs_oficial
        ws['A5'] = ''
        ws['A6'] = 'Listado de Asignaturas Contadas:'
        ws['A7'] = 'Nombre Original'
        ws['B7'] = 'Nombre Normalizado'
        
        row = 8
        for i, (orig, norm) in enumerate(zip(asignaturas, asigs_normalizadas)):
            ws[f'A{row}'] = orig
            ws[f'B{row}'] = norm
            row += 1
        
        # Agregar filas de totales encontradas
        row += 2
        ws[f'A{row}'] = 'Filas de Totales (excluidas):'
        row += 1
        ws[f'A{row}'] = 'Valor'
        for ft in filas_totales:
            row += 1
            ws[f'A{row}'] = ft

# Guardar
output = 'test_conteo_asignaturas.xlsx'
wb.save(output)
print(f"\nGuardado: {output}")
print("Excel generado con la información de prueba")