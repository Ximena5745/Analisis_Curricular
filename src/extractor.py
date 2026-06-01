"""
Módulo de extracción de datos desde archivos Excel microcurriculares.

Este módulo proporciona la clase ExcelExtractor para leer y procesar
archivos Excel con diseños microcurriculares, extrayendo competencias,
resultados de aprendizaje, estrategias pedagógicas y más.
"""

import logging
import re
from pathlib import Path
from typing import Dict, Optional, List, Tuple
import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

import sys
sys.path.append(str(Path(__file__).parent.parent))

from config import EXCEL_SHEETS, HEADER_ROWS, EXPECTED_COLUMNS

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ExcelExtractor:
    """
    Extrae datos de archivos Excel microcurriculares.

    Esta clase maneja la lectura de archivos Excel con diferentes estructuras,
    normaliza nombres de columnas, detecta headers automáticamente y extrae
    información de competencias, resultados de aprendizaje y estrategias.

    Attributes:
        file_path (Path): Ruta al archivo Excel
        workbook (openpyxl.Workbook): Libro de Excel cargado
        programa_nombre (str): Nombre del programa extraído del archivo

    Example:
        >>> extractor = ExcelExtractor('FormatoRA_AdmonEmpresas_PBOG.xlsx')
        >>> data = extractor.extract_all()
        >>> print(f"Competencias: {len(data['competencias'])}")
    """

    def __init__(self, file_path: str):
        """
        Inicializa el extractor con la ruta del archivo.

        Args:
            file_path (str): Ruta al archivo Excel

        Raises:
            FileNotFoundError: Si el archivo no existe
            InvalidFileException: Si el archivo no es un Excel válido
        """
        self.file_path = Path(file_path)

        if not self.file_path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {self.file_path}")

        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            logger.info(f"Archivo cargado: {self.file_path.name}")
        except InvalidFileException as e:
            logger.error(f"Error al cargar archivo: {e}")
            raise

        # Extraer nombre del programa del nombre del archivo
        self.programa_nombre = self._extract_programa_name()
        logger.info(f"Programa detectado: {self.programa_nombre}")

    def _extract_programa_name(self) -> str:
        """
        Extrae el nombre del programa desde el nombre del archivo.

        Formato esperado: FormatoRA_NombrePrograma_PBOG.xlsx

        Returns:
            str: Nombre del programa
        """
        filename = self.file_path.stem  # Sin extensión

        # Intentar extraer nombre entre FormatoRA_ y _PBOG
        match = re.search(r'FormatoRA_(.+?)_[A-Z]{4}', filename)
        if match:
            return match.group(1).replace('_', ' ')

        # Si no coincide, usar el nombre completo sin extensión
        return filename

    def _extract_sede_modalidad(self) -> dict:
        """
        Extrae código de sede y modalidad desde el nombre del archivo.

        FormatoRA_ContPub_PBOG.xlsx -> {'sede': 'Bogotá', 'modalidad': 'Presencial', 'codigo': 'PBOG'}

        Returns:
            dict: codigo, sede, modalidad
        """
        CODIGOS = {
            'PBOG': {'sede': 'Bogotá', 'modalidad': 'Presencial'},
            'VNAL': {'sede': 'Nacional', 'modalidad': 'Virtual'},
            'PMED': {'sede': 'Medellín', 'modalidad': 'Presencial'},
            'HMED': {'sede': 'Medellín', 'modalidad': 'Híbrido'},
            'HBOG': {'sede': 'Bogotá', 'modalidad': 'Híbrido'},
        }
        match = re.search(r'_([A-Z]{4})(?:\.xlsx)?$', self.file_path.stem)
        codigo = match.group(1) if match else 'UNKN'
        return {'codigo': codigo, **CODIGOS.get(codigo, {'sede': 'N/A', 'modalidad': 'N/A'})}

    def _normalize_column_name(self, col_name: str) -> str:
        """
        Normaliza nombres de columnas para comparación.

        - Remueve espacios extras
        - Convierte a minúsculas
        - Remueve tildes
        - Remueve caracteres especiales

        Args:
            col_name (str): Nombre de columna original

        Returns:
            str: Nombre normalizado
        """
        if pd.isna(col_name):
            return ""

        # Convertir a string y lowercase
        col_name = str(col_name).lower().strip()

        # Remover tildes
        replacements = {
            'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
            'ñ': 'n', 'ü': 'u'
        }
        for old, new in replacements.items():
            col_name = col_name.replace(old, new)

        # Remover caracteres especiales excepto espacios
        col_name = re.sub(r'[^a-z0-9\s]', '', col_name)

        # Comprimir espacios múltiples
        col_name = re.sub(r'\s+', ' ', col_name)

        return col_name

    def _find_header_row(self, sheet, expected_columns: List[str],
                         max_rows: int = 10) -> Optional[int]:
        """
        Encuentra automáticamente la fila de headers en una hoja.

        Busca la fila que contiene la mayor cantidad de columnas esperadas.

        Args:
            sheet: Hoja de Excel
            expected_columns: Lista de nombres de columnas esperadas
            max_rows: Máximo de filas a revisar

        Returns:
            Optional[int]: Índice de la fila de header (0-indexed) o None
        """
        # Normalizar columnas esperadas
        expected_normalized = [self._normalize_column_name(col)
                              for col in expected_columns]

        best_match = None
        best_score = 0

        for row_idx in range(max_rows):
            row_values = [cell.value for cell in sheet[row_idx + 1]]
            row_normalized = [self._normalize_column_name(val)
                            for val in row_values]

            # Contar coincidencias
            score = sum(1 for exp in expected_normalized
                       if exp in row_normalized)

            if score > best_score:
                best_score = score
                best_match = row_idx

        if best_match is not None and best_score >= len(expected_columns) * 0.5:
            logger.debug(f"Header encontrado en fila {best_match + 1} "
                        f"(score: {best_score}/{len(expected_columns)})")
            return best_match

        logger.warning(f"No se encontró header válido. Mejor coincidencia: "
                      f"fila {best_match + 1 if best_match else 'N/A'} "
                      f"(score: {best_score}/{len(expected_columns)})")
        return None

    def _find_sheet(self, sheet_name: str) -> str:
        """
        Busca una hoja por nombre con tolerancia a trailing spaces.

        Args:
            sheet_name (str): Nombre configurado de la hoja

        Returns:
            str: Nombre real de la hoja en el archivo

        Raises:
            ValueError: Si no se encuentra ninguna coincidencia
        """
        if sheet_name in self.workbook.sheetnames:
            return sheet_name
        for sn in self.workbook.sheetnames:
            if sn.strip() == sheet_name.strip() or sn.startswith(sheet_name.strip()):
                return sn
        raise ValueError(f"Hoja '{sheet_name}' no encontrada. "
                        f"Hojas disponibles: {self.workbook.sheetnames}")

    def _read_sheet_as_dataframe(self, sheet_name: str,
                                 header_row: Optional[int] = None,
                                 expected_columns: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Lee una hoja de Excel como DataFrame con detección automática de headers.

        Args:
            sheet_name (str): Nombre de la hoja
            header_row (Optional[int]): Fila del header (0-indexed).
                                       Si es None, se detecta automáticamente
            expected_columns (Optional[List[str]]): Columnas esperadas para validación

        Returns:
            pd.DataFrame: Datos de la hoja

        Raises:
            ValueError: Si la hoja no existe o no se encuentra header
        """
        sheet_name = self._find_sheet(sheet_name)

        sheet = self.workbook[sheet_name]

        # Detectar header si no se proporciona
        if header_row is None and expected_columns:
            header_row = self._find_header_row(sheet, expected_columns)
            if header_row is None:
                logger.warning(f"Usando primera fila como header por defecto")
                header_row = 0
        elif header_row is None:
            header_row = 0

        # Leer con pandas
        df = pd.read_excel(
            self.file_path,
            sheet_name=sheet_name,
            header=header_row,
            engine='openpyxl'
        )

        # Limpiar columnas vacías
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        # Remover filas completamente vacías
        df = df.dropna(how='all')

        logger.debug(f"Hoja '{sheet_name}' leída: {len(df)} filas, "
                    f"{len(df.columns)} columnas")

        return df

    def extract_competencias(self) -> pd.DataFrame:
        """
        Extrae competencias de la hoja 'Paso 2 Redacción competen'.

        Returns:
            pd.DataFrame: DataFrame con competencias
                Columnas: No., Verbo competencia, Objeto conceptual,
                         Finalidad, Condición de contexto o referencia,
                         Redacción competencia, Tipo de competencia

        Example:
            >>> df = extractor.extract_competencias()
            >>> print(df[['No.', 'Redacción competencia']].head())
        """
        sheet_name = EXCEL_SHEETS['COMPETENCIAS']
        expected_cols = EXPECTED_COLUMNS['COMPETENCIAS']
        header_row = HEADER_ROWS.get('COMPETENCIAS', 0)

        try:
            df = self._read_sheet_as_dataframe(
                sheet_name,
                header_row=header_row,
                expected_columns=expected_cols
            )

            # Agregar metadatos
            sede_data = self._extract_sede_modalidad()
            df['Programa'] = self.programa_nombre
            df['Sede'] = sede_data['sede']
            df['Modalidad'] = sede_data['modalidad']
            df['Codigo_Sede'] = sede_data['codigo']
            df['Archivo'] = self.file_path.name

            logger.info(f"Extraídas {len(df)} competencias")
            return df

        except Exception as e:
            logger.error(f"Error extrayendo competencias: {e}")
            return pd.DataFrame()

    def extract_resultados_aprendizaje(self) -> pd.DataFrame:
        """
        Extrae resultados de aprendizaje de la hoja 'Paso 3 Redacción RA'.

        Returns:
            pd.DataFrame: DataFrame con resultados de aprendizaje
                Columnas: Competencia por desarrollar, Número de resultado,
                         TipoSaber, SaberAsociado, Taxonomía, Dominio Asociado,
                         Nivel Dominio, Verbo RA, Resultados Aprendizaje

        Example:
            >>> df = extractor.extract_resultados_aprendizaje()
            >>> print(df[['TipoSaber', 'Resultados Aprendizaje']].head())
        """
        sheet_name = EXCEL_SHEETS['RESULTADOS_APRENDIZAJE']
        expected_cols = EXPECTED_COLUMNS['RESULTADOS_APRENDIZAJE']
        header_row = HEADER_ROWS.get('RESULTADOS_APRENDIZAJE', 0)

        try:
            df = self._read_sheet_as_dataframe(
                sheet_name,
                header_row=header_row,
                expected_columns=expected_cols
            )

            # Agregar metadatos
            sede_data = self._extract_sede_modalidad()
            df['Programa'] = self.programa_nombre
            df['Sede'] = sede_data['sede']
            df['Modalidad'] = sede_data['modalidad']
            df['Codigo_Sede'] = sede_data['codigo']
            df['Archivo'] = self.file_path.name

            logger.info(f"Extraídos {len(df)} resultados de aprendizaje")
            return df

        except Exception as e:
            logger.error(f"Error extrayendo RA: {e}")
            return pd.DataFrame()

    def extract_estrategias_meso(self) -> pd.DataFrame:
        """
        Extrae estrategias mesocurriculares de la hoja 'Paso 4 Estrategias mesocurricu'.

        Returns:
            pd.DataFrame: DataFrame con estrategias mesocurriculares
        """
        sheet_name = EXCEL_SHEETS['ESTRATEGIAS_MESO']
        expected_cols = EXPECTED_COLUMNS['ESTRATEGIAS_MESO']
        header_row = HEADER_ROWS.get('ESTRATEGIAS_MESO', 0)

        try:
            df = self._read_sheet_as_dataframe(
                sheet_name,
                header_row=header_row,
                expected_columns=expected_cols
            )

            sede_data = self._extract_sede_modalidad()
            df['Programa'] = self.programa_nombre
            df['Sede'] = sede_data['sede']
            df['Modalidad'] = sede_data['modalidad']
            df['Codigo_Sede'] = sede_data['codigo']
            df['Archivo'] = self.file_path.name

            logger.info(f"Extraídas {len(df)} estrategias mesocurriculares")
            return df

        except Exception as e:
            logger.error(f"Error extrayendo estrategias meso: {e}")
            return pd.DataFrame()

    def extract_estrategias_micro(self) -> pd.DataFrame:
        """
        Extrae estrategias microcurriculares de la hoja 'Paso 5 Estrategias micro'.

        Returns:
            pd.DataFrame: DataFrame con estrategias microcurriculares
        """
        sheet_name = EXCEL_SHEETS['ESTRATEGIAS_MICRO']
        expected_cols = EXPECTED_COLUMNS['ESTRATEGIAS_MICRO']
        header_row = HEADER_ROWS.get('ESTRATEGIAS_MICRO', 0)

        try:
            df = self._read_sheet_as_dataframe(
                sheet_name,
                header_row=header_row,
                expected_columns=expected_cols
            )

            sede_data = self._extract_sede_modalidad()
            df['Programa'] = self.programa_nombre
            df['Sede'] = sede_data['sede']
            df['Modalidad'] = sede_data['modalidad']
            df['Codigo_Sede'] = sede_data['codigo']
            df['Archivo'] = self.file_path.name

            logger.info(f"Extraídas {len(df)} estrategias microcurriculares")
            return df

        except Exception as e:
            logger.error(f"Error extrayendo estrategias micro: {e}")
            return pd.DataFrame()

    def extract_perfil_egreso(self) -> pd.DataFrame:
        """
        Extrae perfil de egreso de la hoja 'Paso1 Analisis perfil egreso'.

        Returns:
            pd.DataFrame: DataFrame con perfil de egreso
                Columnas: Programa, Modalidad, Perfil profesional, Perfil ocupacional,
                         Saber, SaberHacer, SaberSer, Áreas profesionales,
                         Tareas profesionales, Poblaciones actuación, Valor agregado
        """
        sheet_name = EXCEL_SHEETS['PERFIL_EGRESO']
        expected_cols = EXPECTED_COLUMNS['PERFIL_EGRESO']
        header_row = HEADER_ROWS.get('PERFIL_EGRESO', 0)

        try:
            df = self._read_sheet_as_dataframe(
                sheet_name,
                header_row=header_row,
                expected_columns=expected_cols
            )

            sede_data = self._extract_sede_modalidad()
            df['Programa_Archivo'] = self.programa_nombre
            df['Sede'] = sede_data['sede']
            df['Modalidad_Archivo'] = sede_data['modalidad']
            df['Codigo_Sede'] = sede_data['codigo']

            logger.info(f"Extraído perfil de egreso: {len(df)} filas")
            return df

        except Exception as e:
            logger.error(f"Error extrayendo perfil de egreso: {e}")
            return pd.DataFrame()

    def extract_all(self) -> Dict[str, any]:
        """
        Extrae todos los datos del archivo Excel.

        Returns:
            Dict con estructura:
            {
                'metadata': {
                    'programa': str,
                    'archivo': str,
                    'ruta': str
                },
                'competencias': pd.DataFrame,
                'resultados_aprendizaje': pd.DataFrame,
                'estrategias_meso': pd.DataFrame,
                'estrategias_micro': pd.DataFrame
            }

        Example:
            >>> data = extractor.extract_all()
            >>> print(f"Programa: {data['metadata']['programa']}")
            >>> print(f"Competencias: {len(data['competencias'])}")
        """
        logger.info(f"Iniciando extracción completa de {self.file_path.name}")

        data = {
            'metadata': {
                'programa': self.programa_nombre,
                'archivo': self.file_path.name,
                'ruta': str(self.file_path),
                'sede': self._extract_sede_modalidad()['sede'],
                'modalidad': self._extract_sede_modalidad()['modalidad'],
                'codigo_sede': self._extract_sede_modalidad()['codigo']
            },
            'competencias': self.extract_competencias(),
            'resultados_aprendizaje': self.extract_resultados_aprendizaje(),
            'estrategias_meso': self.extract_estrategias_meso(),
            'estrategias_micro': self.extract_estrategias_micro(),
            'perfil_egreso': self.extract_perfil_egreso()
        }

        logger.info(f"Extracción completa finalizada para {self.programa_nombre}")
        return data

    def validate_structure(self) -> Dict[str, any]:
        """
        Valida que el archivo tenga la estructura esperada.

        Returns:
            Dict con estructura:
            {
                'valid': bool,
                'errors': List[str],
                'warnings': List[str],
                'info': Dict
            }

        Example:
            >>> result = extractor.validate_structure()
            >>> if not result['valid']:
            >>>     print("Errores:", result['errors'])
        """
        errors = []
        warnings = []
        info = {}

        # Verificar hojas requeridas
        required_sheets = [
            EXCEL_SHEETS['COMPETENCIAS'],
            EXCEL_SHEETS['RESULTADOS_APRENDIZAJE']
        ]

        for sheet_name in required_sheets:
            try:
                self._find_sheet(sheet_name)
            except ValueError:
                errors.append(f"Hoja requerida no encontrada: '{sheet_name}'")

        # Verificar hojas opcionales
        optional_sheets = [
            EXCEL_SHEETS['ESTRATEGIAS_MESO'],
            EXCEL_SHEETS['ESTRATEGIAS_MICRO']
        ]

        for sheet_name in optional_sheets:
            try:
                self._find_sheet(sheet_name)
            except ValueError:
                warnings.append(f"Hoja opcional no encontrada: '{sheet_name}'")

        # Intentar extraer datos para validar estructura
        try:
            competencias = self.extract_competencias()
            info['num_competencias'] = len(competencias)

            if len(competencias) == 0:
                warnings.append("No se encontraron competencias")
        except Exception as e:
            errors.append(f"Error al extraer competencias: {str(e)}")

        try:
            ra = self.extract_resultados_aprendizaje()
            info['num_resultados_aprendizaje'] = len(ra)

            if len(ra) == 0:
                warnings.append("No se encontraron resultados de aprendizaje")
        except Exception as e:
            errors.append(f"Error al extraer RA: {str(e)}")

        valid = len(errors) == 0

        return {
            'valid': valid,
            'errors': errors,
            'warnings': warnings,
            'info': info
        }

    def get_summary(self) -> str:
        """
        Retorna un resumen legible del contenido extraído.

        Returns:
            str: Resumen formateado
        """
        validation = self.validate_structure()

        summary = f"""
╔═══════════════════════════════════════════════════════════╗
║  RESUMEN DE EXTRACCIÓN                                   ║
╚═══════════════════════════════════════════════════════════╝

Programa: {self.programa_nombre}
Archivo: {self.file_path.name}

Estado: {'✅ VÁLIDO' if validation['valid'] else '❌ INVÁLIDO'}

Datos Extraídos:
  - Competencias: {validation['info'].get('num_competencias', 0)}
  - Resultados de Aprendizaje: {validation['info'].get('num_resultados_aprendizaje', 0)}

"""

        if validation['errors']:
            summary += "❌ Errores:\n"
            for error in validation['errors']:
                summary += f"  - {error}\n"

        if validation['warnings']:
            summary += "\n⚠️  Advertencias:\n"
            for warning in validation['warnings']:
                summary += f"  - {warning}\n"

        return summary


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def extract_from_file(file_path: str) -> Dict:
    """
    Función auxiliar para extraer datos de un archivo.

    Args:
        file_path (str): Ruta al archivo Excel

    Returns:
        Dict: Datos extraídos
    """
    extractor = ExcelExtractor(file_path)
    return extractor.extract_all()


def batch_extract(input_folder: str, pattern: str = "*.xlsx") -> List[Dict]:
    """
    Extrae datos de múltiples archivos en una carpeta.

    Args:
        input_folder (str): Carpeta con archivos Excel
        pattern (str): Patrón de archivos a procesar

    Returns:
        List[Dict]: Lista de datos extraídos de cada archivo
    """
    folder = Path(input_folder)
    files = list(folder.glob(pattern))

    logger.info(f"Encontrados {len(files)} archivos en {input_folder}")

    results = []
    for file_path in files:
        try:
            logger.info(f"Procesando {file_path.name}...")
            data = extract_from_file(str(file_path))
            results.append(data)
        except Exception as e:
            logger.error(f"Error procesando {file_path.name}: {e}")

    logger.info(f"Procesados {len(results)}/{len(files)} archivos exitosamente")
    return results


# ============================================================================
# EJEMPLO DE USO
# ============================================================================

if __name__ == '__main__':
    # Configurar logging más detallado
    logging.basicConfig(level=logging.DEBUG)

    # Ejemplo de uso
    print("╔═══════════════════════════════════════════════════════════╗")
    print("║  EXTRACTOR DE DATOS MICROCURRICULARES                    ║")
    print("╚═══════════════════════════════════════════════════════════╝\n")

    # Reemplazar con ruta real
    example_file = "data/raw/FormatoRA_AdmonEmpresas_PBOG.xlsx"

    if Path(example_file).exists():
        try:
            extractor = ExcelExtractor(example_file)

            # Mostrar resumen
            print(extractor.get_summary())

            # Extraer todos los datos
            data = extractor.extract_all()

            # Mostrar primeras competencias
            if not data['competencias'].empty:
                print("\n📌 Primeras 3 competencias:")
                print(data['competencias'][['No.', 'Redacción competencia']].head(3))

        except Exception as e:
            print(f"❌ Error: {e}")
    else:
        print(f"⚠️  Archivo de ejemplo no encontrado: {example_file}")
        print("    Coloca un archivo Excel en data/raw/ para probarlo")
