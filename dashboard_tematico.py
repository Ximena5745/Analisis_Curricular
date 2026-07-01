"""
DASHBOARD INTERACTIVO - Analisis Tematico Avanzado
Compatible con Streamlit Cloud y ejecucion local.

Ejecutar local: streamlit run dashboard_tematico.py
Deploy cloud:   Subir repo a GitHub y conectar en share.streamlit.io
"""

import streamlit as st
import streamlit.components.v1 as st_components
from streamlit_option_menu import option_menu
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.figure_factory as ff
import plotly.graph_objects as go
from collections import Counter
import re
import json
import unicodedata
from typing import Dict, Optional
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
from src.nucleos_cleaner import filtrar_nucleos_dataframe, limpiar_nucleo, es_nucleo_valido
from src.perfil_coverage_analyzer import analizar_cobertura_perfil_completa
from scipy.stats import entropy
import io
import warnings
warnings.filterwarnings('ignore')

PALETA_DOUBLE_SPLIT = ['#0FFF8B', '#0FBFFF', '#0F83FF', '#FF0F83', '#FF8B0F']
PALETA_AZUL = ['#092196', '#1A2663', '#748BFC', '#A8B7FF', '#DBE1FF']
COLORES_MODALIDAD = {
    'Presencial': '#0F83FF',
    'Virtual': '#748BFC',
    'Híbrido': '#0FFF8B'
}
COLORES_SEDE = {
    'Bogotá': '#0F83FF',
    'Medellín': '#1565C0',
    'Nacional': '#0FFF8B'
}


# _limpiar_nucleo, _es_nucleo_valido, _PATRON_NO_NUCLEO e _INICIO_INVALIDO
# se usan desde src.nucleos_cleaner (importado arriba).


def render_icon_svg(name: str, color: str = "#0f3460", size: int = 28) -> str:
    icons = {
        'document': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M6 2H14L18 6V22H6V2Z" fill="{color}" fill-opacity="0.12"/><path d="M14 2V6H18" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/><path d="M6 2H14L18 6V22H6V2Z" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/><path d="M9 12H15" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/><path d="M9 16H15" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/></svg>''',
        'trend': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M4 16L9 11L13 15L20 8" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/><path d="M4 20H20" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/><path d="M4 4V20" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/></svg>''',
        'bloom': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><rect x="4" y="4" width="16" height="16" rx="4" fill="{color}" fill-opacity="0.12"/><path d="M8 15H16" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/><path d="M8 11H14" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/><path d="M8 7H12" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/></svg>''',
        'search': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><circle cx="11" cy="11" r="6" stroke="{color}" stroke-width="1.8"/><path d="M16.5 16.5L20 20" stroke="{color}" stroke-width="1.8" stroke-linecap="round"/></svg>''',
        'grid': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><rect x="4" y="4" width="6" height="6" rx="1.5" fill="{color}" fill-opacity="0.16" stroke="{color}" stroke-width="1.8"/><rect x="14" y="4" width="6" height="6" rx="1.5" fill="{color}" fill-opacity="0.08" stroke="{color}" stroke-width="1.8"/><rect x="4" y="14" width="6" height="6" rx="1.5" fill="{color}" fill-opacity="0.08" stroke="{color}" stroke-width="1.8"/><rect x="14" y="14" width="6" height="6" rx="1.5" fill="{color}" fill-opacity="0.16" stroke="{color}" stroke-width="1.8"/></svg>''',
        'settings': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><circle cx="12" cy="12" r="3" stroke="{color}" stroke-width="1.8"/><path d="M19.4 15A1.66 1.66 0 0 0 20 13.7 1.67 1.67 0 0 0 19.4 12.4L21 10.6 19.4 8.8 17.5 9.1A1.67 1.67 0 0 0 16 8 1.67 1.67 0 0 0 14.5 9.1L12.6 8.8 11 10.6 12.6 12.4A1.67 1.67 0 0 0 12 13.7 1.67 1.67 0 0 0 12.5 15L11 16.9 12.6 18.7 14.5 18.4A1.67 1.67 0 0 0 16 20 1.67 1.67 0 0 0 17.5 18.9L19.4 19.2 21 17.4 19.4 15Z" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>''',
        'upload': f'''<svg width="{size}" height="{size}" viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg"><path d="M7 16C4.239 16 2 13.761 2 11c0-2.596 1.948-4.74 4.4-4.938C7.082 4.524 8.918 3 11 3c2.747 0 5 2.253 5 5v1h2c2.761 0 5 2.239 5 5s-2.239 5-5 5" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/><path d="M11 19V10M8 13L11 10L14 13" stroke="{color}" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"/></svg>'''
    }
    return icons.get(name, '')


# ============================================================================
# CONFIGURACION DE PAGINA
# ============================================================================

st.set_page_config(
    page_title="Analisis Tematico Microcurricular",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# ESTILOS VISUALES — POLITECNICO GRANCOLOMBIANO
# Paleta: #0F385A (navy) | #1FB2DE (azul) | #42F2F2 (cyan)
#         #FBAF17 (dorado) | #EC0677 (magenta) — tema claro
# ============================================================================

st.markdown("""
<style>
/* ══════════════════════════════════════════════════════════════════════════
   FORZAR MODO CLARO SIEMPRE — pisa variables CSS de Streamlit en dark mode
   ══════════════════════════════════════════════════════════════════════════ */
:root,
:root[data-theme="dark"],
:root[data-theme="light"] {
    color-scheme: light !important;
    --background-color:           #FFFFFF !important;
    --secondary-background-color: #EAF9FD !important;
    --text-color:                 #0F385A !important;
    --font:                       'Segoe UI', 'Inter', sans-serif !important;
    --primary-color:              #1FB2DE !important;
}
/* Cubre la preferencia del sistema operativo en dark mode */
@media (prefers-color-scheme: dark) {
    :root {
        color-scheme: light !important;
        --background-color:           #FFFFFF !important;
        --secondary-background-color: #EAF9FD !important;
        --text-color:                 #0F385A !important;
    }
}

/* ── Fondo y texto de la app principal ──────────────────────────────────── */
html, body,
.stApp, .stApp > div,
.main, .block-container,
[data-testid="stAppViewContainer"],
[data-testid="stHeader"] {
    font-family: 'Segoe UI', 'Inter', sans-serif !important;
    background-color: #FFFFFF !important;
    color: #0F385A !important;
}

/* ── Contenido dentro de tabs, expanders, columnas ─────────────────────── */
.stTabs [data-baseweb="tab-panel"],
[data-testid="stExpander"],
[data-testid="column"] {
    background-color: #FFFFFF !important;
    color: #0F385A !important;
}

/* ── Inputs, selectbox, multiselect ─────────────────────────────────────── */
/* Caja de control (fondo blanco, texto navy) */
[data-baseweb="select"] > div:first-child,
[data-baseweb="base-input"],
.stTextInput input, .stTextArea textarea, .stNumberInput input {
    background-color: #FFFFFF !important;
    color: #0F385A !important;
    border-color: #1FB2DE !important;
}
/* Texto del placeholder e input interno */
[data-baseweb="select"] input,
[data-baseweb="select"] [data-id="datepicker-input"] {
    color: #0F385A !important;
    background-color: transparent !important;
}
/* Chips / tags de items seleccionados en multiselect */
[data-baseweb="tag"] {
    background-color: #EAF9FD !important;
    border: 1px solid #1FB2DE !important;
}
[data-baseweb="tag"] span {
    color: #0F385A !important;
    background-color: transparent !important;
}
[data-baseweb="tag"] button,
[data-baseweb="tag"] [role="button"] {
    color: #0F385A !important;
    background-color: transparent !important;
}
/* Lista desplegable de opciones */
[data-baseweb="menu"],
[data-baseweb="popover"] > div {
    background-color: #FFFFFF !important;
}
[data-baseweb="menu"] li,
[data-baseweb="menu"] [role="option"] {
    color: #0F385A !important;
    background-color: #FFFFFF !important;
}
[data-baseweb="menu"] li:hover,
[data-baseweb="menu"] [role="option"]:hover {
    background-color: #EAF9FD !important;
}
/* Opción seleccionada en el dropdown */
[data-baseweb="menu"] [aria-selected="true"] {
    background-color: #D0F4FA !important;
    color: #0F385A !important;
}

/* ── Métricas y texto general ───────────────────────────────────────────── */
[data-testid="stMetricValue"],
[data-testid="stMetricLabel"],
[data-testid="stMetricDelta"],
.stMarkdown, .stText, p, h1, h2, h3, h4, h5, label {
    color: #0F385A !important;
}

/* ── Barra lateral ──────────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F385A 0%, #071E32 100%);
    border-right: 3px solid #1FB2DE;
}
[data-testid="stSidebar"][data-testid="stSidebar"] {
    overflow-y: auto !important;
}
[data-testid="stSidebar"] * {
    color: #FFFFFF !important;
}
/* Selectbox en sidebar: texto oscuro sobre fondo blanco */
[data-testid="stSidebar"] [data-baseweb="select"] > div:first-child,
[data-testid="stSidebar"] [data-baseweb="select"] input,
[data-testid="stSidebar"] [data-baseweb="select"] [data-id="select-input"] {
    color: #0F385A !important;
    background-color: #FFFFFF !important;
}
/* ── option_menu: fondo y texto ─────────────────────────────────────────── */
/* Contenedor principal del menú */
[data-testid="stSidebar"] .nav,
[data-testid="stSidebar"] .nav-pills,
[data-testid="stSidebar"] ul.nav {
    background-color: transparent !important;
}
/* Cada ítem */
[data-testid="stSidebar"] .nav-link,
[data-testid="stSidebar"] a.nav-link {
    color: rgba(255,255,255,0.85) !important;
    background-color: transparent !important;
    border-left: 3px solid transparent !important;
    border-radius: 6px !important;
    padding: 7px 12px !important;
    margin: 2px 0 !important;
    font-size: 13px !important;
    transition: background 0.2s, border-left 0.2s;
}
/* Hover */
[data-testid="stSidebar"] .nav-link:hover,
[data-testid="stSidebar"] a.nav-link:hover {
    background-color: rgba(31,178,222,0.20) !important;
    color: #FFFFFF !important;
}
/* Activo */
[data-testid="stSidebar"] .nav-link.active,
[data-testid="stSidebar"] a.nav-link.active {
    background-color: rgba(31,178,222,0.30) !important;
    border-left: 3px solid #FBAF17 !important;
    color: #FFFFFF !important;
    font-weight: 600 !important;
}
/* Iconos */
[data-testid="stSidebar"] .nav-link i,
[data-testid="stSidebar"] a.nav-link i {
    color: #42F2F2 !important;
}
/* ── Forzar visibilidad del iframe de option_menu ── */
[data-testid="stSidebar"] iframe {
    display: block !important;
    min-height: 480px !important;
    height: auto !important;
    width: 100% !important;
    border: none !important;
    overflow-y: auto !important;
}
[data-testid="stSidebar"] hr {
    border-color: transparent !important;
    background-color: transparent !important;
}
[data-testid="stSidebar"] .stFileUploader {
    background: rgba(255,255,255,0.07);
    border-radius: 8px;
    padding: 8px;
    border: 1px solid rgba(31,178,222,0.3);
}
/* Zona de drop: fondo oscuro para que el texto blanco sea visible */
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"],
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] > div {
    background-color: rgba(7, 30, 50, 0.85) !important;
    border: 1px dashed rgba(66,242,242,0.5) !important;
    border-radius: 6px !important;
}
/* Todos los textos dentro del file uploader */
[data-testid="stSidebar"] [data-testid="stFileUploader"] label,
[data-testid="stSidebar"] [data-testid="stFileUploader"] span,
[data-testid="stSidebar"] [data-testid="stFileUploader"] small,
[data-testid="stSidebar"] [data-testid="stFileUploader"] p,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] *,
[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] * {
    color: #FFFFFF !important;
    background-color: transparent !important;
}
/* Botón "Browse files" */
[data-testid="stSidebar"] [data-testid="stFileUploader"] button,
[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] {
    background: rgba(31,178,222,0.25) !important;
    border: 1px solid rgba(66,242,242,0.5) !important;
    color: #FFFFFF !important;
    border-radius: 6px !important;
}

/* ── Encabezados principales ─────────────────────────────────────────────── */
h1 {
    color: #0F385A !important;
    border-bottom: 3px solid #1FB2DE;
    padding-bottom: 8px;
}
h2 {
    color: #0F385A !important;
    border-left: 4px solid #1FB2DE;
    padding-left: 10px;
}
h3 {
    color: #0F385A !important;
}

/* ── Tarjetas de métricas ─────────────────────────────────────────────────── */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, #EAF9FD 0%, #F5FEFF 100%);
    border: 1px solid #A8E6F5;
    border-left: 4px solid #1FB2DE;
    border-radius: 10px;
    padding: 14px 16px !important;
    box-shadow: 0 2px 8px rgba(31,178,222,0.12);
}
[data-testid="stMetricValue"] {
    color: #0F385A !important;
    font-weight: 700 !important;
}
[data-testid="stMetricLabel"] {
    color: #1FB2DE !important;
    font-size: 0.85em !important;
    font-weight: 600 !important;
}

/* ── Cuadros info / warning / success / error ────────────────────────────── */
[data-testid="stAlert"][data-baseweb="notification"] {
    border-radius: 10px !important;
    border-left-width: 5px !important;
}
div[data-testid="stAlert"] > div[role="alert"] {
    border-radius: 10px;
}

/* Info → azul Politécnico */
div[data-baseweb="notification"][kind="info"],
div[class*="stInfo"] {
    background-color: #E8F8FD !important;
    border-left: 5px solid #1FB2DE !important;
    color: #0F385A !important;
}

/* Success → verde */
div[data-baseweb="notification"][kind="positive"] {
    background-color: #EAFAF1 !important;
    border-left: 5px solid #27AE60 !important;
}

/* Warning → dorado Politécnico */
div[data-baseweb="notification"][kind="warning"] {
    background-color: #FFFBEA !important;
    border-left: 5px solid #FBAF17 !important;
    color: #5C3D00 !important;
}

/* Error → magenta Politécnico */
div[data-baseweb="notification"][kind="negative"] {
    background-color: #FEE5F2 !important;
    border-left: 5px solid #EC0677 !important;
    color: #5A0030 !important;
}

/* ── Tabs ────────────────────────────────────────────────────────────────── */
[data-testid="stTabs"] [role="tablist"] {
    border-bottom: 2px solid #1FB2DE;
    gap: 4px;
}
[data-testid="stTabs"] [role="tab"] {
    background: #EAF9FD;
    border-radius: 8px 8px 0 0;
    color: #0F385A;
    font-weight: 500;
    padding: 8px 18px;
    border: 1px solid #A8E6F5;
    border-bottom: none;
    transition: background 0.2s;
}
[data-testid="stTabs"] [role="tab"]:hover {
    background: #C5EEF8;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    background: #1FB2DE !important;
    color: #FFFFFF !important;
    border-color: #1FB2DE;
    font-weight: 700 !important;
}

/* ── Botones ─────────────────────────────────────────────────────────────── */
.stButton > button, .stButton button, [data-testid="stButton"] button, .stDownloadButton > button, button[data-testid="stBaseButton-secondary"] {
    background: linear-gradient(135deg, #1976D2, #1565C0) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 8px 22px !important;
    font-weight: 600 !important;
    box-shadow: 0 3px 8px rgba(15,56,90,0.22) !important;
    transition: transform 0.15s, box-shadow 0.15s !important;
}
.stButton > button p, .stButton button p, [data-testid="stButton"] button p, button[data-testid="stBaseButton-secondary"] p {
    color: #ffffff !important;
}
.stButton > button span, .stButton button span {
    color: #ffffff !important;
}
.stButton > button:hover, .stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 5px 14px rgba(15,56,90,0.32) !important;
    background: linear-gradient(135deg, #FBAF17, #D48F00) !important;
}

/* ── Expanders ───────────────────────────────────────────────────────────── */
details {
    border: 1px solid #A8E6F5 !important;
    border-radius: 8px !important;
    margin-bottom: 6px !important;
}
details summary {
    background: linear-gradient(90deg, #EAF9FD, #F5FEFF) !important;
    color: #0F385A !important;
    font-weight: 500;
    border-radius: 8px;
    padding: 8px 14px;
}
details[open] summary {
    border-bottom: 2px solid #1FB2DE;
    border-radius: 8px 8px 0 0;
}

/* ── Selectbox / Multiselect ─────────────────────────────────────────────── */
[data-testid="stMultiSelect"] [data-baseweb="tag"] {
    background: #1FB2DE !important;
    color: white !important;
    border-radius: 5px;
}

/* ── Tabla dataframe ─────────────────────────────────────────────────────── */
[data-testid="stDataFrame"] {
    border: 1px solid #A8E6F5 !important;
    border-radius: 8px !important;
    overflow: hidden;
}

/* ── Divisor horizontal ─────────────────────────────────────────────────── */
hr {
    border: none;
    border-top: 2px solid #D0F0FA;
    margin: 20px 0;
}

/* ── Caption / pequeño texto ─────────────────────────────────────────────── */
[data-testid="stCaptionContainer"] p,
small, .caption {
    color: #4A7A90 !important;
}

/* ── Spinner ─────────────────────────────────────────────────────────────── */
[data-testid="stSpinner"] > div {
    border-top-color: #1FB2DE !important;
}

/* ── Header institucional (banner) ──────────────────────────────────────── */
.poligran-header {
    background: linear-gradient(135deg, #0F385A 0%, #1FB2DE 70%, #42F2F2 100%);
    padding: 20px 28px;
    border-radius: 12px;
    margin-bottom: 18px;
    box-shadow: 0 4px 20px rgba(15,56,90,0.25);
}
.poligran-header h2 {
    color: white !important;
    border: none !important;
    padding: 0 !important;
    margin: 0;
    font-size: 1.5em;
    text-shadow: 0 1px 3px rgba(0,0,0,0.2);
}
.poligran-header p {
    color: rgba(255,255,255,0.9) !important;
    margin: 5px 0 0 0;
    font-size: 0.95em;
}
.poligran-badge {
    background: #FBAF17;
    color: #0F385A;
    font-size: 0.72em;
    font-weight: 700;
    padding: 3px 12px;
    border-radius: 20px;
    letter-spacing: 0.5px;
    display: inline-block;
    margin-top: 8px;
}

/* ── Sección con fondo suave ─────────────────────────────────────────────── */
.section-card {
    background: #F5FDFF;
    border: 1px solid #C5EEF8;
    border-radius: 10px;
    padding: 16px 20px;
    margin: 10px 0;
}
.section-card-accent {
    background: #F5FDFF;
    border: 1px solid #C5EEF8;
    border-left: 4px solid #FBAF17;
    border-radius: 10px;
    padding: 16px 20px;
    margin: 10px 0;
}

/* ── Badge de alerta ────────────────────────────────────────────────────── */
.badge-alta   { background:#EC0677; color:white; padding:2px 8px; border-radius:12px; font-size:0.78em; font-weight:700; }
.badge-media  { background:#FBAF17; color:#0F385A; padding:2px 8px; border-radius:12px; font-size:0.78em; font-weight:700; }
.badge-baja   { background:#42F2F2; color:#0F385A; padding:2px 8px; border-radius:12px; font-size:0.78em; font-weight:700; }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# TENDENCIAS GLOBALES POR DEFECTO
# ============================================================================

TENDENCIAS_DEFAULT = {
    "SOSTENIBILIDAD": {
        "keywords": [
            "sostenibilidad", "sostenible", "desarrollo sostenible",
            "ODS", "objetivos desarrollo", "medio ambiente", "ambiental",
            "ecologia", "verde", "cambio climatico", "huella carbono",
            "economia circular", "recursos naturales",
            "consumo responsable", "energias renovables", "eficiencia energetica",
            "movilidad sostenible", "agua", "biodiversidad",
            "innovacion social", "desarrollo comunitario", "responsabilidad social"
        ],
        "color": "#2ECC71",
        "descripcion": "Sostenibilidad y Desarrollo Sostenible (ODS)"
    },
    "INTELIGENCIA_ARTIFICIAL": {
        "keywords": [
            "inteligencia artificial", "IA", "machine learning", "aprendizaje automatico",
            "deep learning", "redes neuronales", "algoritmos", "big data",
            "data science", "ciencia datos", "analytics", "mineria datos",
            "chatbot", "NLP", "vision artificial"
        ],
        "color": "#3498DB",
        "descripcion": "Inteligencia Artificial y Tecnologias Emergentes"
    },
    "TRANSFORMACION_DIGITAL": {
        "keywords": [
            "transformacion digital", "digitalizacion", "digital",
            "industria 4.0", "automatizacion", "robotica",
            "internet cosas", "IoT", "cloud computing", "nube",
            "ciberseguridad", "blockchain", "tecnologia"
        ],
        "color": "#9B59B6",
        "descripcion": "Transformacion Digital e Industria 4.0"
    },
    "INNOVACION": {
        "keywords": [
            "innovacion", "innovar", "emprendimiento", "emprender",
            "startup", "creatividad", "disrupcion", "disruptivo",
            "Design thinking", "lean", "agil", "prototipo"
        ],
        "color": "#E74C3C",
        "descripcion": "Innovacion y Emprendimiento"
    },
    "ETICA": {
        "keywords": [
            "etica", "valores", "responsabilidad social", "RSE", "RSC",
            "deontologia", "codigo etico", "integridad",
            "transparencia", "buen gobierno", "compliance"
        ],
        "color": "#42F2F2",
        "descripcion": "Etica, Valores y Responsabilidad Social"
    },
    "GLOBALIZACION": {
        "keywords": [
            "globalizacion", "global", "internacional", "glocal",
            "interculturalidad", "multicultural", "diversidad cultural",
            "comercio internacional", "exportacion", "importacion"
        ],
        "color": "#1ABC9C",
        "descripcion": "Globalizacion y Perspectiva Glocal"
    },
    "LIDERAZGO": {
        "keywords": [
            "liderazgo", "trabajo equipo", "colaborativo", "comunicacion",
            "habilidades blandas", "soft skills", "gestion personas",
            "talento humano", "coaching", "mentoring", "negociacion"
        ],
        "color": "#1A5276",
        "descripcion": "Liderazgo y Habilidades Blandas"
    },
    "ANALISIS_DATOS": {
        "keywords": [
            "analisis datos", "estadistica", "data analytics",
            "visualizacion datos", "business intelligence", "BI",
            "dashboard", "metricas", "KPI", "indicadores",
            "toma decisiones basada datos"
        ],
        "color": "#16A085",
        "descripcion": "Analisis de Datos y Business Intelligence"
    },
    "GESTION_CAMBIO": {
        "keywords": [
            "gestion cambio", "adaptabilidad", "resiliencia",
            "agilidad organizacional", "transformacion organizacional",
            "cultura organizacional", "cambio organizacional"
        ],
        "color": "#8E44AD",
        "descripcion": "Gestion del Cambio y Adaptabilidad"
    },
    "CALIDAD": {
        "keywords": [
            "calidad", "mejora continua", "excelencia",
            "ISO", "normas", "estandares", "certificacion",
            "auditoria", "procesos", "eficiencia", "productividad"
        ],
        "color": "#27AE60",
        "descripcion": "Calidad y Mejora Continua"
    }
}

STOPWORDS_ES = set([
    'el', 'la', 'de', 'que', 'y', 'a', 'en', 'un', 'ser', 'se', 'no', 'haber',
    'por', 'con', 'su', 'para', 'como', 'estar', 'tener', 'le', 'lo', 'todo',
    'pero', 'mas', 'hacer', 'o', 'poder', 'decir', 'este', 'ir', 'otro', 'ese',
    'la', 'si', 'me', 'ya', 'ver', 'porque', 'dar', 'cuando', 'muy', 'sin',
    'vez', 'mucho', 'saber', 'que', 'sobre', 'tambien', 'me', 'hasta', 'hay',
    'donde', 'quien', 'desde', 'todos', 'durante', 'uno', 'les', 'ni',
    'contra', 'otros', 'fueron', 'ese', 'eso', 'ante', 'ellos', 'e', 'esto',
    'mi', 'antes', 'algunos', 'unos', 'yo', 'del', 'las', 'los', 'al',
    'una', 'nos', 'te', 'ti'
])


# ============================================================================
# FUNCIONES DE CARGA DE DATOS
# ============================================================================

def normalizar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nombres de columnas eliminando tildes."""
    nuevos_nombres = {}
    for col in df.columns:
        nfkd = unicodedata.normalize('NFKD', str(col))
        sin_acentos = ''.join(c for c in nfkd if not unicodedata.combining(c))
        nuevos_nombres[col] = sin_acentos
    return df.rename(columns=nuevos_nombres)


def _normalize_column_name(name: str) -> str:
    """Normaliza un nombre de columna para comparación tolerante."""
    normalized = unicodedata.normalize('NFKD', str(name))
    normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
    normalized = normalized.lower()
    normalized = re.sub(r'[\s\-_.]+', '', normalized)
    return normalized


def _find_column(df: pd.DataFrame, target: str):
    """Busca una columna en el DataFrame ignorando mayúsculas, espacios, tildes y puntuación."""
    target_norm = _normalize_column_name(target)
    for col in df.columns:
        col_norm = _normalize_column_name(col)
        if col_norm == target_norm:
            return col
    return None


def _normalize_value(value: str) -> str:
    """Normaliza valores de celda para comparación tolerante."""
    normalized = unicodedata.normalize('NFKD', str(value))
    normalized = ''.join(c for c in normalized if not unicodedata.combining(c))
    normalized = normalized.lower()
    normalized = ''.join(c for c in normalized if c.isalnum())
    return normalized


_TAXONOMIAS_MATRIZ_PATH = "data/raw/Taxonomias_MatrizBD.xlsx"

# ── Dominio de cada subcategoría (clave normalizada sin acentos) ──────────────
_SUBCAT_TO_DOMAIN: Dict[str, str] = {
    'conocimiento':  'Cognitivo',
    'comprension':   'Cognitivo',
    'aplicacion':    'Cognitivo',
    'analisis':      'Cognitivo',
    'sintesis':      'Cognitivo',
    'evaluacion':    'Cognitivo',
    'creacion':      'Cognitivo',
    'imitacion':     'Procedimental',
    'manipulacion':  'Procedimental',
    'precision':     'Procedimental',
    'control':       'Procedimental',
    'percepcion':    'Actitudinal',
    'responder':     'Actitudinal',
    'valorar':       'Actitudinal',
    'organizar':     'Actitudinal',
    'caracterizar':  'Actitudinal',
}

# Orden jerárquico (1=más básico … 15=más complejo) para priorizar al detectar
_SUBCAT_ORDER: Dict[str, int] = {
    'conocimiento': 1,  'comprension': 2,  'aplicacion': 3,
    'analisis':     4,  'sintesis':    5,  'evaluacion': 6,
    'imitacion':    7,  'manipulacion': 8, 'precision':  9, 'control': 10,
    'percepcion':  11,  'responder':   12, 'valorar':   13,
    'organizar':   14,  'caracterizar': 15,
}

# Colores por subcategoría (tonos del dominio)
_SUBCAT_COLORS: Dict[str, str] = {
    'Conocimiento': '#AED6F1', 'Comprension': '#5DADE2',
    'Aplicacion':   '#2E86C1', 'Analisis':    '#1A5276',
    'Sintesis':     '#7D3C98', 'Evaluacion':  '#922B21',
    'Imitacion':    '#ABEBC6', 'Manipulacion': '#58D68D',
    'Precision':    '#1E8449', 'Control':     '#0B5345',
    'Percepcion':   '#B3DCF0', 'Responder':   '#5DADE2',
    'Valorar':      '#2E86C1', 'Organizar':   '#1A5276',
    'Caracterizar': '#0F385A',
}

_DOMAIN_COLORS: Dict[str, str] = {
    'Cognitivo':      '#2E86C1',
    'Procedimental':  '#1E8449',
    'Actitudinal':    '#42F2F2',
    'No identificado': '#BDC3C7',
}

# Mapeo explícito: nombre_subcategoria (normalizado sin acentos) → nivel Bloom
_SUBCAT_TO_BLOOM: Dict[str, str] = {
    'conocimiento':  'Recordar',
    'comprension':   'Comprender',
    'aplicacion':    'Aplicar',
    'analisis':      'Analizar',
    'sintesis':      'Crear',
    'evaluacion':    'Evaluar',
    'creacion':      'Crear',
    # Procedimental → Aplicar (uso práctico de habilidades)
    'imitacion':     'Aplicar',
    'manipulacion':  'Aplicar',
    'precision':     'Aplicar',
    'control':       'Evaluar',
    # Actitudinal → distribuir entre niveles inferiores
    'percepcion':    'Recordar',
    'responder':     'Comprender',
    'valorar':       'Evaluar',
    'organizar':     'Analizar',
    'caracterizar':  'Crear',
}


def _stem_es(word: str) -> str:
    """Extrae el stem aproximado de un verbo español quitando terminaciones comunes."""
    w = word.strip().lower()
    # Longer suffixes first to avoid premature truncation
    for suf in ('ando', 'iendo', 'ados', 'idos', 'ado', 'ido',
                'amos', 'emos', 'imos', 'aron', 'ieron',
                'ará', 'erá', 'irá', 'ara', 'era', 'ira',
                'ar', 'er', 'ir',
                'as', 'es', 'an', 'en',
                'a', 'e', 'o'):
        if w.endswith(suf) and len(w) - len(suf) >= 4:
            return w[:-len(suf)]
    return w


def leer_taxonomias_bloom(uploaded_files=None) -> Dict[str, list]:
    """
    Carga verbos desde data/raw/Taxonomias_MatrizBD.xlsx, hoja 'Verbos'.
    Estructura: id_verbo | verbo | id_subcategoria | nombre_subcategoria |
                id_dominio | nombre_dominio
    Retorna: {nombre_subcategoria_normalizado: [verbos...]}
    """
    import os

    def _norm(s: str) -> str:
        return unicodedata.normalize('NFKD', str(s).strip().lower()).encode('ascii', 'ignore').decode('ascii')

    taxonomias: Dict[str, list] = {}

    if os.path.exists(_TAXONOMIAS_MATRIZ_PATH):
        try:
            df_v = pd.read_excel(_TAXONOMIAS_MATRIZ_PATH, sheet_name='Verbos', engine='openpyxl')
            # Normalize column names for safe access
            df_v.columns = [_norm(c) for c in df_v.columns]

            col_verbo  = next((c for c in df_v.columns if c == 'verbo'), None)
            col_subcat = next((c for c in df_v.columns if 'nombre_subcategoria' in c), None)

            if col_verbo and col_subcat:
                for _, row in df_v.iterrows():
                    subcat = _norm(str(row[col_subcat]))
                    verbo  = _norm(str(row[col_verbo]))
                    if subcat in ('', 'nan') or verbo in ('', 'nan') or len(verbo) < 3:
                        continue
                    taxonomias.setdefault(subcat, []).append(verbo)
        except Exception:
            taxonomias = {}

    return {nivel: list(dict.fromkeys(verbos)) for nivel, verbos in taxonomias.items()}


def extract_modality_sede(filename):
    """Extrae modalidad y sede del nombre del archivo."""
    base_name = filename.replace('.xlsx', '').replace('.xls', '')
    codigo = base_name[-4:] if len(base_name) >= 4 else ""
    
    sede_map = {
        'PBOG': 'Bogotá', 'PMED': 'Medellín', 'VNAL': 'Nacional',
        'HBOG': 'Bogotá', 'HMED': 'Medellín', 'HVAL': 'Virtual',
    }
    modalidad_map = {'P': 'Presencial', 'V': 'Virtual', 'H': 'Híbrido'}
    
    modalidad_codigo = codigo[0] if codigo else ""
    modalidad = modalidad_map.get(modalidad_codigo, 'No identificado')
    sede = sede_map.get(codigo, codigo if codigo else 'No identificado')
    
    return {'modalidad': modalidad, 'sede': sede, 'codigo': codigo}


def procesar_archivos(uploaded_files) -> pd.DataFrame:
    """Procesa archivos Excel subidos y consolida datos."""
    all_data = []
    failed_files = []  # Registrar archivos fallidos

    for uploaded_file in uploaded_files:
        nombre = uploaded_file.name
        metadata = extract_modality_sede(nombre)
        
        # Extraer nombre del programa (primero del archivo, luego intentar leer celda A3 del perfil)
        programa_nombre = (
            nombre
            .replace("FormatoRA_", "")
            .replace("FormatoRA-", "")
            .replace("_PBOG", "")
            .replace("_VNAL", "")
            .replace("_PMED", "")
            .replace("_HBOG", "")
            .replace("_HMED", "")
            .replace("_HVAL", "")
            .replace(".xlsx", "")
            .replace(".xls", "")
        )
        programa_nombre = programa_nombre.strip()

        # Si el archivo contiene la hoja de perfil, usar el valor real del programa si está disponible.
        try:
            uploaded_file.seek(0)
            df_perfil = pd.read_excel(
                uploaded_file,
                sheet_name='Paso1 Analisis perfil egreso',
                header=None,
                nrows=10,
                engine='openpyxl'
            )
            if df_perfil is not None and len(df_perfil) > 2 and len(df_perfil.columns) > 0:
                val = df_perfil.iloc[2, 0]
                if val is not None and str(val).strip():
                    programa_nombre = str(val).strip()
        except Exception:
            pass  # Mantener nombre derivado del archivo si no se puede leer la hoja        
        try:
            uploaded_file.seek(0)
            df = pd.read_excel(
                uploaded_file,
                sheet_name='Paso 5 Estrategias micro',
                header=1,
                engine='openpyxl'
            )
            if df is not None and not df.empty:
                df = normalizar_columnas(df)
                nivel_col = _find_column(df, 'Nivel')
                if nivel_col is not None:
                    df['Nivel'] = (
                        df[nivel_col]
                        .astype(str)
                        .str.strip()
                        .replace({'nan': 'No identificado', '': 'No identificado'})
                        .str.title()
                    )
                componente_col = _find_column(df, 'Componente academico')
                if componente_col is not None:
                    df['Componente academico'] = (
                        df[componente_col]
                        .astype(str)
                        .str.strip()
                        .replace({'nan': 'No identificado', '': 'No identificado'})
                    )

                required_fields = ['Tipo de Saber', 'Resultado de aprendizaje', 'Nombre asignatura o modulo', 'Indicadores de logro asignatura o modulo', 'Creditos', 'Semestre']
                found_cols = {col: _find_column(df, col) for col in required_fields}
                missing_fields = [col for col, found in found_cols.items() if found is None]
                if missing_fields:
                    st.warning(f"{nombre}: Faltan columnas requeridas - {', '.join(missing_fields)}")

                df['Programa'] = programa_nombre
                df['Modalidad'] = metadata['modalidad']
                df['Sede'] = metadata['sede']
                df['Codigo'] = metadata['codigo']
                all_data.append(df)
            else:
                failed_files.append({'nombre': nombre, 'causa': 'Archivo vacío'})
        except Exception as e:
            error_msg = str(e)
            if 'multiple' in error_msg.lower() or 'found' in error_msg.lower():
                causa = "Hoja no encontrada o formato inválido"
            elif 'empty' in error_msg.lower():
                causa = "Archivo vacío o sin datos"
            elif 'permission' in error_msg.lower():
                causa = "Archivo en uso por otra aplicación"
            else:
                causa = error_msg[:80] if len(error_msg) > 80 else error_msg
            failed_files.append({'nombre': nombre, 'causa': causa})
            continue

    if not all_data:
        return pd.DataFrame(), failed_files

    df_consolidado = pd.concat(all_data, ignore_index=True)

    # Derivar Nivel cuando no exista como columna explícita
    if 'Nivel' not in df_consolidado.columns:
        niveles_por_programa = {
            prog: _detectar_nivel(g)
            for prog, g in df_consolidado.groupby('Programa', sort=False)
        }
        df_consolidado['Nivel'] = df_consolidado['Programa'].map(niveles_por_programa)
    else:
        df_consolidado['Nivel'] = (
            df_consolidado['Nivel'].astype(str)
            .str.strip()
            .replace({'nan': 'No identificado', '': 'No identificado'})
            .str.title()
        )
        # Rellenar niveles no identificados con detección automática
        mask_no_id = df_consolidado['Nivel'].isin(['No identificado', 'nan', ''])
        if mask_no_id.any():
            niveles_por_programa = {
                prog: _detectar_nivel(g)
                for prog, g in df_consolidado[mask_no_id].groupby('Programa', sort=False)
            }
            df_consolidado.loc[mask_no_id, 'Nivel'] = df_consolidado.loc[mask_no_id, 'Programa'].map(niveles_por_programa)

    # Filtrar Tipo de Saber vacio
    df_consolidado = df_consolidado[df_consolidado['Tipo de Saber'].notna()]
    df_consolidado = df_consolidado[
        df_consolidado['Tipo de Saber'].astype(str).str.strip() != ''
    ]

    # Normalizar Tipo de Saber (tolerante a espacios, guiones, acentos, mayúsculas)
    def _norm_tipo_saber(valor: str) -> str:
        v = unicodedata.normalize('NFKD', str(valor).strip().lower()).encode('ascii', 'ignore').decode('ascii')
        v = re.sub(r'[\s\-_/]+', '', v)   # quitar espacios, guiones, barras
        if v in ('saber', 'conocimiento', 'conceptual', 'teorico', 'conocer'):
            return 'Saber'
        if 'saberhacer' in v or v in ('hacer', 'practica', 'procedimental', 'habilidad', 'proceder'):
            return 'SaberHacer'
        if 'saberser' in v or v in ('ser', 'actitud', 'valor', 'etica', 'actitudinal', 'humano'):
            return 'SaberSer'
        return valor  # mantener original si no se reconoce

    df_consolidado['Tipo de Saber'] = (
        df_consolidado['Tipo de Saber'].astype(str).str.strip()
        .map(_norm_tipo_saber)
    )
    # Descartar filas cuyo Tipo de Saber no sea uno de los tres válidos
    df_consolidado = df_consolidado[
        df_consolidado['Tipo de Saber'].isin({'Saber', 'SaberHacer', 'SaberSer'})
    ]

    # Preparar texto combinado
    if 'Proceso Responsable' not in df_consolidado.columns:
        df_consolidado['Proceso Responsable'] = ''
    df_consolidado['Texto_Completo'] = (
        df_consolidado['Resultado de aprendizaje'].fillna('') + ' ' +
        df_consolidado['Nombre asignatura o modulo'].fillna('') + ' ' +
        df_consolidado['Indicadores de logro asignatura o modulo'].fillna('') + ' ' +
        df_consolidado['Nucleos tematicos'].fillna('') + ' ' +
        df_consolidado['Proceso Responsable'].fillna('')
    ).str.lower().str.strip()

    return df_consolidado, failed_files


def obtener_tendencias() -> Dict:
    """Obtiene tendencias desde session_state o usa las por defecto."""
    if 'tendencias' not in st.session_state:
        st.session_state['tendencias'] = json.loads(json.dumps(TENDENCIAS_DEFAULT))
    return st.session_state['tendencias']


# ============================================================================
# FUNCIONES DE ANALISIS
# ============================================================================

def _split_nucleos(texto: str) -> list:
    """Separa núcleos temáticos. NO divide por coma porque las comas son
    parte de la descripción del núcleo (ej: 'Ciudadanía: relaciones, dinámicas...')."""
    partes = re.split(r'[;\n\|]+', texto)
    # También dividir en ítems numerados tipo "1. Núcleo" o "2) Núcleo"
    resultado = []
    for p in partes:
        sub = re.split(r'(?<!\d)(?=\d+[\.\)]\s+\w)', p)
        resultado.extend(sub)
    return resultado


def analizar_cobertura(df: pd.DataFrame) -> Dict:
    """Analisis de cobertura y densidad tematica usando pipeline de nucleos_cleaner."""
    # Pipeline de filtrado
    df_filtrado = filtrar_nucleos_dataframe(df, columna='Nucleos tematicos')

    # Extraer todos los núcleos válidos del DataFrame filtrado
    nucleos_list = []
    todos_rechazados = []
    for _, row in df_filtrado.iterrows():
        nucleos_list.extend(row.get('nucleos_validos', []))
        todos_rechazados.extend(row.get('nucleos_rechazados', []))

    nucleos_counter = Counter(nucleos_list)

    # Guardar rechazados en session_state para auditoría
    if 'nucleos_rechazados_pipeline' not in st.session_state:
        st.session_state['nucleos_rechazados_pipeline'] = []
    if todos_rechazados:
        nuevos = [r for r in todos_rechazados
                  if r not in st.session_state['nucleos_rechazados_pipeline']]
        st.session_state['nucleos_rechazados_pipeline'].extend(nuevos)

    # Densidad por asignatura
    densidad = df.groupby('Nombre asignatura o modulo')['Nucleos tematicos'].apply(
        lambda x: len(_split_nucleos(' '.join(x.fillna('').astype(str))))
    ).sort_values(ascending=False)

    # Shannon entropy
    if len(nucleos_counter) > 1:
        frecuencias = np.array(list(nucleos_counter.values()))
        probabilidades = frecuencias / frecuencias.sum()
        shannon = entropy(probabilidades, base=2)
        max_ent = np.log2(len(nucleos_counter))
        diversidad = (shannon / max_ent) * 100
    else:
        shannon = 0
        diversidad = 0

    # Matriz programa x nucleo (top 20)
    top_20 = [n for n, _ in nucleos_counter.most_common(20)]
    matriz = pd.DataFrame(0, index=df['Programa'].unique(), columns=top_20)
    for _, row in df_filtrado.iterrows():
        programa = row['Programa']
        for nucleo in row.get('nucleos_validos', []):
            if nucleo in top_20:
                matriz.loc[programa, nucleo] += 1

    return {
        'nucleos_counter': nucleos_counter,
        'nucleos_unicos': len(nucleos_counter),
        'total_menciones': sum(nucleos_counter.values()),
        'densidad_asignatura': densidad,
        'promedio_densidad': densidad.mean(),
        'shannon': shannon,
        'diversidad': diversidad,
        'matriz_cobertura': matriz,
        'total_rechazados': len(todos_rechazados)
    }


def analizar_tendencias(df: pd.DataFrame, tendencias: Dict) -> Dict:
    """Detecta tendencias globales en los datos.

    Cobertura = % de asignaturas unicas que abordan la tendencia
    (no % de programas). Esto refleja cuantas materias del total
    trabajan cada tendencia.
    """
    programas = df['Programa'].unique()
    asig_col = 'Nombre asignatura o modulo'
    total_asigs = df[asig_col].nunique()

    matriz = pd.DataFrame(0, index=programas, columns=list(tendencias.keys()))
    detalle = {t: {p: [] for p in programas} for t in tendencias}
    # Sets de asignaturas unicas que tocan cada tendencia
    asig_sets = {tid: set() for tid in tendencias}

    for _, row in df.iterrows():
        programa = row['Programa']
        texto = str(row.get('Texto_Completo', '')).lower()
        asig_raw = row.get(asig_col, '')
        asig_str = (
            str(asig_raw).strip()
            if pd.notna(asig_raw) and str(asig_raw).strip() not in ('nan', '')
            else ''
        )

        for tid, tinfo in tendencias.items():
            # Recolectar TODAS las keywords que coinciden en este texto
            kws_match = [kw for kw in tinfo['keywords'] if kw.lower() in texto]
            if not kws_match:
                continue
            # Contar la fila UNA sola vez para no inflar la matriz
            matriz.loc[programa, tid] += 1
            if asig_str:
                asig_sets[tid].add(asig_str)
            # Guardar un hallazgo por cada keyword coincidente
            _ra   = str(row.get('Resultado de aprendizaje', ''))
            _nuc  = str(row.get('Nucleos tematicos', ''))
            _ind  = str(row.get('Indicadores de logro asignatura o modulo', ''))
            _proc = str(row.get('Proceso Responsable', ''))
            for kw in kws_match:
                campos = []
                textos = {}
                if kw.lower() in _ra.lower():
                    campos.append('RA')
                    textos['RA'] = _ra
                if kw.lower() in _nuc.lower():
                    campos.append('Nucleos')
                    textos['Nucleos'] = _nuc
                if kw.lower() in _ind.lower():
                    campos.append('Indicadores')
                    textos['Indicadores'] = _ind
                if kw.lower() in _proc.lower():
                    campos.append('Proceso')
                    textos['Proceso'] = _proc
                detalle[tid][programa].append({
                    'keyword': kw,
                    'campos': campos,
                    'textos': textos,
                    'asignatura': asig_str if asig_str else 'Sin nombre'
                })

    # Cobertura = % de asignaturas unicas que abordan la tendencia
    cobertura = {}
    asig_counts = {}
    for tid in tendencias:
        n_asigs = len(asig_sets[tid])
        asig_counts[tid] = n_asigs
        cobertura[tid] = (n_asigs / total_asigs * 100) if total_asigs > 0 else 0

    ausentes = [tid for tid, pct in cobertura.items() if pct == 0]

    return {
        'matriz': matriz,
        'cobertura': cobertura,
        'detalle': detalle,
        'ausentes': ausentes,
        'asig_sets': asig_sets,
        'asig_counts': asig_counts,
        'total_asigs': total_asigs
    }


def analizar_nlp(df: pd.DataFrame) -> Dict:
    """Analisis de mineria de texto con TF-IDF (terminos y n-gramas).
    La similitud entre asignaturas se calcula de forma interactiva en la pagina.
    """
    vectorizer = TfidfVectorizer(
        max_features=100, min_df=2, max_df=0.8,
        stop_words=list(STOPWORDS_ES), ngram_range=(1, 3)
    )
    tfidf_matrix = vectorizer.fit_transform(df['Texto_Completo'])
    features = vectorizer.get_feature_names_out()
    tfidf_sum = tfidf_matrix.sum(axis=0).A1
    top_idx = tfidf_sum.argsort()[::-1][:30]
    top_terminos = {features[i]: float(tfidf_sum[i]) for i in top_idx}

    # TF-IDF por programa
    top_por_programa = {}
    for programa in df['Programa'].unique():
        df_p = df[df['Programa'] == programa]
        if len(df_p) < 5:
            continue
        vec_p = TfidfVectorizer(
            max_features=50, min_df=2, max_df=0.85,
            stop_words=list(STOPWORDS_ES), ngram_range=(1, 2)
        )
        try:
            tfidf_p = vec_p.fit_transform(df_p['Texto_Completo'])
            feat_p = vec_p.get_feature_names_out()
            sum_p = tfidf_p.sum(axis=0).A1
            idx_p = sum_p.argsort()[::-1][:20]
            top_por_programa[programa] = {feat_p[i]: float(sum_p[i]) for i in idx_p}
        except Exception:
            continue

    # N-gramas
    vec_ng = CountVectorizer(
        ngram_range=(2, 3), max_features=30,
        stop_words=list(STOPWORDS_ES), min_df=2
    )
    try:
        ng_matrix = vec_ng.fit_transform(df['Texto_Completo'])
        ng_count = ng_matrix.sum(axis=0).A1
        ng_names = vec_ng.get_feature_names_out()
        top_ngrams = sorted(
            [(ng_names[i], int(ng_count[i])) for i in range(len(ng_names))],
            key=lambda x: x[1], reverse=True
        )[:20]
    except Exception:
        top_ngrams = []

    return {
        'top_terminos': top_terminos,
        'top_por_programa': top_por_programa,
        'top_ngrams': dict(top_ngrams)
    }


def calcular_similitud_asignaturas(df_filtrado: pd.DataFrame) -> Dict:
    """Calcula similitud coseno entre asignaturas del dataframe filtrado."""
    asig_col = 'Nombre asignatura o modulo'
    df_asig = df_filtrado.groupby(asig_col)['Texto_Completo'].apply(
        lambda x: ' '.join(x)
    ).reset_index()

    if len(df_asig) < 3:
        return {'similitud_df': pd.DataFrame(), 'par_similar': None, 'n_asignaturas': len(df_asig)}

    vec_asig = TfidfVectorizer(max_features=50, stop_words=list(STOPWORDS_ES))
    tfidf_asig = vec_asig.fit_transform(df_asig['Texto_Completo'])
    sim = cosine_similarity(tfidf_asig)
    similitud_df = pd.DataFrame(
        sim,
        index=df_asig[asig_col],
        columns=df_asig[asig_col]
    )
    sim_copy = sim.copy()
    np.fill_diagonal(sim_copy, 0)
    max_idx = np.unravel_index(sim_copy.argmax(), sim_copy.shape)
    par_similar = {
        'asig1': df_asig.iloc[max_idx[0]][asig_col],
        'asig2': df_asig.iloc[max_idx[1]][asig_col],
        'similitud': float(sim_copy[max_idx])
    }
    return {
        'similitud_df': similitud_df,
        'par_similar': par_similar,
        'n_asignaturas': len(df_asig)
    }


# ============================================================================
# PAGINAS DEL DASHBOARD
# ============================================================================

def _creditos_unicos_programa(grupo: pd.DataFrame) -> int:
    """Suma créditos únicos por asignatura (sin duplicar por número de RAs)."""
    asig_col = 'Nombre asignatura o modulo'
    cred_col = next(
        (c for c in grupo.columns if c.lower().replace('é', 'e') in ('creditos', 'credito')),
        None
    )
    if asig_col not in grupo.columns or cred_col is None:
        return 0
    # Un crédito por asignatura única (primer valor no nulo)
    cred = (
        grupo.dropna(subset=[asig_col])
        .groupby(asig_col)[cred_col]
        .first()
    )
    cred_num = pd.to_numeric(cred, errors='coerce').fillna(0)
    return int(cred_num[cred_num > 0].sum())


def _creditos_por_bloque(grupo: pd.DataFrame) -> Dict[str, int]:
    """
    Desglosa créditos únicos por bloque (Institucional, Disciplinar, Electivo).
    Cada subject se cuenta una sola vez (primer valor no nulo de créditos).
    Usa _find_column para tolerar variaciones de tildes y espacios.
    """
    asig_col_real = _find_column(grupo, 'Nombre asignatura o modulo')
    cred_col = next(
        (c for c in grupo.columns if _normalize_column_name(c) in ('creditos', 'credito')),
        None
    )
    if asig_col_real is None or cred_col is None:
        return {'Institucional': 0, 'Disciplinar': 0, 'Electivo': 0, 'Total': 0}

    # Una fila por asignatura única (toma la primera fila con nombre)
    asig_df = grupo.dropna(subset=[asig_col_real]).copy()
    asig_df[cred_col] = pd.to_numeric(asig_df[cred_col], errors='coerce').fillna(0)

    credits_by_assignment = (
        asig_df.groupby(asig_col_real, as_index=False)[cred_col]
        .first()
    )

    def _sum_bloque(col_alias: str) -> int:
        col_real = _find_column(asig_df, col_alias)
        if col_real is None:
            return 0
        mask = asig_df[col_real].astype(str).apply(_normalize_value).isin(
            {'x', 'si', 's', 'sí', '1', 'true', 't', 'v', 'verdadero'}
        )
        asignaturas_validas = asig_df.loc[mask, asig_col_real].dropna().unique()
        if len(asignaturas_validas) == 0:
            return 0
        return int(
            credits_by_assignment[credits_by_assignment[asig_col_real].isin(asignaturas_validas)][cred_col].sum()
        )

    inst  = _sum_bloque('B.Institucional')
    disc  = _sum_bloque('B.Disciplinar')
    elec  = _sum_bloque('B.Electivo')
    total = int(credits_by_assignment[credits_by_assignment[cred_col] > 0][cred_col].sum())
    return {'Institucional': inst, 'Disciplinar': disc, 'Electivo': elec, 'Total': total}


def _detectar_nivel(grupo: pd.DataFrame) -> str:
    """Detecta automáticamente Pregrado o Posgrado según las columnas de bloque/componente
    presentes en el DataFrame. No requiere columna 'Nivel' explícita.
    Usa _find_column para tolerar variaciones de tildes, espacios y mayúsculas.

    - Columnas B.Institucional/B.Disciplinar/B.Electivo con datos → Pregrado
    - Columnas C.Fundamentacion/C.Profundizacion con datos → Posgrado
    - Ambos → Mixto
    """
    marcadores = {'x', 'si', 'sí', 's', '1', 'true', 't', 'v', 'verdadero'}

    def _tiene_datos(col_alias: str) -> bool:
        col_real = _find_column(grupo, col_alias)
        if col_real is None:
            return False
        return grupo[col_real].astype(str).apply(_normalize_value).isin(marcadores).any()

    tiene_b = any(_tiene_datos(c) for c in [
        'B.Institucional', 'Institucional',
        'B.Disciplinar', 'Disciplinar',
        'B.Electivo', 'Electivo'
    ])
    tiene_c = any(_tiene_datos(c) for c in [
        'C.Fundamentacion', 'Fundamentacion',
        'C.Profundizacion', 'Profundizacion'
    ])

    if tiene_b and not tiene_c:
        return 'Pregrado'
    elif tiene_c and not tiene_b:
        return 'Posgrado'
    elif tiene_b and tiene_c:
        return 'Mixto'
    return 'No identificado'


def _creditos_por_componente(grupo: pd.DataFrame) -> Dict[str, int]:
    """Desglosa créditos únicos por componente (Fundamentacion, Profundizacion).
    Usado para programas de Posgrado. Cada asignatura se cuenta una sola vez.
    Usa _find_column para tolerar variaciones de tildes y espacios.
    """
    asig_col_real = _find_column(grupo, 'Nombre asignatura o modulo')
    cred_col = next(
        (c for c in grupo.columns if _normalize_column_name(c) in ('creditos', 'credito')),
        None
    )
    if asig_col_real is None or cred_col is None:
        return {'Fundamentacion': 0, 'Profundizacion': 0, 'Total': 0}

    asig_df = grupo.dropna(subset=[asig_col_real]).copy()
    asig_df[cred_col] = pd.to_numeric(asig_df[cred_col], errors='coerce').fillna(0)

    credits_by_assignment = (
        asig_df.groupby(asig_col_real, as_index=False)[cred_col]
        .first()
    )

    def _sum_componente(col_aliases) -> int:
        if isinstance(col_aliases, str):
            col_aliases = [col_aliases]
        col_real = None
        for alias in col_aliases:
            col_real = _find_column(asig_df, alias)
            if col_real is not None:
                break
        if col_real is None:
            return 0
        mask = asig_df[col_real].astype(str).apply(_normalize_value).isin(
            {'x', 'si', 's', 'sí', '1', 'true', 't', 'v', 'verdadero'}
        )
        asignaturas_validas = asig_df.loc[mask, asig_col_real].dropna().unique()
        if len(asignaturas_validas) == 0:
            return 0
        return int(
            credits_by_assignment[credits_by_assignment[asig_col_real].isin(asignaturas_validas)][cred_col].sum()
        )

    fund  = _sum_componente(['C.Fundamentacion', 'Fundamentacion'])
    prof  = _sum_componente(['C.Profundizacion', 'Profundizacion'])
    total = int(credits_by_assignment[credits_by_assignment[cred_col] > 0][cred_col].sum())
    return {'Fundamentacion': fund, 'Profundizacion': prof, 'Total': total}


def _contar_semestres_validos(df_grupo: pd.DataFrame) -> int:
    """
    Obtiene el número máximo de semestres desde la columna Semestre.
    
    La duración del programa es el valor MÁXIMO encontrado en Semestre,
    no el conteo de valores únicos.
    
    Ejemplo:
    - Si Semestre tiene [1, 1, 2, 2, 3, 3] → Retorna 3 (máximo)
    - Si Semestre tiene [1, 2, 4] → Retorna 4 (máximo)
    
    Args:
        df_grupo: dataframe del programa (ya filtrado por programa/modalidad/sede/nivel)
    
    Returns:
        Número máximo de semestres (duración del programa)
    """
    if df_grupo.empty:
        return 0
    
    # Buscar columna Semestre
    sem_col = _find_column(df_grupo, 'Semestre')
    if sem_col is None:
        return 0
    
    # Obtener valores numéricos de Semestre
    semestres = pd.to_numeric(df_grupo[sem_col], errors='coerce')
    semestres = semestres.dropna()
    
    if len(semestres) == 0:
        return 0
    
    # Retornar el MÁXIMO
    max_sem = int(semestres.max())
    return max_sem


def leer_totales_programa(uploaded_files) -> Dict[str, Dict[str, int]]:
    """
    Lee los totales OFICIALES de créditos desde las filas de resumen declaradas
    al final de cada Excel (ej: 'Total créditos programa', 'Total créditos bloque
    institucional', etc.).

    Estrategia: escanea TODAS las celdas buscando las etiquetas de totales;
    el valor oficial se lee de la columna de Créditos en la misma fila.
    El número de filas varía por programa, por lo que NO se usa posición fija.
    """
    def _norm(s: str) -> str:
        return unicodedata.normalize('NFKD', str(s).strip().lower()).encode('ascii', 'ignore').decode('ascii')

    def _leer_nombre_programa(f):
        """Lee el nombre del programa desde la hoja de perfil."""
        try:
            f.seek(0)
            df_perfil = pd.read_excel(f, sheet_name='Paso1 Analisis perfil egreso', header=None, nrows=10, engine='openpyxl')
            if df_perfil is not None and len(df_perfil) > 2 and len(df_perfil.columns) > 0:
                val = df_perfil.iloc[2, 0]
                if val is not None and str(val).strip():
                    return str(val).strip()
        except:
            pass
        return None

    totales: Dict[str, Dict[str, int]] = {}

    for f in uploaded_files:
        nombre = f.name
        prog_key = (nombre
                .replace("FormatoRA_", "").replace("FormatoRA-", "")
                .replace("_PBOG", "").replace("_VNAL", "").replace("_PMED", "")
                .replace("_HBOG", "").replace("_HMED", "").replace("_HVAL", "")
                .replace(".xlsx", "").replace(".xls", "")).strip()
        
        prog_real = _leer_nombre_programa(f)
        claves_programa = [prog_key]
        if prog_real and prog_real not in claves_programa:
            claves_programa.append(prog_real)
        
        try:
            f.seek(0)
            raw = pd.read_excel(
                f, sheet_name='Paso 5 Estrategias micro',
                header=None, engine='openpyxl'
            )

            # Encontrar columna de Créditos por encabezado (fila 1 = Excel fila 2)
            cred_col_idx = next(
                (i for i, v in enumerate(raw.iloc[1])
                 if pd.notna(v) and 'dito' in _norm(str(v))),
                9  # fallback: columna J (índice 9)
            )

            pt: Dict[str, int] = {}
            nrows, ncols = raw.shape

            for r in range(nrows):
                for c in range(ncols):
                    cell = raw.iloc[r, c]
                    if not pd.notna(cell):
                        continue
                    cn = _norm(str(cell))
                    
                    # Buscar asignaturas: usar misma lógica que test_generar_excel.py
                    # Prioridad: 1) Total modulos, 2) Total asignaturas del programa, 3) Total materias
                    if 'asignaturas' not in pt:
                        if ('total modulos' in cn or 
                            'total asignaturas del programa' in cn or 
                            ('total materias' in cn and 'asignatura' in cn)):
                            next_col = c + 1
                            if next_col < ncols:
                                raw_val = raw.iloc[r, next_col]
                                try:
                                    if isinstance(raw_val, (int, float)):
                                        val = int(raw_val)
                                    elif pd.notna(raw_val):
                                        val = int(float(str(raw_val).strip()))
                                    else:
                                        val = 0
                                except:
                                    val = 0
                                if val > 0:
                                    pt['asignaturas'] = val
                    
                    # Leer el crédito de la misma fila en la columna de Créditos
                    raw_val = (raw.iloc[r, cred_col_idx]
                               if cred_col_idx < ncols else None)
                    try:
                        val = int(float(raw_val)) if pd.notna(raw_val) else 0
                    except (ValueError, TypeError):
                        val = 0

                    if 'total creditos programa' in cn or 'total creditos del programa' in cn:
                        pt['total'] = val
                    elif 'bloque inst' in cn:
                        pt['institucional'] = val
                    elif 'bloque disc' in cn:
                        pt['disciplinar'] = val
                    elif 'bloque elec' in cn:
                        pt['electivo'] = val
                    elif 'componente fund' in cn or 'fundamentacion' in cn:
                        pt['fundamentacion'] = val
                    elif 'componente prof' in cn or 'profundizacion' in cn:
                        pt['profundizacion'] = val

            for clave in claves_programa:
                totales[clave] = pt
                # Debug: mostrar qué se leyó
                print(f"LEER_TOTALES: clave={clave}, pt={pt}")
        except Exception as e:
            print(f"LEER_TOTALES ERROR: {e}")
            for clave in claves_programa:
                totales[clave] = {}

    return totales


def pagina_inicio(df: pd.DataFrame, totales_oficiales: Optional[Dict] = None):
    """Pagina de inicio con metricas generales."""
    st.title("Análisis Temático Microcurricular")
    st.markdown("---")
    st.info(
        "**Bienvenido/a.** Este dashboard analiza los microcurrículos de los programas "
        "académicos en tres dimensiones principales: **cobertura temática** (qué temas se trabajan "
        "y con qué profundidad), **tendencias globales** (alineación con retos del mundo actual) y "
        "**análisis de texto** (conceptos clave, similitudes y redundancias). "
        "Navega por las secciones del menú lateral para explorar cada dimensión."
    )

    if 'Nivel' in df.columns:
        df = df.copy()
        df['Nivel'] = df['Nivel'].astype(str).str.strip().replace({'nan': 'No identificado', '': 'No identificado'}).str.title()
        unique_programs = df[['Programa', 'Modalidad', 'Sede', 'Nivel']].drop_duplicates()
    else:
        unique_programs = df[['Programa', 'Modalidad', 'Sede']].drop_duplicates()
    asignaturas = df['Nombre asignatura o modulo'].nunique()
    total_registros = len(df)
    total_palabras = df['Texto_Completo'].str.split().str.len().sum()
    count_modalidad = unique_programs['Modalidad'].fillna('No identificado').value_counts()
    presencial_count = int(count_modalidad.get('Presencial', 0))
    virtual_count = int(count_modalidad.get('Virtual', 0))
    hibrido_count = int(count_modalidad.get('Híbrido', 0))
    if 'Nivel' in unique_programs.columns:
        count_nivel = unique_programs['Nivel'].fillna('No identificado').value_counts()
        pregrado_count = int(count_nivel.get('Pregrado', 0))
        posgrado_count = int(count_nivel.get('Posgrado', 0))
    else:
        pregrado_count = 0
        posgrado_count = 0

    col1, col2, col3, col4 = st.columns(4)
    col1.metric(
        "Programas cargados", len(unique_programs),
        help="Número de programas distintos considerando Programa + Modalidad + Sede."
    )
    col2.metric(
        "Presenciales", presencial_count,
        help="Cantidad de programas cargados en modalidad presencial."
    )
    col3.metric(
        "Virtuales", virtual_count,
        help="Cantidad de programas cargados en modalidad virtual."
    )
    col4.metric(
        "Híbridos", hibrido_count,
        help="Cantidad de programas cargados en modalidad híbrida."
    )

    if 'Nivel' in unique_programs.columns:
        col5, col6 = st.columns(2)
        col5.metric(
            "Pregrado", pregrado_count,
            help="Cantidad de programas de nivel Pregrado."
        )
        col6.metric(
            "Posgrado", posgrado_count,
            help="Cantidad de programas de nivel Posgrado."
        )
        st.markdown("---")
    if 'Nivel' in unique_programs.columns:
        col_a, col_b, col_c = st.columns(3)
    else:
        col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("Programas por modalidad")
        st.caption(
            "Cantidad de programas distintos por modalidad y sede. Cada programa es una combinación "
            "única de Programa + Modalidad + Sede."
        )
        programas_x_modalidad = (
            unique_programs
            .groupby(['Modalidad', 'Sede'])
            .size()
            .reset_index(name='Programas')
        )
        fig = px.bar(
            programas_x_modalidad,
            x='Modalidad',
            y='Programas',
            color='Sede',
            text='Programas',
            labels={'Programas': 'N° de programas', 'Modalidad': 'Modalidad'},
            category_orders={'Modalidad': ['Presencial', 'Virtual', 'Híbrido']},
            color_discrete_map=COLORES_SEDE
        )
        fig.update_layout(showlegend=True, height=320, barmode='stack')
        fig.update_traces(textposition='inside', textfont=dict(size=11, color='white'), textangle=0)
        totales = programas_x_modalidad.groupby('Modalidad')['Programas'].sum().reset_index()
        max_val = totales['Programas'].max()
        for i, row in totales.iterrows():
            fig.add_annotation(
                x=row['Modalidad'],
                y=row['Programas'] + max_val * 0.1,
                text=f"<b>{row['Programas']}</b>",
                showarrow=False,
                font=dict(size=14, color='#0F83FF')
            )
        st.plotly_chart(fig, use_container_width=True)

    with col_b:
        if 'Nivel' in unique_programs.columns:
            st.subheader("Programas por Nivel")
            st.caption("Distribución de programas por nivel académico.")
            niveles_df = (
                unique_programs
                .groupby('Nivel')
                .size()
                .reset_index(name='Programas')
                .sort_values('Programas', ascending=False)
            )
            fig_nivel = px.bar(
                niveles_df,
                x='Nivel',
                y='Programas',
                color='Nivel',
                text='Programas',
                color_discrete_map={
                    'Pregrado': '#0F83FF',
                    'Posgrado': '#0FFF8B',
                    'No identificado': '#8c8c8c'
                },
                labels={'Programas': 'N° de programas'}
            )
            max_val = niveles_df['Programas'].max()
            fig_nivel.update_layout(
                showlegend=False, 
                height=320,
                yaxis=dict(range=[0, max_val * 1.3])
            )
            fig_nivel.update_traces(textposition='outside')
            st.plotly_chart(fig_nivel, use_container_width=True)
        else:
            st.subheader("Distribución Tipo de Saber (todos los programas)")
            st.caption(
                "**Saber** = conocimiento teórico/conceptual. "
                "**SaberHacer** = habilidades prácticas y procedimentales. "
                "**SaberSer** = actitudes, valores y dimensión ética. "
                "Un currículo equilibrado debería contemplar los tres tipos."
            )
            tipo_saber = df['Tipo de Saber'].value_counts().reset_index()
            tipo_saber.columns = ['Tipo', 'Cantidad']
            fig = px.pie(tipo_saber, values='Cantidad', names='Tipo',
                         color_discrete_sequence=PALETA_AZUL)
            fig.update_layout(height=320)
            st.plotly_chart(fig, use_container_width=True)

    with col_c:
        if 'Nivel' in unique_programs.columns:
            st.subheader("Distribución Tipo de Saber (todos los programas)")
            st.caption(
                "**Saber** = conocimiento teórico/conceptual. "
                "**SaberHacer** = habilidades prácticas y procedimentales. "
                "**SaberSer** = actitudes, valores y dimensión ética. "
                "Un currículo equilibrado debería contemplar los tres tipos."
            )
            tipo_saber = df['Tipo de Saber'].value_counts().reset_index()
            tipo_saber.columns = ['Tipo', 'Cantidad']
            fig = px.pie(tipo_saber, values='Cantidad', names='Tipo',
                         color_discrete_sequence=PALETA_AZUL)
            fig.update_layout(height=320)
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("Resumen por Programa")
    st.caption(
        "Los créditos oficiales se leen directamente desde las filas de totales declaradas "
        "al final de cada Excel ('Total créditos programa', 'Total créditos bloque...'). "
        "La columna **Diferencia** compara el total oficial con la suma de los tres bloques."
    )

    resumen_rows = []
    grupos_resumen = ['Programa', 'Modalidad', 'Sede', 'Nivel']
    
    unique_groups = df[grupos_resumen].drop_duplicates()
    
    for _, row in unique_groups.iterrows():
        prog = row['Programa']
        modalidad = row['Modalidad']
        sede = row['Sede']
        nivel_detectado = row['Nivel']
        
        g = df[(df['Programa'] == prog) & (df['Modalidad'] == modalidad) & (df['Sede'] == sede) & (df['Nivel'] == nivel_detectado)]
        
        st.session_state[f'debug_grupo_{prog}_{modalidad}_{sede}_{nivel_detectado}_filas'] = f"Grupo {prog} | {modalidad} | {sede} | {nivel_detectado}: {len(g)} filas"

# Totales oficiales del Excel (footer rows)
        # Buscar en totales_oficiales por coincidencia exacta o parcial del nombre
        of = None
        totales_keys = list((totales_oficiales or {}).keys())
        
        # Primero buscar coincidencia exacta
        if prog in totales_keys:
            of = totales_oficiales[prog]
        else:
            # Buscar coincidencia parcial
            for k in totales_keys:
                if prog.lower() in k.lower() or k.lower() in prog.lower():
                    of = totales_oficiales[k]
                    break
                # También verificar coincidencia con palabras clave
                prog_words = prog.lower().split()
                k_words = k.lower().split()
                if any(w in k_words for w in prog_words if len(w) > 3):
                    of = totales_oficiales[k]
                    break
        
        if not of:
            of = {}
        
        cr_total = int(of.get('total', 0)) if of else 0
        asigs_oficial = int(of.get('asignaturas', 0)) if of else 0

        if nivel_detectado == 'Posgrado':
            # ── Posgrado: usar componentes C.* ──────────────────────────────
            cr_fund = of.get('fundamentacion', 0)
            cr_prof = of.get('profundizacion', 0)

            if cr_total == 0:
                cp = _creditos_por_componente(g)
                cr_total = cp['Total']
                cr_fund  = cp['Fundamentacion']
                cr_prof  = cp['Profundizacion']

            suma_componentes = cr_fund + cr_prof
            diferencia = cr_total - suma_componentes

            # Usar asignaturas oficiales del Excel (ya leído desde totales_programa)
            # NO recalcular aquí porque el dataframe está filtrado por Tipo de Saber
            asigs_mostrar = asigs_oficial if asigs_oficial else 0
            row = {
                'Programa':             prog,
                'Modalidad':            modalidad,
                'Sede':                 sede,
                'Nivel':                nivel_detectado,
                'Asignaturas':          asigs_mostrar,
                'Asig. Oficial':       asigs_oficial if asigs_oficial else None,
                'Semestres':            _contar_semestres_validos(g),
                'Cr. Total':            cr_total,
                'Cr. Fundamentacion':   cr_fund,
                'Cr. Profundizacion':   cr_prof,
                'Suma componentes':     suma_componentes,
                'Diferencia':           diferencia,
            }
        else:
            # ── Pregrado (o No identificado/Mixto): usar bloques B.* ─────────
            cr_inst = of.get('institucional', 0)
            cr_disc = of.get('disciplinar',   0)
            cr_elec = of.get('electivo',      0)

            if cr_total == 0:
                bl = _creditos_por_bloque(g)
                cr_total = bl['Total']
                cr_inst  = bl['Institucional']
                cr_disc  = bl['Disciplinar']
                cr_elec  = bl['Electivo']

            suma_bloques = cr_inst + cr_disc + cr_elec
            diferencia   = cr_total - suma_bloques

            # Usar asignaturas oficiales del Excel (ya leído desde totales_programa)
            # NO recalcular aquí porque el dataframe está filtrado por Tipo de Saber
            asigs_mostrar = asigs_oficial if asigs_oficial else 0
            row = {
                'Programa':          prog,
                'Modalidad':         modalidad,
                'Sede':              sede,
                'Nivel':             nivel_detectado,
                'Asignaturas':       asigs_mostrar,
                'Asig. Oficial':     asigs_oficial if asigs_oficial else None,
                'Semestres':         _contar_semestres_validos(g),
                'Cr. Total':         cr_total,
                'B. Institucional': cr_inst,
                'B. Disciplinar':   cr_disc,
                'B. Electivo':      cr_elec,
                'Suma bloques':      suma_bloques,
                'Diferencia':        diferencia,
            }

        resumen_rows.append(row)

    resumen = pd.DataFrame(resumen_rows)

    # Convertir columnas numéricas a int (llenar NaNs con 0 primero)
    for col in resumen.columns:
        if col not in ['Programa', 'Modalidad', 'Sede', 'Nivel']:
            resumen[col] = resumen[col].fillna(0).astype(int)

    # Reordenar columnas: columnas comunes primero, luego las específicas por nivel
    cols_comunes = ['Programa', 'Modalidad', 'Sede', 'Nivel', 'Asignaturas', 'Asig. Oficial', 'Semestres', 'Cr. Total']
    cols_pregrado = ['B. Institucional', 'B. Disciplinar', 'B. Electivo', 'Suma bloques']
    cols_posgrado = ['Cr. Fundamentacion', 'Cr. Profundizacion', 'Suma componentes']
    cols_extra = ['Diferencia']
    cols_presentes = [c for c in cols_comunes + cols_pregrado + cols_posgrado + cols_extra if c in resumen.columns]
    resumen = resumen[cols_presentes]

    # ── Ordenar por Nivel (Pregrado primero, luego Posgrado) ────────────────────
    nivel_order = {'Pregrado': 0, 'Posgrado': 1, 'Mixto': 2, 'No identificado': 3}
    resumen['_nivel_sort'] = resumen['Nivel'].map(nivel_order).fillna(4)
    resumen = resumen.sort_values(['_nivel_sort', 'Programa', 'Modalidad', 'Sede']).drop('_nivel_sort', axis=1)
    resumen = resumen.reset_index(drop=True)

    # ── Paginación ────────────────────────────────────────────────────────────────
    rows_per_page = 10
    total_rows = len(resumen)
    total_pages = (total_rows + rows_per_page - 1) // rows_per_page

    # Inicializar session_state para página actual
    if 'pagina_resumen' not in st.session_state:
        st.session_state.pagina_resumen = 0

    # Controles de paginación
    col_prev, col_info, col_next = st.columns([1, 2, 1])
    with col_prev:
        if st.button("⬅️ Anterior", disabled=st.session_state.pagina_resumen == 0):
            st.session_state.pagina_resumen -= 1
            st.rerun()

    with col_info:
        st.markdown(f"<div style='text-align:center'><b>Página {st.session_state.pagina_resumen + 1} de {total_pages}</b> ({total_rows} registros)</div>", unsafe_allow_html=True)

    with col_next:
        if st.button("Siguiente ➡️", disabled=st.session_state.pagina_resumen >= total_pages - 1):
            st.session_state.pagina_resumen += 1
            st.rerun()

    # Obtener datos de la página actual
    start_idx = st.session_state.pagina_resumen * rows_per_page
    end_idx = start_idx + rows_per_page
    resumen_pagina = resumen.iloc[start_idx:end_idx]

    def _color_dif(val):
        if val == 0:
            return 'background-color:#d4edda;color:#155724'
        return 'background-color:#f8d7da;color:#721c24'

    st.markdown("""
    <style>
    .Resumen por Programa .stDataFrame thead th {
        background-color: #1E88E5 !important;
        color: white !important;
        font-weight: bold !important;
        font-size: 14px !important;
    }
    .Resumen por Programa .stDataFrame tbody td {
        font-size: 13px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.dataframe(
        resumen_pagina.style.map(_color_dif, subset=['Diferencia']),
        use_container_width=True, 
        hide_index=True,
        column_config={
            "Programa": st.column_config.TextColumn(
                "Programa",
                pinned=True
            )
        }
    )
    
    programas_con_dif = resumen[resumen['Diferencia'] != 0][['Programa', 'Modalidad', 'Sede', 'Diferencia']].sort_values('Diferencia', key=abs, ascending=False)
    if not programas_con_dif.empty:
        st.warning(f"**{len(programas_con_dif)} programa(s) con diferencias que deben validarse:**")
        for _, row in programas_con_dif.iterrows():
            st.caption(f"• {row['Programa']} ({row['Modalidad']} - {row['Sede']}): Diferencia = {row['Diferencia']}")
    
    st.caption(
        "**Pregrado** → columnas B. Institucional / Disciplinar / Electivo.  "
        "**Posgrado** → columnas Cr. Fundamentacion / Profundizacion.  "
        "**Diferencia = 0** coincide | **≠ 0** revisar el Excel."
    )

    # debug_keys = [k for k in st.session_state.keys() if k.startswith('debug_')]
    # if debug_keys:
    #     with st.expander("DEBUG: Conteo de Asignaturas"):
    #         for k in sorted(debug_keys):
    #             st.code(st.session_state[k])


def pagina_cobertura(df: pd.DataFrame, resultados: Dict):
    """Pagina de cobertura y densidad tematica."""
    st.title("Cobertura y Densidad Temática")
    st.markdown("---")
    st.info(
        "**¿Qué mide esta sección?** Analiza los **núcleos temáticos** declarados en los "
        "microcurrículos: qué temas se trabajan, con qué frecuencia, y qué tan distribuidos "
        "están entre asignaturas. Un currículo con alta **diversidad** temática cubre más "
        "áreas del conocimiento."
    )

    # ── Revisión y depuración de núcleos ─────────────────────────────────────
    todos_nucleos_ord = resultados['nucleos_counter'].most_common()
    if 'nucleos_excluidos' not in st.session_state:
        st.session_state['nucleos_excluidos'] = []

    with st.expander("🔍 Revisar y depurar núcleos detectados", expanded=False):
        st.caption(
            "Revisa los núcleos identificados automáticamente. Si alguno no es un núcleo "
            "temático válido (fragmentos, errores de formato, etc.), exclúyelo del análisis. "
            "Los cambios se aplican inmediatamente a todos los gráficos de esta página."
        )
        df_revision = pd.DataFrame(
            [{'Núcleo': n, 'Menciones': c} for n, c in todos_nucleos_ord]
        )
        st.dataframe(df_revision, use_container_width=True, hide_index=True, height=250)

        opciones = [n for n, _ in todos_nucleos_ord]
        excluidos_sel = st.multiselect(
            "Seleccionar núcleos a excluir:",
            options=opciones,
            default=[x for x in st.session_state['nucleos_excluidos'] if x in opciones],
            placeholder="Escribe para buscar un núcleo…",
            key='ms_nucleos_excluidos'
        )
        st.session_state['nucleos_excluidos'] = excluidos_sel
        if excluidos_sel:
            st.warning(f"**{len(excluidos_sel)}** núcleo(s) excluido(s) del análisis.")
        else:
            st.success("Todos los núcleos detectados están incluidos en el análisis.")

    # Aplicar exclusiones al counter y recalcular métricas derivadas
    excluidos = set(st.session_state.get('nucleos_excluidos', []))
    nucleos_counter = Counter({k: v for k, v in resultados['nucleos_counter'].items()
                               if k not in excluidos})
    nucleos_unicos   = len(nucleos_counter)
    total_menciones  = sum(nucleos_counter.values())
    if nucleos_unicos > 1:
        _frec = np.array(list(nucleos_counter.values()), dtype=float)
        _prob = _frec / _frec.sum()
        _h    = entropy(_prob, base=2)
        diversidad = (_h / np.log2(nucleos_unicos)) * 100
    else:
        diversidad = 0.0
    # Reconstruir matriz filtrada (top 20)
    top_20_fil = [n for n, _ in nucleos_counter.most_common(20)]
    matriz = pd.DataFrame(0, index=df['Programa'].unique(), columns=top_20_fil)
    for _, row in df.iterrows():
        programa = row['Programa']
        nucleos_raw = str(row.get('Nucleos tematicos', ''))
        if nucleos_raw and nucleos_raw != 'nan':
            for nuc in [limpiar_nucleo(n.strip()) for n in _split_nucleos(nucleos_raw)
                        if es_nucleo_valido(n.strip())[0]]:
                if nuc in top_20_fil and nuc not in excluidos:
                    matriz.loc[programa, nuc] += 1

    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric(
        "Núcleos Únicos", nucleos_unicos,
        help="Número de temas o núcleos temáticos diferentes identificados en todos los programas"
    )
    col2.metric(
        "Total Menciones", f"{total_menciones:,}",
        help="Cuántas veces en total aparecen núcleos temáticos (pueden repetirse entre asignaturas)"
    )
    col3.metric(
        "Promedio por Asignatura", f"{resultados['promedio_densidad']:.1f}",
        help="Promedio de núcleos temáticos declarados por asignatura"
    )
    col4.metric(
        "Diversidad Temática", f"{diversidad:.1f}/100",
        help=(
            "Índice de diversidad (Entropía de Shannon normalizada). "
            "100 = máxima diversidad (todos los temas con igual peso). "
            "0 = mínima diversidad (un solo tema domina todo)."
        )
    )

    st.markdown("---")

    # ── Matriz mejorada ──────────────────────────────────────────────────────
    st.subheader("Presencia de Núcleos Temáticos por Programa (Top 20)")
    st.caption(
        "Cada celda muestra cuántas veces aparece ese núcleo temático en las filas "
        "del programa. Los núcleos se extraen de la columna «Núcleos temáticos» separados "
        "por coma o punto y coma. Los más frecuentes (como términos cortos) pueden provenir "
        "de asignaturas del bloque institucional compartido entre programas."
    )
    if not matriz.empty:
        # Truncar nombres de nucleos para mejor legibilidad
        max_len = 35
        col_cortos = {
            c: (c[:max_len] + '…' if len(c) > max_len else c)
            for c in matriz.columns
        }
        matriz_display = matriz.rename(columns=col_cortos)

        fig = px.imshow(
            matriz_display.values,
            x=matriz_display.columns.tolist(),
            y=matriz_display.index.tolist(),
            color_continuous_scale='Blues',
            text_auto=True,
            aspect='auto',
            labels={'color': 'Menciones'}
        )
        fig.update_layout(
            height=max(300, len(matriz_display) * 60 + 150),
            xaxis_tickangle=45,
            xaxis_title="Núcleo Temático (top 20 más frecuentes)",
            yaxis_title="Programa"
        )
        st.plotly_chart(fig, use_container_width=True)

        # Mostrar tabla de nombres completos para referencia
        with st.expander("Ver nombres completos de los núcleos (referencia)"):
            nombres_df = pd.DataFrame([
                {'Abreviado': v, 'Nombre completo': k}
                for k, v in col_cortos.items()
                if k != v
            ])
            if not nombres_df.empty:
                st.dataframe(nombres_df, use_container_width=True, hide_index=True)
            else:
                st.caption("Todos los nombres son cortos y se muestran completos.")

    st.markdown("---")

    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("Top 20 Núcleos Temáticos")
        st.caption(
            "Frecuencia = cuántas veces aparece ese núcleo en todos los registros. "
            "Los núcleos muy cortos (1-2 palabras) suelen ser temas generales del bloque institucional."
        )
        top_nucleos = nucleos_counter.most_common(20)
        df_nucleos = pd.DataFrame(top_nucleos, columns=['Nucleo', 'Frecuencia'])
        # Truncar para visualización
        df_nucleos['Nucleo_display'] = df_nucleos['Nucleo'].apply(
            lambda x: x[:45] + '…' if len(x) > 45 else x
        )
        fig = px.bar(df_nucleos, y='Nucleo_display', x='Frecuencia',
                     orientation='h', color='Frecuencia',
                     color_continuous_scale='Blues',
                     labels={'Nucleo_display': 'Núcleo Temático', 'Frecuencia': 'Menciones'},
                     hover_data={'Nucleo': True, 'Frecuencia': True, 'Nucleo_display': False})
        fig.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig, use_container_width=True)

    with col_b:
        st.subheader("Densidad Temática (Top 20 Asignaturas)")
        st.caption(
            "Asignaturas con más núcleos temáticos declarados. Alta densidad = mayor "
            "explicitación de contenidos. Asignaturas muy densas pueden tener temas duplicados."
        )
        densidad_top = resultados['densidad_asignatura'].head(20)
        df_densidad = pd.DataFrame({
            'Asignatura': densidad_top.index,
            'Nucleos': densidad_top.values
        })
        df_densidad['Asig_display'] = df_densidad['Asignatura'].apply(
            lambda x: x[:40] + '…' if len(str(x)) > 40 else x
        )
        fig = px.bar(df_densidad, y='Asig_display', x='Nucleos',
                     orientation='h', color='Nucleos',
                     color_continuous_scale='Blues',
                     labels={'Asig_display': 'Asignatura', 'Nucleos': 'N° Núcleos'},
                     hover_data={'Asignatura': True, 'Nucleos': True, 'Asig_display': False})
        fig.update_layout(height=600, yaxis={'categoryorder': 'total ascending'})
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("Explorar Núcleos por Programa")
    st.caption("Selecciona un programa para ver sus núcleos temáticos más frecuentes.")
    programa_sel = st.selectbox("Seleccionar programa:", df['Programa'].unique(),
                                key='sel_prog_cobertura')
    df_prog = df[df['Programa'] == programa_sel]

    nucleos_prog = []
    for _, row in df_prog.iterrows():
        nucleos_raw = str(row.get('Nucleos tematicos', ''))
        if nucleos_raw and nucleos_raw != 'nan':
            nucleos_prog.extend([
                limpiar_nucleo(n.strip())
                for n in _split_nucleos(nucleos_raw)
                if es_nucleo_valido(n.strip())[0] and limpiar_nucleo(n.strip()) not in excluidos
            ])

    if nucleos_prog:
        counter_prog = Counter(nucleos_prog)
        df_prog_nucleos = pd.DataFrame(
            counter_prog.most_common(15),
            columns=['Nucleo', 'Frecuencia']
        )
        df_prog_nucleos['Nucleo_display'] = df_prog_nucleos['Nucleo'].apply(
            lambda x: x[:50] + '…' if len(x) > 50 else x
        )
        fig = px.bar(
            df_prog_nucleos, x='Nucleo_display', y='Frecuencia',
            color='Frecuencia', color_continuous_scale='Viridis',
            title=f"Top 15 Núcleos Temáticos — {programa_sel}",
            labels={'Nucleo_display': 'Núcleo Temático', 'Frecuencia': 'Menciones'},
            hover_data={'Nucleo': True, 'Frecuencia': True, 'Nucleo_display': False}
        )
        fig.update_layout(xaxis_tickangle=45, height=450)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info(f"No se encontraron núcleos temáticos para {programa_sel}.")


def pagina_tendencias(df: pd.DataFrame, tendencias: Dict, resultados: Dict):
    """Pagina de alineacion con tendencias globales."""
    st.title("Alineacion con Tendencias Globales")
    st.markdown("---")

    st.info(
        "**¿Qué mide esta sección?** Identifica qué tendencias del mundo actual "
        "están presentes en los planes de estudio. La **cobertura** indica el "
        f"porcentaje de asignaturas (sobre {resultados.get('total_asigs', '?')} en total) "
        "que abordan cada tendencia, detectada mediante palabras clave en los Resultados "
        "de Aprendizaje, Núcleos Temáticos e Indicadores de Logro."
    )

    total_tend = len(tendencias)
    detectadas = total_tend - len(resultados['ausentes'])
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Tendencias Evaluadas", total_tend,
                help="Número total de tendencias globales configuradas para buscar")
    col2.metric("Tendencias Detectadas", detectadas,
                help="Tendencias presentes en al menos una asignatura")
    col3.metric("Tendencias Ausentes", len(resultados['ausentes']),
                help="Tendencias sin ninguna asignatura que las aborde — son brechas curriculares")
    col4.metric("Total Asignaturas", resultados.get('total_asigs', '?'),
                help="Número de asignaturas únicas analizadas en todos los programas")

    st.markdown("---")

    # Helper de export — informe gerencial con formato
    def _build_listados_excel(dataframe: pd.DataFrame, tendencia_nombre: str = '', detalle_tend: dict = None) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
        import datetime

        # ── Paleta ──────────────────────────────────────────────────────────
        AZUL_OSCURO  = '0F385A'
        AZUL_MEDIO   = '1565A7'
        AZUL_CLARO   = '1FB2DE'
        GRIS_FILA    = 'EBF5FB'
        BLANCO       = 'FFFFFF'
        AMARILLO_ACC = 'FBAF17'

        def fill(hex_color):
            return PatternFill('solid', fgColor=hex_color)

        def borde_fino():
            s = Side(style='thin', color='D0D7DE')
            return Border(left=s, right=s, top=s, bottom=s)

        def borde_header():
            s = Side(style='medium', color=AZUL_OSCURO)
            return Border(left=s, right=s, top=s, bottom=s)

        def apply_header_row(ws, row_idx, labels, col_widths):
            for ci, label in enumerate(labels, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=label)
                cell.fill = fill(AZUL_OSCURO)
                cell.font = Font(name='Calibri', bold=True, color=BLANCO, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = borde_header()
                ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
            ws.row_dimensions[row_idx].height = 28

        def apply_data_rows(ws, data_rows, start_row):
            for ri, row_data in enumerate(data_rows):
                row_idx = start_row + ri
                bg = GRIS_FILA if ri % 2 == 0 else BLANCO
                for ci, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=ci, value=value)
                    cell.fill = fill(bg)
                    cell.font = Font(name='Calibri', size=10, color='1A1A2E')
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                    cell.border = borde_fino()
                ws.row_dimensions[row_idx].height = 18

        def write_title_block(ws, title, subtitle, n_cols, fecha_str):
            # Fila 1: título principal
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=n_cols)
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.fill = fill(AZUL_OSCURO)
            title_cell.font = Font(name='Calibri', bold=True, color=BLANCO, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 34
            ws.row_dimensions[2].height = 34

            # Fila 3: subtítulo
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=n_cols - 1)
            sub_cell = ws.cell(row=3, column=1, value=subtitle)
            sub_cell.fill = fill(AZUL_MEDIO)
            sub_cell.font = Font(name='Calibri', italic=True, color=BLANCO, size=10)
            sub_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

            # Fecha en última columna de fila 3
            fecha_cell = ws.cell(row=3, column=n_cols, value=fecha_str)
            fecha_cell.fill = fill(AZUL_MEDIO)
            fecha_cell.font = Font(name='Calibri', color=AMARILLO_ACC, size=10, bold=True)
            fecha_cell.alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[3].height = 20

            # Fila 4: separador
            for ci in range(1, n_cols + 1):
                sep = ws.cell(row=4, column=ci, value='')
                sep.fill = fill(AZUL_CLARO)
            ws.row_dimensions[4].height = 4

        # ── Preparar datos ───────────────────────────────────────────────────
        df_programas = (
            dataframe[['Modalidad', 'Nivel', 'Programa']]
            .drop_duplicates()
            .sort_values(['Modalidad', 'Nivel', 'Programa'])
            .reset_index(drop=True)
        )
        asig_col = next(
            (c for c in dataframe.columns if 'nombre asignatura' in c.lower()),
            'Nombre asignatura o modulo'
        )
        df_asignaturas = (
            dataframe[['Programa', asig_col]]
            .dropna(subset=[asig_col])
            .drop_duplicates()
            .rename(columns={asig_col: 'Nombre de asignatura', 'Programa': 'Nombre del programa'})
            .sort_values(['Nombre del programa', 'Nombre de asignatura'])
            .reset_index(drop=True)
        )

        # Construir lookup (prog, asig) → {keywords, campos, textos} desde detalle_tend
        CAMPO_LABEL = {
            'RA': 'Resultado de aprendizaje',
            'Nucleos': 'Núcleos temáticos',
            'Indicadores': 'Indicadores de logro',
            'Proceso': 'Proceso Responsable',
        }
        lookup_det = {}
        if detalle_tend:
            for prog, hallazgos in detalle_tend.items():
                for h in hallazgos:
                    asig = str(h.get('asignatura', '')).strip()
                    key = (str(prog).strip(), asig)
                    if key not in lookup_det:
                        lookup_det[key] = {'keywords': set(), 'campos': set(), 'textos': {}}
                    kw = h.get('keyword', '')
                    if kw:
                        lookup_det[key]['keywords'].add(kw)
                    for c in h.get('campos', []):
                        lookup_det[key]['campos'].add(c)
                    # Acumular textos: si el campo ya existe y el texto es distinto, no duplicar
                    for c, t in h.get('textos', {}).items():
                        existing = lookup_det[key]['textos'].get(c, '')
                        if t and t != existing:
                            # Concatenar si es texto diferente del mismo campo
                            lookup_det[key]['textos'][c] = (existing + '\n---\n' + t) if existing else t

        fecha_str = datetime.date.today().strftime('%d/%m/%Y')
        tend_label = f'Tendencia: {tendencia_nombre}' if tendencia_nombre else 'Todos los programas cargados'
        wb = Workbook()

        # ═══════════════════════════════════════════════════════════════════
        # HOJA 1 — Programas
        # ═══════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Programas'
        ws1.sheet_view.showGridLines = False

        cols_prog  = ['Modalidad', 'Nivel', 'Programa']
        widths_prog = [18, 18, 52]
        write_title_block(ws1, 'Listado de Programas Académicos', tend_label, len(cols_prog), fecha_str)
        apply_header_row(ws1, 5, cols_prog, widths_prog)

        # Fila de resumen en color acento antes de los datos
        ws1.merge_cells(start_row=6, start_column=1, end_row=6, end_column=len(cols_prog))
        res_cell = ws1.cell(row=6, column=1, value=f'Total de programas: {len(df_programas)}')
        res_cell.fill = fill(AMARILLO_ACC)
        res_cell.font = Font(name='Calibri', bold=True, color=AZUL_OSCURO, size=10)
        res_cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws1.row_dimensions[6].height = 16

        apply_data_rows(ws1, df_programas.values.tolist(), start_row=7)
        ws1.freeze_panes = 'A7'

        # ═══════════════════════════════════════════════════════════════════
        # HOJA 2 — Asignaturas
        # ═══════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet('Asignaturas')
        ws2.sheet_view.showGridLines = False

        N_ASIG_COLS = 5
        cols_asig   = ['Nombre del programa', 'Nombre de asignatura',
                       'Keywords identificadas', 'Campos detectados',
                       'Campo y fragmento detectado']
        widths_asig = [38, 42, 28, 26, 60]
        write_title_block(ws2, 'Listado de Asignaturas por Programa', tend_label, N_ASIG_COLS, fecha_str)
        apply_header_row(ws2, 5, cols_asig, widths_asig)

        ws2.merge_cells(start_row=6, start_column=1, end_row=6, end_column=N_ASIG_COLS)
        res2 = ws2.cell(row=6, column=1, value=f'Total de asignaturas únicas: {len(df_asignaturas)}')
        res2.fill = fill(AMARILLO_ACC)
        res2.font = Font(name='Calibri', bold=True, color=AZUL_OSCURO, size=10)
        res2.alignment = Alignment(horizontal='right', vertical='center', indent=1)
        ws2.row_dimensions[6].height = 16

        # Agrupar visualmente por programa: fila de subgrupo cada vez que cambia
        prev_prog = None
        data_start = 7
        for ri, row in enumerate(df_asignaturas.values.tolist()):
            prog = str(row[0]).strip()
            asig = str(row[1]).strip()
            row_idx = data_start + ri
            if prog != prev_prog:
                # Fila separadora de programa (abarca todas las columnas)
                ws2.insert_rows(row_idx)
                ws2.merge_cells(start_row=row_idx, start_column=1,
                                end_row=row_idx, end_column=N_ASIG_COLS)
                gc = ws2.cell(row=row_idx, column=1, value=prog)
                gc.fill = fill(AZUL_MEDIO)
                gc.font = Font(name='Calibri', bold=True, color=BLANCO, size=10)
                gc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                gc.border = borde_fino()
                ws2.row_dimensions[row_idx].height = 20
                data_start += 1
                prev_prog = prog

            # Enriquecer con datos de detalle
            det = lookup_det.get((prog, asig), {})
            kw_str     = '; '.join(sorted(det.get('keywords', [])))
            campos_str = '; '.join(
                CAMPO_LABEL.get(c, c) for c in sorted(det.get('campos', []))
            )
            frag_parts = [
                f"{CAMPO_LABEL.get(c, c)}:\n{t}"
                for c, t in sorted(det.get('textos', {}).items())
            ]
            frag_str = '\n\n'.join(frag_parts)

            row_idx2 = data_start + ri
            bg = GRIS_FILA if ri % 2 == 0 else BLANCO
            extended_row = [prog, asig, kw_str, campos_str, frag_str]
            for ci, value in enumerate(extended_row, start=1):
                cell = ws2.cell(row=row_idx2, column=ci, value=value)
                cell.fill = fill(bg)
                cell.font = Font(name='Calibri', size=10, color='1A1A2E')
                wrap = ci >= 3  # wrap desde Keywords en adelante
                cell.alignment = Alignment(
                    horizontal='left', vertical='top', wrap_text=wrap
                )
                cell.border = borde_fino()
            # Altura proporcional al contenido: ~15pt por cada línea estimada
            n_lineas = max(1, frag_str.count('\n') + 1) if frag_str else 1
            ws2.row_dimensions[row_idx2].height = max(18, min(n_lineas * 15, 400))

        ws2.freeze_panes = 'A7'

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    tab_global, tab_programa = st.tabs([
        "🌍 Vista Global — Todos los Programas",
        "🏫 Análisis por Programa"
    ])

    # ── TAB 1: VISTA GLOBAL ───────────────────────────────────────────────────
    with tab_global:
        st.subheader("Cobertura de Tendencias por Programa")
        st.caption(
            "Cada celda muestra el número de asignaturas únicas que abordan esa tendencia. "
            "Tonos más oscuros indican mayor cobertura."
        )
        # Construir matriz programa × tendencia
        detalle = resultados['detalle']
        programas_ord = sorted({str(p) for tid_d in detalle.values() for p in tid_d})
        tend_labels = {
            tid: (tendencias[tid]['descripcion'][:40] + '…' if len(tendencias[tid]['descripcion']) > 40 else tendencias[tid]['descripcion'])
            for tid in tendencias
        }
        matrix_rows = {}
        for tid, por_prog in detalle.items():
            label = tend_labels.get(tid, tid)
            matrix_rows[label] = {}
            for prog in programas_ord:
                registros = por_prog.get(prog, [])
                matrix_rows[label][str(prog)] = len({r['asignatura'] for r in registros if r['asignatura'] != 'Sin nombre'})

        if matrix_rows:
            df_heat = pd.DataFrame(matrix_rows).T  # filas=tendencias, cols=programas
            # Ordenar tendencias por total descendente
            df_heat['_total'] = df_heat.sum(axis=1)
            df_heat = df_heat.sort_values('_total', ascending=False).drop(columns='_total')

            z = df_heat.values.tolist()
            x_labels = [str(p)[:30] for p in df_heat.columns]
            y_labels = df_heat.index.tolist()

            # Texto en celdas: solo mostrar valores > 0 para no saturar
            text_matrix = [
                [str(int(v)) if v > 0 else '' for v in row]
                for row in z
            ]

            fig_heat = go.Figure(go.Heatmap(
                z=z,
                x=x_labels,
                y=y_labels,
                text=text_matrix,
                texttemplate='%{text}',
                textfont=dict(size=9, color='white'),
                colorscale=[
                    [0.0,  '#F0F4F8'],
                    [0.01, '#C8E6F5'],
                    [0.3,  '#1FB2DE'],
                    [0.7,  '#1565A7'],
                    [1.0,  '#0F385A'],
                ],
                showscale=True,
                colorbar=dict(
                    title=dict(text='Asignaturas', font=dict(size=12)),
                    thickness=14,
                    len=0.8,
                ),
                hoverongaps=False,
                hovertemplate='<b>%{y}</b><br>%{x}<br>%{z} asignatura(s)<extra></extra>',
            ))

            n_progs = len(x_labels)
            n_tends = len(y_labels)
            cell_h = max(28, min(48, 900 // n_tends))
            cell_w = max(18, min(60, 1400 // n_progs))
            fig_height = max(380, cell_h * n_tends + 120)

            fig_heat.update_layout(
                height=fig_height,
                margin=dict(l=10, r=20, t=30, b=120),
                xaxis=dict(
                    tickangle=40,
                    tickfont=dict(size=9),
                    side='bottom',
                    fixedrange=False,
                ),
                yaxis=dict(
                    tickfont=dict(size=10),
                    autorange='reversed',
                ),
                plot_bgcolor='white',
                paper_bgcolor='white',
                font=dict(family='Inter, Arial, sans-serif'),
            )
            st.plotly_chart(fig_heat, use_container_width=True)
        else:
            st.info("No se detectaron tendencias en los datos cargados.")

        st.markdown("---")

        col_a, col_b = st.columns(2)
        with col_a:
            st.subheader("Cobertura por Tendencia (% de asignaturas)")
            st.caption("Porcentaje de asignaturas únicas que abordan cada tendencia.")
            cobertura_data = []
            for tid, pct in resultados['cobertura'].items():
                nombre = tendencias[tid]['descripcion'] if tid in tendencias else tid
                n_asigs = resultados.get('asig_counts', {}).get(tid, 0)
                cobertura_data.append({'Tendencia': nombre[:38], 'Cobertura': round(pct, 1), 'Asignaturas': n_asigs})
            df_cob = pd.DataFrame(cobertura_data).sort_values('Cobertura', ascending=True)
            bar_h_cob = max(380, len(df_cob) * 32 + 80)
            fig_cob = px.bar(
                df_cob, y='Tendencia', x='Cobertura', orientation='h',
                color='Cobertura',
                color_continuous_scale=[[0, '#F8D7DA'], [0.3, '#FBAF17'], [0.6, '#8BC34A'], [1, '#2E7D32']],
                range_color=[0, 100],
                text='Cobertura',
                hover_data=['Asignaturas'],
                labels={'Cobertura': '% Asignaturas', 'Tendencia': ''},
            )
            fig_cob.update_traces(texttemplate='%{text:.1f}%', textposition='outside', textfont_size=10)
            fig_cob.add_vline(x=50, line_dash='dash', line_color='#0F385A', line_width=1.5,
                              annotation_text='50 %', annotation_font_size=10,
                              annotation_position='top right')
            fig_cob.update_layout(
                height=bar_h_cob,
                margin=dict(l=10, r=60, t=20, b=20),
                xaxis=dict(range=[0, 115], showgrid=True, gridcolor='#F0F0F0', ticksuffix='%'),
                yaxis=dict(tickfont=dict(size=10)),
                coloraxis_showscale=False,
                plot_bgcolor='white', paper_bgcolor='white',
                font=dict(family='Inter, Arial, sans-serif'),
            )
            st.plotly_chart(fig_cob, use_container_width=True)

        with col_b:
            st.subheader("N° de Asignaturas por Tendencia")
            st.caption("Número absoluto de asignaturas que mencionan cada tendencia.")
            menciones_data = []
            for tid in tendencias:
                nombre = tendencias[tid]['descripcion']
                n_asigs = resultados.get('asig_counts', {}).get(tid, 0)
                menciones_data.append({'Tendencia': nombre[:38], 'Asignaturas': n_asigs})
            df_menciones = pd.DataFrame(menciones_data).sort_values('Asignaturas', ascending=True)
            bar_h_men = max(380, len(df_menciones) * 32 + 80)
            fig_men = px.bar(
                df_menciones, y='Tendencia', x='Asignaturas', orientation='h',
                color='Asignaturas',
                color_continuous_scale=[[0, '#C8E6F5'], [0.5, '#1FB2DE'], [1, '#0F385A']],
                text='Asignaturas',
                labels={'Asignaturas': 'N° Asignaturas únicas', 'Tendencia': ''},
            )
            fig_men.update_traces(texttemplate='%{text}', textposition='outside', textfont_size=10)
            fig_men.update_layout(
                height=bar_h_men,
                margin=dict(l=10, r=50, t=20, b=20),
                xaxis=dict(showgrid=True, gridcolor='#F0F0F0'),
                yaxis=dict(tickfont=dict(size=10)),
                coloraxis_showscale=False,
                plot_bgcolor='white', paper_bgcolor='white',
                font=dict(family='Inter, Arial, sans-serif'),
            )
            st.plotly_chart(fig_men, use_container_width=True)

        if resultados['ausentes']:
            st.markdown("---")
            st.subheader("⚠️ Brechas Curriculares Detectadas")
            st.warning(
                "Las siguientes tendencias **no fueron detectadas en ninguna asignatura**. "
                "Considere si los programas deberían incorporarlas:"
            )
            for tid in resultados['ausentes']:
                if tid in tendencias:
                    st.markdown(f"- **{tendencias[tid]['descripcion']}**")

        st.markdown("---")
        st.subheader("Detalle por Tendencia (todos los programas)")
        st.caption("Expande cada programa para ver qué asignaturas abordan la tendencia seleccionada.")
        col_sel, col_btn = st.columns([3, 1])
        with col_sel:
            tend_sel = st.selectbox(
                "Seleccionar tendencia:",
                list(tendencias.keys()),
                format_func=lambda x: tendencias[x]['descripcion'],
                key="tend_sel_global"
            )

        if tend_sel in resultados['detalle']:
            n_asigs_tend = resultados.get('asig_counts', {}).get(tend_sel, 0)
            total_asigs = resultados.get('total_asigs', 1)
            st.markdown(
                f"**{tendencias[tend_sel]['descripcion']}** — presente en "
                f"**{n_asigs_tend} de {total_asigs}** asignaturas "
                f"({n_asigs_tend/total_asigs*100:.1f}% de cobertura)"
            )

            # Construir pares exactos (Programa, Asignatura) desde resultados['detalle']
            # — misma fuente que usa el heatmap, garantiza consistencia de conteos
            asig_col_exp = next(
                (c for c in df.columns if 'nombre asignatura' in c.lower()),
                'Nombre asignatura o modulo'
            )
            pares_rows = []
            for prog, hallazgos in resultados['detalle'][tend_sel].items():
                asigs_vistas = set()
                for h in hallazgos:
                    a = str(h.get('asignatura', '')).strip()
                    if a and a not in ('Sin nombre', 'nan'):
                        asigs_vistas.add(a)
                for a in asigs_vistas:
                    pares_rows.append({'Programa': str(prog).strip(), asig_col_exp: a})

            if pares_rows:
                df_pares = pd.DataFrame(pares_rows).drop_duplicates()
                # Normalizar espacios antes del merge para evitar descuadres
                df_norm = df.copy()
                df_norm['Programa'] = df_norm['Programa'].astype(str).str.strip()
                df_norm[asig_col_exp] = df_norm[asig_col_exp].astype(str).str.strip()
                df_filtrado = df_norm.merge(df_pares, on=['Programa', asig_col_exp], how='inner')
            else:
                df_filtrado = df.iloc[0:0]
            tend_nombre = tendencias[tend_sel]['descripcion'][:30]
            with col_btn:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                st.download_button(
                    label="📥 Descargar Excel",
                    data=_build_listados_excel(df_filtrado, tend_nombre, detalle_tend=resultados['detalle'][tend_sel]),
                    file_name=f"listado_{tend_sel}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help=f"Descarga programas y asignaturas que cubren: {tend_nombre}",
                )
            tiene_hallazgos = False
            for programa, hallazgos in resultados['detalle'][tend_sel].items():
                if hallazgos:
                    tiene_hallazgos = True
                    asig_set = set()
                    for h in hallazgos:
                        asig = h.get('asignatura', '')
                        if asig and asig not in ('Sin nombre', 'nan', ''):
                            asig_set.add(asig)
                    with st.expander(f"📚 {programa} — {len(asig_set)} asignatura(s)"):
                        for asig in sorted(asig_set):
                            campos_asig = []
                            for h in hallazgos:
                                if h.get('asignatura') == asig:
                                    campos_asig.extend(h.get('campos', []))
                            campos_txt = ', '.join(dict.fromkeys(campos_asig)) or 'Texto general'
                            st.markdown(f"- **{asig}** _(detectada en: {campos_txt})_")
            if not tiene_hallazgos:
                st.info("Esta tendencia no fue detectada en ningún programa.")

    # ── TAB 2: POR PROGRAMA ───────────────────────────────────────────────────
    with tab_programa:
        st.subheader("Análisis de Tendencias por Programa")
        st.caption(
            "Selecciona un programa para ver en detalle cuántas asignaturas cubren cada tendencia "
            "y cuáles son las asignaturas que las abordan."
        )

        programas_disp = sorted(df['Programa'].unique().tolist())
        prog_sel = st.selectbox("Seleccionar programa:", programas_disp, key="tend_prog_sel")

        df_prog = df[df['Programa'] == prog_sel]
        asig_col_t = 'Nombre asignatura o modulo'
        total_asigs_prog = df_prog[asig_col_t].nunique()

        # Calcular cobertura específica para este programa
        cob_prog = []
        for tid, tinfo in tendencias.items():
            kws = [unicodedata.normalize('NFKD', k.lower()).encode('ascii', 'ignore').decode('ascii')
                   for k in tinfo['keywords']]
            asigs_con_tend = set()
            for _, row in df_prog.iterrows():
                texto = str(row.get('Texto_Completo', '')).lower()
                texto_n = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
                asig = str(row.get(asig_col_t, '')).strip()
                if asig and asig not in ('nan', '') and any(k in texto_n for k in kws):
                    asigs_con_tend.add(asig)
            pct = len(asigs_con_tend) / total_asigs_prog * 100 if total_asigs_prog > 0 else 0
            cob_prog.append({
                'Tendencia': tinfo['descripcion'][:40],
                'ID': tid,
                'Asignaturas': len(asigs_con_tend),
                'Cobertura': round(pct, 1),
                'Lista': sorted(asigs_con_tend)
            })

        df_cob_prog = pd.DataFrame(cob_prog).sort_values('Cobertura', ascending=False)

        # Métricas del programa
        n_con_tend = (df_cob_prog['Asignaturas'] > 0).sum()
        col_p1, col_p2, col_p3 = st.columns(3)
        col_p1.metric("Asignaturas del programa", total_asigs_prog)
        col_p2.metric("Tendencias detectadas", f"{n_con_tend}/{len(tendencias)}")
        col_p3.metric("Cobertura promedio", f"{df_cob_prog['Cobertura'].mean():.1f}%")

        st.markdown("---")

        # Gráfico horizontal de cobertura para el programa
        fig_prog = px.bar(
            df_cob_prog.sort_values('Cobertura', ascending=True),
            y='Tendencia', x='Cobertura', orientation='h',
            color='Cobertura',
            color_continuous_scale=[[0, '#EC0677'], [0.3, '#42F2F2'], [0.6, '#1FB2DE'], [1, '#0F385A']],
            range_color=[0, 100],
            hover_data=['Asignaturas'],
            title=f"Cobertura de tendencias en {prog_sel}",
            labels={'Cobertura': '% Asignaturas', 'Tendencia': ''}
        )
        fig_prog.add_vline(x=50, line_dash="dot", line_color="#FBAF17", opacity=0.8,
                           annotation_text="50%", annotation_position="top right")
        fig_prog.update_layout(height=480, showlegend=False)
        st.plotly_chart(fig_prog, use_container_width=True)

        # Detalle: asignaturas por tendencia del programa
        st.markdown("---")
        st.subheader(f"Asignaturas de **{prog_sel}** por Tendencia")
        tend_sel_prog = st.selectbox(
            "Seleccionar tendencia:",
            df_cob_prog['ID'].tolist(),
            format_func=lambda x: tendencias[x]['descripcion'] if x in tendencias else x,
            key="tend_sel_prog"
        )
        row_sel = df_cob_prog[df_cob_prog['ID'] == tend_sel_prog].iloc[0] if not df_cob_prog.empty else None
        if row_sel is not None:
            asigs_lista = row_sel['Lista']
            if asigs_lista:
                st.markdown(
                    f"**{len(asigs_lista)} asignatura(s)** de {prog_sel} abordan "
                    f"**{tendencias[tend_sel_prog]['descripcion']}** ({row_sel['Cobertura']:.1f}%):"
                )
                for a in asigs_lista:
                    st.markdown(f"- {a}")
            else:
                st.info(f"Ninguna asignatura de **{prog_sel}** aborda esta tendencia.")


def pagina_nlp(df: pd.DataFrame, resultados: Dict):
    """Pagina de mineria de texto."""
    st.title("Mineria de Texto y Analisis Semantico")
    st.markdown("---")
    st.info(
        "**¿Qué es esto?** Aplica técnicas de Procesamiento de Lenguaje Natural (NLP) "
        "sobre los textos de los programas para identificar los conceptos más relevantes, "
        "frases repetidas y qué tan similares son las asignaturas entre sí en cuanto a contenido."
    )

    col_a, col_b = st.columns(2)

    with col_a:
        st.subheader("Términos Clave Globales (TF-IDF)")
        st.caption(
            "TF-IDF mide la importancia de un término: alta frecuencia en el documento "
            "pero baja en todos los demás. Revela los conceptos más distintivos del currículo."
        )
        n_terms = st.slider("Número de términos:", 10, 30, 20, key='slider_tfidf')
        top_items = list(resultados['top_terminos'].items())[:n_terms]
        df_terms = pd.DataFrame(top_items, columns=['Termino', 'Score'])
        df_terms = df_terms.sort_values('Score', ascending=True)

        fig = px.bar(df_terms, y='Termino', x='Score',
                     orientation='h', color='Score',
                     color_continuous_scale='Plasma',
                     labels={'Score': 'Relevancia TF-IDF', 'Termino': 'Término'})
        fig.update_layout(height=max(400, n_terms * 25))
        st.plotly_chart(fig, use_container_width=True)

    with col_b:
        st.subheader("Frases Frecuentes (N-gramas)")
        st.caption(
            "N-gramas son combinaciones de 2 o 3 palabras que aparecen juntas frecuentemente. "
            "Revelan conceptos compuestos y enfoques pedagógicos recurrentes."
        )
        n_ng = st.slider("Número de frases:", 5, 20, 15, key='slider_ngrams')
        ng_items = list(resultados['top_ngrams'].items())[:n_ng]
        df_ng = pd.DataFrame(ng_items, columns=['N-grama', 'Frecuencia'])
        df_ng = df_ng.sort_values('Frecuencia', ascending=True)

        fig = px.bar(df_ng, y='N-grama', x='Frecuencia',
                     orientation='h', color='Frecuencia',
                     color_continuous_scale='Magma',
                     labels={'Frecuencia': 'Veces que aparece', 'N-grama': 'Frase'})
        fig.update_layout(height=max(400, n_ng * 25))
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    st.subheader("Términos Clave por Programa")
    st.caption("Compara qué conceptos son más relevantes y distintivos en cada programa académico.")
    if resultados['top_por_programa']:
        tabs = st.tabs(list(resultados['top_por_programa'].keys()))
        for tab, (programa, terminos) in zip(tabs, resultados['top_por_programa'].items()):
            with tab:
                items = list(terminos.items())[:15]
                df_p = pd.DataFrame(items, columns=['Termino', 'Score'])
                df_p = df_p.sort_values('Score', ascending=True)
                fig = px.bar(df_p, y='Termino', x='Score',
                             orientation='h', color='Score',
                             color_continuous_scale='Teal',
                             title=f"Términos más relevantes — {programa}")
                fig.update_layout(height=450)
                st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ── Similitud entre Asignaturas (interactivo) ──────────────────────────
    st.subheader("Similitud entre Asignaturas")
    st.info(
        "**¿Para qué sirve esto?** Detecta asignaturas con contenidos muy similares entre sí, "
        "lo que puede indicar redundancia o falta de diferenciación curricular. "
        "Una similitud superior al 80% merece revisión. "
        "Puede comparar 2 o más programas a la vez."
    )

    programas_disponibles = sorted(df['Programa'].unique().tolist())
    col_prog_sim, col_top_sim = st.columns([3, 1])
    with col_prog_sim:
        programas_sel = st.multiselect(
            "Seleccionar programas a comparar (2 o más):",
            programas_disponibles,
            default=programas_disponibles,
            help="Elige qué programas incluir en el análisis de similitud"
        )
    with col_top_sim:
        max_asigs_heat = st.number_input(
            "Máx. asignaturas en heatmap:",
            min_value=5, max_value=80, value=40, step=5,
            help="Limita el heatmap a las N asignaturas con mayor similitud promedio"
        )

    if len(programas_sel) < 2:
        st.warning("Selecciona al menos 2 programas para el análisis de similitud.")
    else:
        df_sim_filtrado = df[df['Programa'].isin(programas_sel)]
        with st.spinner("Calculando similitud..."):
            res_sim = calcular_similitud_asignaturas(df_sim_filtrado)

        if res_sim['n_asignaturas'] < 3:
            st.warning("Se necesitan al menos 3 asignaturas en los programas seleccionados.")
        else:
            if res_sim['par_similar']:
                par = res_sim['par_similar']
                st.info(
                    f"**Par más similar:** {par['asig1']} ↔ {par['asig2']} "
                    f"(Similitud: {par['similitud']:.1%})"
                )

            sim_df = res_sim['similitud_df']
            if not sim_df.empty:
                # Reducir a top-N asignaturas por similitud promedio (excl. diagonal)
                if len(sim_df) > max_asigs_heat:
                    sim_vals_full = sim_df.values.copy()
                    np.fill_diagonal(sim_vals_full, 0)
                    avg_sim = sim_vals_full.mean(axis=1)
                    top_idx = np.argsort(avg_sim)[::-1][:int(max_asigs_heat)]
                    top_names = [sim_df.index[i] for i in sorted(top_idx)]
                    sim_df_heat = sim_df.loc[top_names, top_names]
                    st.caption(
                        f"📊 Mostrando las **{int(max_asigs_heat)} asignaturas** con mayor similitud promedio "
                        f"(de {res_sim['n_asignaturas']} totales). Ajusta el límite en el campo superior."
                    )
                else:
                    sim_df_heat = sim_df

                tab_grupos, tab_pares = st.tabs(["🔍 Grupos Temáticos", "📋 Pares Más Similares"])

                with tab_grupos:
                    st.caption(
                        "Las asignaturas se agrupan automáticamente según similitud de contenido. "
                        "Grupos con alta cohesión comparten vocabulario y temáticas similares."
                    )
                    n_max = min(10, max(2, len(sim_df) // 2))
                    n_clusters = st.slider(
                        "Número de grupos:", min_value=2, max_value=n_max,
                        value=min(5, n_max), key="n_clusters_nlp"
                    )

                    # Clustering aglomerativo sobre matriz de distancias
                    dist_matrix = 1.0 - sim_df.values.clip(0, 1)
                    np.fill_diagonal(dist_matrix, 0.0)
                    clustering = AgglomerativeClustering(
                        n_clusters=n_clusters, metric='precomputed', linkage='average'
                    )
                    labels = clustering.fit_predict(dist_matrix)

                    PALETA = ['#0F385A', '#1FB2DE', '#42F2F2', '#FBAF17', '#EC0677',
                              '#2E7D32', '#6A1B9A', '#E65100', '#00695C', '#AD1457']

                    for cid in range(n_clusters):
                        idx = [i for i, lb in enumerate(labels) if lb == cid]
                        subjects = sorted([sim_df.index[i] for i in idx])
                        if not subjects:
                            continue

                        # Cohesión intra-grupo (similitud promedio)
                        if len(idx) > 1:
                            sub_sim = sim_df.values[np.ix_(idx, idx)].copy()
                            np.fill_diagonal(sub_sim, 0)
                            cohesion = sub_sim.sum() / (len(idx) * (len(idx) - 1))
                        else:
                            cohesion = 1.0

                        color = PALETA[cid % len(PALETA)]
                        label = f"Grupo {cid + 1} &nbsp;·&nbsp; {len(subjects)} asignaturas &nbsp;·&nbsp; Cohesión: {cohesion:.0%}"

                        with st.expander(f"Grupo {cid + 1} — {len(subjects)} asignaturas (cohesión {cohesion:.0%})"):
                            cols = st.columns(2)
                            for k, subj in enumerate(subjects):
                                cols[k % 2].markdown(
                                    f"<div style='padding:4px 8px;margin:2px 0;border-radius:4px;"
                                    f"border-left:3px solid {color};background:#F5FDFF;font-size:0.88em'>"
                                    f"{subj}</div>",
                                    unsafe_allow_html=True
                                )

                    st.caption(
                        "💡 Cohesión alta = asignaturas muy similares entre sí (posible redundancia). "
                        "Ajusta el número de grupos para obtener agrupaciones más o menos detalladas."
                    )

                with tab_pares:
                    sim_vals = sim_df.values.copy()
                    np.fill_diagonal(sim_vals, 0)
                    pares = []
                    for i in range(len(sim_vals)):
                        for j in range(i + 1, len(sim_vals)):
                            pares.append({
                                'Asignatura 1': str(sim_df.index[i]),
                                'Asignatura 2': str(sim_df.columns[j]),
                                'Similitud': round(float(sim_vals[i, j]), 4)
                            })
                    df_pares = (
                        pd.DataFrame(pares)
                        .sort_values('Similitud', ascending=False)
                        .head(40)
                    )
                    def highlight_similar(val):
                        if isinstance(val, float):
                            if val >= 0.8:
                                return 'background-color:#FEE5F2;color:#EC0677;font-weight:bold'
                            if val >= 0.6:
                                return 'background-color:#FFFBEA;color:#8C6000'
                        return ''
                    st.dataframe(
                        df_pares.style.map(highlight_similar, subset=['Similitud']),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.caption("🔴 ≥ 0.80 = posible redundancia · 🟡 0.60–0.79 = revisar diferenciación")

    st.markdown("---")
    st.subheader("Buscar Términos en los Datos")
    st.caption("Escribe un concepto o frase para encontrar en qué asignaturas y programas aparece.")
    termino_buscar = st.text_input("Término o frase a buscar:")
    if termino_buscar:
        mask = df['Texto_Completo'].str.contains(termino_buscar.lower(), na=False)
        res_busq = df[mask][['Programa', 'Nombre asignatura o modulo',
                             'Tipo de Saber', 'Semestre']].drop_duplicates()
        st.markdown(f"**{len(res_busq)} registros encontrados** para «{termino_buscar}»")
        st.dataframe(res_busq, use_container_width=True, hide_index=True)


def pagina_tipo_saber(df: pd.DataFrame):
    """Pagina de analisis profundo del Tipo de Saber."""
    # Garantizar que solo entran los tres tipos válidos
    df = df[df['Tipo de Saber'].isin({'Saber', 'SaberHacer', 'SaberSer'})].copy()

    st.title("Análisis de Tipo de Saber")
    st.markdown("---")
    st.info(
        "El **Tipo de Saber** clasifica cada estrategia microcurricular según la dimensión "
        "del aprendizaje que desarrolla:\n\n"
        "- 🔵 **Saber** (conocimiento): comprensión conceptual, teórica y declarativa. "
        "Ej: «conoce los fundamentos de…», «identifica los principios de…»\n"
        "- 🟠 **SaberHacer** (habilidad): competencias procedimentales y prácticas. "
        "Ej: «aplica técnicas de…», «elabora un plan de…», «resuelve problemas de…»\n"
        "- 🟣 **SaberSer** (actitud): valores, ética, ciudadanía y dimensión humana. "
        "Ej: «valora la importancia de…», «actúa con responsabilidad en…»\n\n"
        "**Referencia:** Un currículo por competencias equilibrado debería tener énfasis en "
        "SaberHacer (≥40%), complementado con Saber (30-40%) y SaberSer (≥15%)."
    )

    COLORES_SABER = {
        'Saber': '#092196',
        'SaberHacer': '#748BFC',
        'SaberSer': '#A8B7FF'
    }
    REFS_SABER = {
        'Saber':     (25, 45),
        'SaberHacer': (35, 60),
        'SaberSer':  (10, 30)
    }

    # ── Distribución global + Diagnóstico ───────────────────────────────────
    st.markdown("---")
    st.subheader("Distribución Global de Tipo de Saber")

    totales = df['Tipo de Saber'].value_counts()
    total = totales.sum()

    col_donut, col_diag = st.columns([1, 1])
    with col_donut:
        total_tipo = df['Tipo de Saber'].value_counts().reset_index()
        total_tipo.columns = ['Tipo', 'Registros']
        fig_donut = px.pie(
            total_tipo, values='Registros', names='Tipo',
            color='Tipo', color_discrete_map=COLORES_SABER,
            hole=0.45,
            title='Composición global (todos los programas)'
        )
        fig_donut.update_traces(
            texttemplate='%{label}<br><b>%{percent:.1%}</b>',
            textfont_size=12
        )
        fig_donut.update_layout(height=380, showlegend=False,
                                margin=dict(t=50, b=10, l=10, r=10))
        st.plotly_chart(fig_donut, use_container_width=True)

    with col_diag:
        st.markdown("**Diagnóstico de Equilibrio**")
        st.caption("Referencia recomendada para currículos por competencias")
        for tipo, (ref_min, ref_max) in REFS_SABER.items():
            n = int(totales.get(tipo, 0))
            pct = n / total * 100 if total > 0 else 0
            if ref_min <= pct <= ref_max:
                icono, color = "✅", "#27AE60"
            elif pct > ref_max:
                icono, color = "⬆️", "#FBAF17"
            else:
                icono, color = "⬇️", "#EC0677"
            bar_pct = int(pct * 2)
            st.markdown(
                f"<div style='margin:10px 0;padding:10px 14px;border-radius:8px;"
                f"border-left:4px solid {COLORES_SABER[tipo]};background:#F5FDFF'>"
                f"<b style='color:{COLORES_SABER[tipo]}'>{tipo}</b> "
                f"<span style='float:right;color:{color};font-weight:700'>{icono} {pct:.1f}%</span><br>"
                f"<div style='background:#E0F0F8;border-radius:4px;height:8px;margin:6px 0'>"
                f"<div style='background:{COLORES_SABER[tipo]};width:{min(bar_pct,100)}%;height:8px;border-radius:4px'></div></div>"
                f"<small style='color:#666'>Referencia: {ref_min}–{ref_max}% &nbsp;|&nbsp; {n} registros</small>"
                f"</div>",
                unsafe_allow_html=True
            )

    # ── Radar + Scatter de balance por programa ─────────────────────────────
    st.markdown("---")
    st.subheader("Perfil de Competencias por Programa")
    st.caption(
        "**Radar:** compara simultáneamente los 3 tipos de saber entre programas — "
        "el área ideal es amplia y equilibrada. "
        "**Mapa de balance:** posición óptima = SaberHacer alto (eje X) y SaberSer alto (eje Y)."
    )

    pivot = (
        df.groupby(['Programa', 'Tipo de Saber'])
        .size()
        .reset_index(name='Registros')
    )
    total_prog = pivot.groupby('Programa')['Registros'].transform('sum')
    pivot['Porcentaje'] = (pivot['Registros'] / total_prog * 100).round(1)
    pivot_wide = pivot.pivot_table(index='Programa', columns='Tipo de Saber',
                                   values='Porcentaje', fill_value=0).reset_index()

    tab_radar, tab_barra = st.tabs([
        "🕸️ Radar de Competencias",
        "📊 Comparativo Detallado"
    ])

    with tab_radar:
        tipos_disp = [t for t in ['Saber', 'SaberHacer', 'SaberSer'] if t in pivot_wide.columns]
        fig_radar = go.Figure()
        palette_radar = PALETA_DOUBLE_SPLIT
        for i, row in pivot_wide.iterrows():
            vals = [row.get(t, 0) for t in tipos_disp]
            vals_closed = vals + [vals[0]]
            cats_closed = tipos_disp + [tipos_disp[0]]
            fig_radar.add_trace(go.Scatterpolar(
                r=vals_closed,
                theta=cats_closed,
                fill='toself',
                fillcolor=f"rgba({int(palette_radar[i % len(palette_radar)].lstrip('#')[0:2], 16)},"
                          f"{int(palette_radar[i % len(palette_radar)].lstrip('#')[2:4], 16)},"
                          f"{int(palette_radar[i % len(palette_radar)].lstrip('#')[4:6], 16)},0.15)",
                line=dict(color=palette_radar[i % len(palette_radar)], width=2),
                name=str(row['Programa'])
            ))
        # Zona de referencia
        ref_vals = [35, 47.5, 20]
        fig_radar.add_trace(go.Scatterpolar(
            r=ref_vals + [ref_vals[0]],
            theta=tipos_disp + [tipos_disp[0]],
            fill='toself',
            fillcolor='rgba(251,175,23,0.07)',
            line=dict(color='#FBAF17', width=1, dash='dot'),
            name='Zona referencia'
        ))
        fig_radar.update_layout(
            polar=dict(radialaxis=dict(visible=True, range=[0, 100],
                                       tickfont=dict(size=10), ticksuffix='%')),
            height=450,
            showlegend=True,
            legend=dict(orientation='h', y=-0.15),
            title="Radar de tipos de saber por programa (zona dorada = rango ideal)"
        )
        st.plotly_chart(fig_radar, use_container_width=True)
        st.caption("La zona dorada punteada representa el rango de referencia ideal.")

    with tab_barra:
        fig_bar = px.bar(
            pivot.sort_values('Tipo de Saber'),
            x='Porcentaje', y='Programa', color='Tipo de Saber',
            barmode='group', orientation='h',
            color_discrete_map=COLORES_SABER,
            text='Porcentaje',
            title='Comparativo detallado por programa y tipo de saber',
            labels={'Porcentaje': '%', 'Tipo de Saber': 'Tipo'}
        )
        fig_bar.update_traces(texttemplate='%{text:.0f}%', textposition='outside')
        for tipo, (ref_min, ref_max) in REFS_SABER.items():
            fig_bar.add_vline(x=ref_min, line_dash='dot',
                              line_color=COLORES_SABER.get(tipo, 'gray'), opacity=0.4)
        fig_bar.update_layout(height=max(300, len(pivot_wide) * 80),
                              xaxis=dict(range=[0, 100], title='% de registros'),
                              yaxis_title='')
        st.plotly_chart(fig_bar, use_container_width=True)
        st.caption("Las líneas punteadas verticales marcan el límite mínimo de referencia para cada tipo.")

    # ── Por semestre ────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Evolución de Tipo de Saber por Semestre")
    st.caption(
        "¿Cómo cambia el enfoque del aprendizaje a lo largo de la carrera? "
        "Se espera que los semestres avanzados tengan mayor proporción de SaberHacer y SaberSer."
    )

    programas_disponibles = sorted(df['Programa'].unique().tolist())
    prog_sel_saber = st.multiselect(
        "Seleccionar programas:",
        programas_disponibles,
        default=programas_disponibles,
        key='saber_prog_sel'
    )
    df_sem = df[df['Programa'].isin(prog_sel_saber)] if prog_sel_saber else df

    # Filtrar semestres numéricos
    df_sem = df_sem.copy()
    df_sem['Semestre_num'] = pd.to_numeric(df_sem['Semestre'], errors='coerce')
    df_sem_valid = df_sem.dropna(subset=['Semestre_num'])

    if not df_sem_valid.empty:
        pivot_sem = (
            df_sem_valid.groupby(['Semestre_num', 'Tipo de Saber'])
            .size()
            .reset_index(name='Registros')
        )
        total_sem = pivot_sem.groupby('Semestre_num')['Registros'].transform('sum')
        pivot_sem['Porcentaje'] = (pivot_sem['Registros'] / total_sem * 100).round(1)
        pivot_sem['Semestre'] = pivot_sem['Semestre_num'].astype(int).astype(str)

        fig = px.line(
            pivot_sem.sort_values('Semestre_num'),
            x='Semestre', y='Porcentaje', color='Tipo de Saber',
            color_discrete_map=COLORES_SABER,
            markers=True,
            title='Evolución del Tipo de Saber por semestre',
            labels={'Porcentaje': '%', 'Semestre': 'Semestre'}
        )
        fig.update_layout(height=400, yaxis_range=[0, 100])
        st.plotly_chart(fig, use_container_width=True)
        st.caption(
            "💡 Si SaberSer es muy bajo en todos los semestres, el programa puede estar "
            "descuidando la formación en valores y ciudadanía."
        )
    else:
        st.info("No hay datos de semestres numéricos para los programas seleccionados.")

    # ── Por asignatura ──────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("Tipo de Saber Predominante por Asignatura")
    st.caption(
        "Cada asignatura se clasifica según el tipo de saber más frecuente en sus registros. "
        "Útil para identificar si una asignatura es principalmente teórica, práctica o formativa."
    )

    prog_asig_sel = st.selectbox(
        "Seleccionar programa:",
        sorted(df['Programa'].unique()),
        key='saber_asig_prog'
    )
    df_asig_saber = df[df['Programa'] == prog_asig_sel]

    pivot_asig = (
        df_asig_saber.groupby(['Nombre asignatura o modulo', 'Tipo de Saber'])
        .size()
        .reset_index(name='Registros')
    )
    total_asig = pivot_asig.groupby('Nombre asignatura o modulo')['Registros'].transform('sum')
    pivot_asig['Porcentaje'] = (pivot_asig['Registros'] / total_asig * 100).round(1)

    # Ordenar por SaberHacer descendente para identificar asignaturas más prácticas
    orden = (
        pivot_asig[pivot_asig['Tipo de Saber'] == 'SaberHacer']
        .sort_values('Porcentaje', ascending=False)['Nombre asignatura o modulo']
        .tolist()
    )
    todas = pivot_asig['Nombre asignatura o modulo'].unique().tolist()
    orden_completo = orden + [a for a in todas if a not in orden]

    # ── Tabla semáforo por asignatura ───────────────────────────────────────
    tabla_wide = pivot_asig.pivot_table(
        index='Nombre asignatura o modulo',
        columns='Tipo de Saber',
        values='Porcentaje',
        fill_value=0
    ).round(1).reset_index()
    tabla_wide.columns.name = None

    # Asegurar que existen las columnas aunque no haya datos
    for col in ['Saber', 'SaberHacer', 'SaberSer']:
        if col not in tabla_wide.columns:
            tabla_wide[col] = 0.0

    # Ordenar igual que antes (por SaberHacer desc)
    tabla_wide = tabla_wide.set_index('Nombre asignatura o modulo').loc[
        [a for a in orden_completo if a in tabla_wide['Nombre asignatura o modulo'].values]
    ].reset_index()

    st.caption("Distribución porcentual de cada tipo de saber por asignatura, ordenadas de mayor a menor SaberHacer.")

    cols_mostrar = ['Nombre asignatura o modulo', 'Saber', 'SaberHacer', 'SaberSer']
    tabla_display = tabla_wide[[c for c in cols_mostrar if c in tabla_wide.columns]].rename(
        columns={'Nombre asignatura o modulo': 'Asignatura'}
    )
    st.dataframe(
        tabla_display.style.format(
            {c: '{:.1f}%' for c in ['Saber', 'SaberHacer', 'SaberSer'] if c in tabla_display.columns}
        ),
        use_container_width=True,
        hide_index=True,
        height=min(600, len(tabla_display) * 35 + 60)
    )


def pagina_resumen_ejecutivo(df: pd.DataFrame, tendencias: Dict) -> None:
    """Pagina de resumen ejecutivo con hallazgos y recomendaciones para toma de decisiones."""
    st.title("📋 Resumen Ejecutivo")
    st.markdown("---")
    st.info(
        "Esta sección sintetiza los hallazgos más relevantes del análisis microcurricular "
        "para facilitar la **toma de decisiones** en comités académicos, procesos de "
        "autoevaluación y rediseños curriculares. Incluye alertas automáticas, fortalezas "
        "detectadas y recomendaciones priorizadas."
    )

    programas = df['Programa'].unique().tolist()
    asig_col = 'Nombre asignatura o modulo'
    total_asigs = df[asig_col].nunique()
    total_registros = len(df)

    # ── Métricas generales ──────────────────────────────────────────────────
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Programas analizados", len(programas))
    col2.metric("Asignaturas únicas", total_asigs)
    col3.metric("Registros procesados", f"{total_registros:,}")
    col4.metric("Semestres detectados",
                df['Semestre'].dropna().nunique())

    st.markdown("---")

    # ── Sección 1: Alertas priorizadas ─────────────────────────────────────
    st.subheader("🚦 Alertas y Brechas Detectadas")
    st.caption(
        "Situaciones que requieren atención según criterios de calidad académica. "
        "Ordenadas por prioridad."
    )

    alertas = []

    # 1a. Tipo de Saber — desequilibrios
    totales_saber = df['Tipo de Saber'].value_counts()
    total_ts = totales_saber.sum()
    for tipo, (ref_min, ref_max) in {
        'SaberHacer': (35, 65),
        'SaberSer': (10, 30),
        'Saber': (20, 50)
    }.items():
        pct = totales_saber.get(tipo, 0) / total_ts * 100 if total_ts > 0 else 0
        if pct < ref_min:
            alertas.append({
                'Prioridad': 'Alta' if tipo in ('SaberHacer', 'SaberSer') else 'Media',
                'Categoría': 'Tipo de Saber',
                'Hallazgo': f'**{tipo}** es bajo ({pct:.1f}%) — referencia: {ref_min}–{ref_max}%',
                'Recomendación': (
                    'Revisar asignaturas con 0% de SaberHacer e incorporar actividades prácticas.'
                    if tipo == 'SaberHacer' else
                    'Incorporar resultados de aprendizaje orientados a valores, ética y ciudadanía.'
                    if tipo == 'SaberSer' else
                    'Verificar si el currículo tiene suficiente fundamentación teórica.'
                )
            })
        elif pct > ref_max:
            alertas.append({
                'Prioridad': 'Media',
                'Categoría': 'Tipo de Saber',
                'Hallazgo': f'**{tipo}** es alto ({pct:.1f}%) — referencia: {ref_min}–{ref_max}%',
                'Recomendación': f'Considerar rebalancear hacia los otros tipos de saber.'
            })

    # 1b. Semestres sin datos
    df_sem_num = df.copy()
    df_sem_num['Semestre_num'] = pd.to_numeric(df_sem_num['Semestre'], errors='coerce')
    pct_sin_semestre = df_sem_num['Semestre_num'].isna().mean() * 100
    if pct_sin_semestre > 20:
        alertas.append({
            'Prioridad': 'Media',
            'Categoría': 'Completitud de datos',
            'Hallazgo': f'{pct_sin_semestre:.0f}% de registros sin semestre numérico válido',
            'Recomendación': 'Estandarizar el campo Semestre en los archivos Excel (usar números: 1, 2, 3…)'
        })

    # 1c. Tendencias globales ausentes
    resultados_tend = analizar_tendencias(df, tendencias)
    if resultados_tend['ausentes']:
        desc_ausentes = [tendencias[t]['descripcion'] for t in resultados_tend['ausentes'] if t in tendencias]
        alertas.append({
            'Prioridad': 'Alta',
            'Categoría': 'Tendencias Globales',
            'Hallazgo': f'{len(desc_ausentes)} tendencia(s) sin cobertura: {", ".join(desc_ausentes[:3])}{"…" if len(desc_ausentes) > 3 else ""}',
            'Recomendación': 'Revisar si estas tendencias son relevantes para el perfil del egresado y actualizar los contenidos.'
        })

    # 1d. Tendencias con muy baja cobertura (<20%)
    for tid, pct in resultados_tend['cobertura'].items():
        if 0 < pct < 20 and tid in tendencias:
            alertas.append({
                'Prioridad': 'Baja',
                'Categoría': 'Tendencias Globales',
                'Hallazgo': f'Baja cobertura de **{tendencias[tid]["descripcion"]}** ({pct:.1f}%)',
                'Recomendación': 'Solo unas pocas asignaturas abordan esta tendencia. Considerar ampliar su presencia.'
            })

    # 1e. Asignaturas sin núcleos temáticos
    sin_nucleos = df[df['Nucleos tematicos'].isna() | (df['Nucleos tematicos'].astype(str).str.strip() == '') |
                     (df['Nucleos tematicos'].astype(str).str.strip() == 'nan')]
    asigs_sin_nucleos = sin_nucleos[asig_col].nunique() if not sin_nucleos.empty else 0
    if asigs_sin_nucleos > 0:
        pct_sin = asigs_sin_nucleos / total_asigs * 100
        alertas.append({
            'Prioridad': 'Media',
            'Categoría': 'Completitud de datos',
            'Hallazgo': f'{asigs_sin_nucleos} asignaturas ({pct_sin:.0f}%) sin núcleos temáticos declarados',
            'Recomendación': 'Completar la columna «Núcleos temáticos» en el formato Excel para un análisis más preciso.'
        })

    if alertas:
        orden_prioridad = {'Alta': 0, 'Media': 1, 'Baja': 2}
        df_alertas = pd.DataFrame(alertas).sort_values(
            'Prioridad', key=lambda x: x.map(orden_prioridad)
        )
        for _, a in df_alertas.iterrows():
            icono = '🔴' if a['Prioridad'] == 'Alta' else ('🟡' if a['Prioridad'] == 'Media' else '🟢')
            with st.expander(f"{icono} [{a['Prioridad']}] {a['Categoría']}: {a['Hallazgo']}", expanded=(a['Prioridad'] == 'Alta')):
                st.markdown(f"**💡 Recomendación:** {a['Recomendación']}")
    else:
        st.success("✅ No se detectaron alertas críticas en los programas analizados.")

    # ── Sección 2: Fortalezas detectadas ───────────────────────────────────
    st.markdown("---")
    st.subheader("✅ Fortalezas Detectadas")
    st.caption("Aspectos positivos identificados en el análisis.")

    fortalezas = []
    tend_detectadas = len(tendencias) - len(resultados_tend['ausentes'])
    if tend_detectadas >= len(tendencias) * 0.8:
        fortalezas.append(f"Alta alineación con tendencias globales: **{tend_detectadas}/{len(tendencias)}** tendencias presentes")
    saberhacer_pct = totales_saber.get('SaberHacer', 0) / total_ts * 100 if total_ts > 0 else 0
    if saberhacer_pct >= 35:
        fortalezas.append(f"Buen énfasis en **SaberHacer** ({saberhacer_pct:.1f}%) — favorece la formación por competencias")

    # Diversidad temática
    nucleos_list = []
    for _, row in df.iterrows():
        raw = str(row.get('Nucleos tematicos', ''))
        if raw and raw not in ('nan', ''):
            nucleos_list.extend([limpiar_nucleo(n.strip()) for n in _split_nucleos(raw) if es_nucleo_valido(n.strip())[0]])
    n_unicos = len(set(nucleos_list))
    if n_unicos > total_asigs * 2:
        fortalezas.append(f"Alta diversidad temática: **{n_unicos}** núcleos únicos para {total_asigs} asignaturas")

    if asigs_sin_nucleos == 0:
        fortalezas.append("**100%** de asignaturas con núcleos temáticos declarados")

    if fortalezas:
        for f_item in fortalezas:
            st.markdown(f"- {f_item}")
    else:
        st.info("No se identificaron fortalezas destacadas con los umbrales actuales.")

    # ── Sección 3: Resumen por programa ────────────────────────────────────
    st.markdown("---")
    st.subheader("📊 Perfil por Programa")
    st.caption("Resumen de los indicadores clave por programa para soporte a decisiones curriculares.")

    for prog in sorted(programas):
        df_prog = df[df['Programa'] == prog]
        n_asigs = df_prog[asig_col].nunique()
        n_reg = len(df_prog)
        ts = df_prog['Tipo de Saber'].value_counts()
        ts_total = ts.sum()
        sh_pct = ts.get('SaberHacer', 0) / ts_total * 100 if ts_total > 0 else 0
        ss_pct = ts.get('SaberSer', 0) / ts_total * 100 if ts_total > 0 else 0
        tend_prog = sum(1 for tid in tendencias if resultados_tend['asig_counts'].get(tid, 0) > 0)

        with st.expander(f"📚 {prog} — {n_asigs} asignaturas, {n_reg} registros"):
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Asignaturas", n_asigs)
            c2.metric("SaberHacer", f"{sh_pct:.1f}%",
                      delta="✅ OK" if sh_pct >= 35 else "⚠️ Bajo",
                      delta_color="off")
            c3.metric("SaberSer", f"{ss_pct:.1f}%",
                      delta="✅ OK" if ss_pct >= 10 else "⚠️ Bajo",
                      delta_color="off")
            c4.metric("Tendencias cubiertas", f"{tend_prog}/{len(tendencias)}")

    # ── Sección 4: Descarga ─────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("⬇️ Descargar Datos del Análisis")
    st.caption("Exporta los datos filtrados o el resumen de alertas para presentaciones y reportes.")

    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        csv_datos = df[[
            'Programa', 'Nombre asignatura o modulo', 'Semestre',
            'Tipo de Saber', 'Resultado de aprendizaje', 'Nucleos tematicos'
        ]].to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📄 Descargar datos consolidados (CSV)",
            data=csv_datos,
            file_name="datos_microcurriculares.csv",
            mime="text/csv",
            help="Descarga todos los registros procesados en formato Excel-compatible"
        )
    with col_dl2:
        if alertas:
            csv_alertas = pd.DataFrame(alertas).to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="🚨 Descargar alertas (CSV)",
                data=csv_alertas,
                file_name="alertas_curriculares.csv",
                mime="text/csv",
                help="Descarga el listado de alertas y recomendaciones detectadas"
            )


def pagina_bloom_integracion(df: pd.DataFrame, taxonomias_externas: Dict | None = None):
    """Taxonomía por dominios/subcategorías de la BD y Mapa de Integración."""
    import math as _math

    st.title("🎓 Taxonomías Curriculares e Integración")
    st.markdown("---")
    st.info(
        "Clasifica los Resultados de Aprendizaje según los **dominios** (Cognitivo, Procedimental, Actitudinal) "
        "y **subcategorías** definidas en la base de datos taxonómica (646 verbos). "
        "Explore la distribución, la progresión por semestre y la articulación temática entre asignaturas."
    )

    tab_dom, tab_sub, tab_prog, tab_ra, tab_mapa = st.tabs([
        "🎯 Por Dominio",
        "📊 Por Subcategoría",
        "📈 Progresión",
        "🔍 Explorar RAs",
        "🕸️ Mapa de Integración",
    ])

    # ── Construcción del lookup de verbos desde la BD taxonómica ─────────────
    # verb_to_tax: {verbo_norm: (dominio, subcategoria_display, orden)}
    verb_to_tax: Dict[str, tuple] = {}
    subcat_display_map: Dict[str, str] = {}   # norm → display name

    if taxonomias_externas:
        for subcat_norm, verbos_bd in taxonomias_externas.items():
            dominio  = _SUBCAT_TO_DOMAIN.get(subcat_norm, 'Cognitivo')
            orden    = _SUBCAT_ORDER.get(subcat_norm, 0)
            # Display name: title-case sin acentos ASCII → conservar original
            subcat_display = subcat_norm.replace('comprension', 'Comprensión')\
                                        .replace('aplicacion',  'Aplicación')\
                                        .replace('analisis',    'Análisis')\
                                        .replace('sintesis',    'Síntesis')\
                                        .replace('evaluacion',  'Evaluación')\
                                        .replace('creacion',    'Creación')\
                                        .replace('imitacion',   'Imitación')\
                                        .replace('manipulacion','Manipulación')\
                                        .replace('precision',   'Precisión')\
                                        .replace('percepcion',  'Percepción')\
                                        .title()
            subcat_display_map[subcat_norm] = subcat_display
            for v in verbos_bd:
                v_n = unicodedata.normalize('NFKD', v).encode('ascii', 'ignore').decode('ascii')
                if v_n not in verb_to_tax or orden > verb_to_tax[v_n][2]:
                    verb_to_tax[v_n] = (dominio, subcat_display, orden)
        st.success(
            f"✅ BD cargada: **{len(verb_to_tax)}** verbos únicos · "
            f"**{len(subcat_display_map)}** subcategorías · "
            f"**3** dominios (Cognitivo / Procedimental / Actitudinal)"
        )
    else:
        st.warning(
            "⚠️ No se encontró **Taxonomias_MatrizBD.xlsx** en `data/raw/`. "
            "La clasificación usará solo los verbos base del diccionario interno."
        )

    def detectar_taxonomia(texto: str):
        """Retorna (dominio, subcategoria) usando la BD taxonómica con stem matching."""
        if not texto or str(texto).strip() in ('nan', ''):
            return ('No identificado', 'No identificado')
        t = unicodedata.normalize('NFKD', str(texto).lower()).encode('ascii', 'ignore').decode('ascii')
        words_t = re.findall(r'\b[a-z]{3,}\b', t)
        text_stems = [(w, _stem_es(w)) for w in words_t]
        best_orden = -1
        best_dom   = 'No identificado'
        best_sub   = 'No identificado'
        for v_n, (dom, sub, orden) in verb_to_tax.items():
            if orden <= best_orden:
                continue
            if re.search(r'\b' + re.escape(v_n) + r'\b', t):
                best_orden, best_dom, best_sub = orden, dom, sub
                continue
            v_stem = _stem_es(v_n)
            if len(v_stem) < 4:
                continue
            for word, word_stem in text_stems:
                if word_stem == v_stem or word.startswith(v_stem) or v_n.startswith(word_stem):
                    best_orden, best_dom, best_sub = orden, dom, sub
                    break
        return (best_dom, best_sub)

    df_tax = df.copy()
    df_tax[['Dominio', 'Subcategoria']] = df_tax['Resultado de aprendizaje'].apply(
        lambda x: pd.Series(detectar_taxonomia(x))
    )

    dominios_todos   = ['Cognitivo', 'Procedimental', 'Actitudinal', 'No identificado']
    programas_todos  = sorted(df_tax['Programa'].unique().tolist())

    # ── Filtros globales (fuera de los tabs) ──────────────────────────────────
    with st.expander("🔧 Filtros", expanded=False):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            dom_filter = st.multiselect(
                "Dominio:", dominios_todos, default=[], key="tax_dom_filter"
            )
        with col_f2:
            prog_filter = st.multiselect(
                "Programa:", programas_todos, default=[], key="tax_prog_filter"
            )

    df_filt = df_tax.copy()
    if dom_filter:
        df_filt = df_filt[df_filt['Dominio'].isin(dom_filter)]
    if prog_filter:
        df_filt = df_filt[df_filt['Programa'].isin(prog_filter)]

    total_ras = len(df_filt)

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 1 — POR DOMINIO
    # ─────────────────────────────────────────────────────────────────────────
    with tab_dom:
        st.subheader("Distribución por Dominio Taxonómico")
        st.caption(
            "**Cognitivo**: qué sabe el estudiante (conocimiento, comprensión, análisis…). "
            "**Procedimental**: qué sabe hacer (imitación → control). "
            "**Actitudinal**: cómo actúa y valora (percepción → caracterizar)."
        )

        conteo_dom = df_filt['Dominio'].value_counts().reindex(dominios_todos, fill_value=0)

        col_d1, col_d2 = st.columns([1, 1])
        with col_d1:
            fig_dom_pie = go.Figure(go.Pie(
                labels=conteo_dom.index.tolist(),
                values=conteo_dom.values.tolist(),
                marker_colors=[_DOMAIN_COLORS.get(d, '#BDC3C7') for d in conteo_dom.index],
                hole=0.38,
                textinfo='label+percent',
                pull=[0.04 if d == 'No identificado' else 0 for d in conteo_dom.index]
            ))
            fig_dom_pie.update_layout(
                title="% de RAs por dominio",
                height=370, showlegend=False,
                margin=dict(t=50, b=0, l=0, r=0)
            )
            st.plotly_chart(fig_dom_pie, use_container_width=True)

        with col_d2:
            st.markdown("**Resumen de dominios:**")
            for dom in dominios_todos:
                n   = int(conteo_dom.get(dom, 0))
                pct = n / total_ras * 100 if total_ras > 0 else 0
                color = _DOMAIN_COLORS.get(dom, '#BDC3C7')
                st.markdown(
                    f"<div style='margin:8px 0;padding:8px 12px;border-left:4px solid {color};"
                    f"border-radius:4px;background:#f8f9fa'>"
                    f"<b style='color:{color}'>{dom}</b><br>"
                    f"<span style='font-size:1.3em;font-weight:bold'>{n}</span> RAs "
                    f"<span style='color:#666'>({pct:.1f}%)</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

        # Distribución por programa
        st.markdown("---")
        st.subheader("Distribución por Dominio y Programa")
        prog_dom = (
            df_filt.groupby(['Programa', 'Dominio'])
            .size().reset_index(name='RAs')
        )
        total_prog = prog_dom.groupby('Programa')['RAs'].transform('sum')
        prog_dom['pct'] = prog_dom['RAs'] / total_prog * 100
        fig_bar_dom = px.bar(
            prog_dom, x='Programa', y='pct', color='Dominio',
            color_discrete_map=_DOMAIN_COLORS,
            category_orders={'Dominio': dominios_todos},
            barmode='stack', text_auto='.0f',
            labels={'pct': '% RAs', 'Programa': ''},
            title="Composición por dominio (%) por programa"
        )
        fig_bar_dom.update_layout(height=400, xaxis_tickangle=15, yaxis_title="% de RAs",
                                   legend=dict(orientation='h', y=-0.25))
        st.plotly_chart(fig_bar_dom, use_container_width=True)

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 2 — POR SUBCATEGORÍA
    # ─────────────────────────────────────────────────────────────────────────
    with tab_sub:
        st.subheader("Distribución por Subcategoría")
        st.caption("Seleccione uno o más dominios para filtrar las subcategorías mostradas.")

        dom_sub_filter = st.multiselect(
            "Dominio a mostrar:",
            ['Cognitivo', 'Procedimental', 'Actitudinal'],
            default=['Cognitivo', 'Procedimental', 'Actitudinal'],
            key="tax_dom_sub_filter"
        )
        df_sub = df_filt[df_filt['Dominio'].isin(dom_sub_filter)] if dom_sub_filter else df_filt
        df_sub = df_sub[df_sub['Subcategoria'] != 'No identificado']

        conteo_sub = (
            df_sub.groupby(['Dominio', 'Subcategoria'])
            .size().reset_index(name='RAs')
            .sort_values(['Dominio', 'RAs'], ascending=[True, False])
        )

        if conteo_sub.empty:
            st.info("No hay datos para las subcategorías seleccionadas.")
        else:
            fig_sub_bar = px.bar(
                conteo_sub, x='RAs', y='Subcategoria', color='Dominio',
                color_discrete_map=_DOMAIN_COLORS,
                orientation='h', text='RAs',
                labels={'RAs': 'Resultados de Aprendizaje', 'Subcategoria': ''},
                title="RAs clasificados por subcategoría taxonómica"
            )
            fig_sub_bar.update_layout(height=max(350, len(conteo_sub) * 30),
                                       yaxis={'categoryorder': 'total ascending'},
                                       legend=dict(orientation='h', y=-0.15))
            st.plotly_chart(fig_sub_bar, use_container_width=True)

            # Tabla resumen con % por subcategoría
            total_sub = int(conteo_sub['RAs'].sum())
            conteo_sub['%'] = (conteo_sub['RAs'] / total_sub * 100).round(1)
            st.dataframe(
                conteo_sub[['Dominio', 'Subcategoria', 'RAs', '%']]
                .reset_index(drop=True),
                use_container_width=True, hide_index=True
            )

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 3 — PROGRESIÓN POR SEMESTRE
    # ─────────────────────────────────────────────────────────────────────────
    with tab_prog:
        st.subheader("Progresión por Semestre")
        st.caption(
            "Un currículo progresivo muestra dominios procedimentales y actitudinales crecientes "
            "hacia semestres superiores, junto con subcategorías cognitivas más complejas."
        )

        prog_prog_filter = st.multiselect(
            "Programa (vacío = todos):",
            programas_todos, default=[], key="tax_prog_prog_filter"
        )
        vista_prog = st.radio(
            "Agrupar por:", ["Dominio", "Subcategoría"], horizontal=True, key="tax_vista_prog"
        )

        df_p = df_filt[df_filt['Programa'].isin(prog_prog_filter)] if prog_prog_filter else df_filt
        df_p = df_p.copy()
        df_p['Semestre_num'] = pd.to_numeric(df_p['Semestre'], errors='coerce')
        df_p = df_p.dropna(subset=['Semestre_num'])
        df_p['Semestre_num'] = df_p['Semestre_num'].astype(int)

        if df_p.empty:
            st.warning("No hay datos con semestre numérico para mostrar la progresión.")
        else:
            group_col = 'Dominio' if vista_prog == "Dominio" else 'Subcategoria'
            pivot_p = (
                df_p.groupby(['Semestre_num', group_col])
                .size().reset_index(name='count')
            )
            total_g = pivot_p.groupby('Semestre_num')['count'].transform('sum')
            pivot_p['pct'] = pivot_p['count'] / total_g * 100
            sems_ord = sorted(df_p['Semestre_num'].unique())

            color_map_prog = _DOMAIN_COLORS if vista_prog == "Dominio" else {
                row[group_col]: _SUBCAT_COLORS.get(row[group_col], '#BDC3C7')
                for _, row in pivot_p.iterrows()
            }
            cats_presentes = pivot_p[group_col].unique().tolist()

            fig_p = go.Figure()
            for cat in cats_presentes:
                datos_cat = pivot_p[pivot_p[group_col] == cat]
                val_map   = {int(r['Semestre_num']): r['pct'] for _, r in datos_cat.iterrows()}
                y_vals    = [val_map.get(s, 0) for s in sems_ord]
                fig_p.add_trace(go.Scatter(
                    x=sems_ord, y=y_vals, name=cat,
                    mode='lines+markers',
                    line=dict(color=color_map_prog.get(cat, '#BDC3C7'), width=2.5),
                    marker=dict(size=7),
                    stackgroup='one', groupnorm='percent'
                ))
            fig_p.update_layout(
                title=f"Composición por {group_col} (%) por semestre",
                xaxis_title="Semestre", yaxis_title="% de RAs",
                height=450,
                legend=dict(orientation='h', y=-0.3),
                xaxis=dict(tickmode='array', tickvals=sems_ord)
            )
            st.plotly_chart(fig_p, use_container_width=True)

    # ─────────────────────────────────────────────────────────────────────────
    # TAB 4 — EXPLORAR RAs
    # ─────────────────────────────────────────────────────────────────────────
    with tab_ra:
        st.subheader("Explorar Resultados de Aprendizaje")

        col_e1, col_e2, col_e3 = st.columns(3)
        with col_e1:
            dom_e = st.selectbox(
                "Dominio:", ['Todos'] + dominios_todos, key="tax_dom_e"
            )
        with col_e2:
            subcats_disp = sorted(df_tax['Subcategoria'].unique().tolist())
            sub_e = st.selectbox(
                "Subcategoría:", ['Todas'] + subcats_disp, key="tax_sub_e"
            )
        with col_e3:
            prog_e = st.selectbox(
                "Programa:", ['Todos'] + programas_todos, key="tax_prog_e"
            )

        df_e = df_tax.copy()
        if dom_e != 'Todos':
            df_e = df_e[df_e['Dominio'] == dom_e]
        if sub_e != 'Todas':
            df_e = df_e[df_e['Subcategoria'] == sub_e]
        if prog_e != 'Todos':
            df_e = df_e[df_e['Programa'] == prog_e]

        cols_show = ['Programa', 'Semestre', 'Nombre asignatura o modulo',
                     'Resultado de aprendizaje', 'Dominio', 'Subcategoria']
        cols_show = [c for c in cols_show if c in df_e.columns]
        df_e_show = df_e[cols_show].drop_duplicates().sort_values(['Programa', 'Semestre'])
        st.markdown(f"**{len(df_e_show)} RAs** coinciden con los filtros seleccionados.")
        st.dataframe(df_e_show, use_container_width=True, hide_index=True)

    # ─────────────────────────────────────────────────────────────────────────
    # ─────────────────────────────────────────────────────────────────────────
    with tab_mapa:
        st.subheader("Mapa de Integración entre Asignaturas")
        st.caption(
            "Visualiza qué asignaturas **comparten núcleos temáticos**. "
            "Un nodo = una asignatura; una línea = al menos un núcleo en común. "
            "El grosor de la línea indica la cantidad de núcleos compartidos. "
            "Útil para detectar redundancias, solapamientos u oportunidades de articulación curricular."
        )

        asig_col = 'Nombre asignatura o modulo'

        # Filtro por programa
        prog_mapa = st.multiselect(
            "Filtrar por programa:",
            sorted(df['Programa'].unique().tolist()),
            default=sorted(df['Programa'].unique().tolist()),
            key="mapa_prog_filter"
        )
        df_mapa = df[df['Programa'].isin(prog_mapa)] if prog_mapa else df

        # Construir diccionario: asignatura -> set de núcleos
        nucleos_dict: Dict[str, set] = {}
        prog_asig: Dict[str, str] = {}
        for _, row in df_mapa.iterrows():
            asig = str(row.get(asig_col, '')).strip()
            if not asig or asig in ('nan', ''):
                continue
            raw_nuc = str(row.get('Nucleos tematicos', '')).strip()
            if raw_nuc and raw_nuc not in ('nan', ''):
                nset = {
                    unicodedata.normalize('NFKD', limpiar_nucleo(n.strip()).lower()).encode('ascii', 'ignore').decode('ascii')
                    for n in _split_nucleos(raw_nuc)
                    if es_nucleo_valido(n.strip())[0]
                }
            else:
                nset = set()
            if asig not in nucleos_dict:
                nucleos_dict[asig] = set()
                prog_asig[asig] = str(row.get('Programa', ''))
            nucleos_dict[asig] |= nset

        subjects = sorted(nucleos_dict.keys())
        n_subjects = len(subjects)

        if n_subjects < 2:
            st.warning("Se necesitan al menos 2 asignaturas con núcleos temáticos para construir el mapa.")
        else:
            # Construir pares con núcleos compartidos
            pares_compartidos = []
            for i in range(n_subjects):
                for j in range(i + 1, n_subjects):
                    s1, s2 = subjects[i], subjects[j]
                    shared = nucleos_dict[s1] & nucleos_dict[s2]
                    if shared:
                        pares_compartidos.append({
                            'Asignatura 1': s1,
                            'Asignatura 2': s2,
                            'Programa 1': prog_asig[s1],
                            'Programa 2': prog_asig[s2],
                            'Núcleos compartidos': len(shared),
                            'Temas compartidos': ', '.join(sorted(shared)[:5]) + ('…' if len(shared) > 5 else '')
                        })

            df_pares = pd.DataFrame(pares_compartidos)

            if df_pares.empty:
                st.info("No se encontraron núcleos temáticos compartidos entre las asignaturas seleccionadas.")
            else:
                # Grado de conectividad
                degree: Counter = Counter()
                for _, row in df_pares.iterrows():
                    degree[row['Asignatura 1']] += int(row['Núcleos compartidos'])
                    degree[row['Asignatura 2']] += int(row['Núcleos compartidos'])

                connected_set = set(df_pares['Asignatura 1'].tolist() + df_pares['Asignatura 2'].tolist())
                isolated_subjects = [s for s in subjects if s not in connected_set]

                col_m1, col_m2, col_m3 = st.columns(3)
                col_m1.metric("Asignaturas en el mapa", n_subjects)
                col_m2.metric("Pares con núcleos compartidos", len(df_pares))
                col_m3.metric("Asignaturas sin conexiones", len(isolated_subjects),
                              delta="aisladas" if isolated_subjects else "Todas conectadas",
                              delta_color="off")
                st.markdown("---")

                # ── Red de integración (grafo circular) ──────────────────────
                st.subheader("Red de Integración Temática")
                max_nodes = 40
                if n_subjects > max_nodes:
                    st.info(
                        f"Hay **{n_subjects}** asignaturas. Mostrando las **{max_nodes}** "
                        f"con mayor conectividad temática."
                    )
                    top_subjects = [s for s, _ in degree.most_common(max_nodes)]
                else:
                    top_subjects = subjects

                n_top = len(top_subjects)
                idx_map = {s: i for i, s in enumerate(top_subjects)}
                angles_list = [2 * _math.pi * i / n_top for i in range(n_top)]
                x_pos = {s: _math.cos(a) for s, a in zip(top_subjects, angles_list)}
                y_pos = {s: _math.sin(a) for s, a in zip(top_subjects, angles_list)}

                # Color por programa
                progs_unicos = sorted({prog_asig[s] for s in top_subjects})
                palette = px.colors.qualitative.Set2 + px.colors.qualitative.Pastel
                prog_color_map = {p: palette[i % len(palette)] for i, p in enumerate(progs_unicos)}

                # Aristas
                max_shared = int(df_pares['Núcleos compartidos'].max()) if not df_pares.empty else 1
                edge_traces = []
                for _, row in df_pares.iterrows():
                    s1, s2 = row['Asignatura 1'], row['Asignatura 2']
                    if s1 not in idx_map or s2 not in idx_map:
                        continue
                    w = max(1, min(8, int(row['Núcleos compartidos']) / max_shared * 8))
                    alpha = round(0.15 + (int(row['Núcleos compartidos']) / max_shared) * 0.65, 2)
                    edge_traces.append(go.Scatter(
                        x=[x_pos[s1], x_pos[s2], None],
                        y=[y_pos[s1], y_pos[s2], None],
                        mode='lines',
                        line=dict(width=w, color=f'rgba(100,100,100,{alpha})'),
                        hoverinfo='none',
                        showlegend=False
                    ))

                # Nodos por programa (una traza por programa para la leyenda)
                node_traces = []
                for prog in progs_unicos:
                    subs_prog = [s for s in top_subjects if prog_asig[s] == prog]
                    node_traces.append(go.Scatter(
                        x=[x_pos[s] for s in subs_prog],
                        y=[y_pos[s] for s in subs_prog],
                        mode='markers+text',
                        name=prog[:35],
                        text=[s[:22] + ('…' if len(s) > 22 else '') for s in subs_prog],
                        textposition='top center',
                        textfont=dict(size=8),
                        marker=dict(
                            size=[10 + min(degree.get(s, 0), 25) for s in subs_prog],
                            color=prog_color_map[prog],
                            line=dict(width=1, color='white')
                        ),
                        hovertext=[
                            f"<b>{s}</b><br>Programa: {prog_asig[s]}<br>"
                            f"Peso de conexiones: {degree.get(s, 0)}<br>"
                            f"Núcleos únicos: {len(nucleos_dict.get(s, set()))}"
                            for s in subs_prog
                        ],
                        hoverinfo='text'
                    ))

                fig_net = go.Figure(data=edge_traces + node_traces)
                fig_net.update_layout(
                    title="Red de integración temática (nodos = asignaturas, líneas = núcleos compartidos)",
                    height=660,
                    showlegend=True,
                    legend=dict(title="Programa", orientation='v', x=1.01, font=dict(size=10)),
                    xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                    plot_bgcolor='white',
                    margin=dict(t=50, b=20, l=20, r=20)
                )
                st.plotly_chart(fig_net, use_container_width=True)
                st.caption(
                    "**Lectura:** El tamaño del nodo refleja el peso total de conexiones temáticas. "
                    "Nodos muy grandes son asignaturas articuladoras del currículo. "
                    "Nodos aislados (sin líneas) pueden indicar asignaturas sin integración con otras."
                )

                st.markdown("---")

                # ── Heatmap de núcleos compartidos (si hay pocas asignaturas) ─
                if n_subjects <= 35:
                    st.subheader("Matriz de Integración (Núcleos Compartidos)")
                    st.caption(
                        "Cada celda muestra cuántos núcleos temáticos comparten dos asignaturas. "
                        "Valores altos indican alta integración (o posible redundancia a revisar)."
                    )
                    mat = pd.DataFrame(0, index=subjects, columns=subjects)
                    for _, row in df_pares.iterrows():
                        mat.loc[row['Asignatura 1'], row['Asignatura 2']] = int(row['Núcleos compartidos'])
                        mat.loc[row['Asignatura 2'], row['Asignatura 1']] = int(row['Núcleos compartidos'])
                    short = {s: (s[:30] + '…' if len(s) > 30 else s) for s in subjects}
                    mat_disp = mat.rename(index=short, columns=short)
                    fig_heat = px.imshow(
                        mat_disp.values,
                        x=mat_disp.columns.tolist(),
                        y=mat_disp.index.tolist(),
                        color_continuous_scale='Blues',
                        aspect='auto',
                        title="Núcleos temáticos compartidos entre asignaturas",
                        labels=dict(color="Núcleos comunes")
                    )
                    fig_heat.update_layout(
                        height=max(400, n_subjects * 18),
                        xaxis_tickangle=45,
                        xaxis=dict(tickfont=dict(size=9)),
                        yaxis=dict(tickfont=dict(size=9))
                    )
                    st.plotly_chart(fig_heat, use_container_width=True)
                    st.markdown("---")

                # ── Top pares con más núcleos compartidos ─────────────────────
                st.subheader("Pares con Mayor Integración Temática")
                st.caption(
                    "Asignaturas que comparten más núcleos temáticos — "
                    "candidatas a articulación explícita o revisión de solapamiento."
                )
                df_top = df_pares.sort_values('Núcleos compartidos', ascending=False).head(20)
                st.dataframe(df_top, use_container_width=True, hide_index=True)

                # ── Asignaturas aisladas ───────────────────────────────────────
                if isolated_subjects:
                    with st.expander(f"⚠️ Asignaturas sin integración temática ({len(isolated_subjects)})"):
                        st.caption(
                            "Estas asignaturas no comparten ningún núcleo temático con otras. "
                            "Pueden ser especializadas (normal) o indicar falta de articulación curricular."
                        )
                        for s in sorted(isolated_subjects):
                            st.markdown(f"- **{s}** — Programa: {prog_asig.get(s, 'N/A')}")


def _cargar_cobertura_perfil_por_programa(uploaded_files) -> Dict:
    """Lee archivos y ejecuta analisis de cobertura perfil por programa."""
    import hashlib

    cache_key = 'cobertura_perfil_cache'
    _cache_ver = 'v2_hibrido_bm25'
    files_sig = hashlib.md5(
        (_cache_ver + ',' + ','.join(sorted(f.name for f in uploaded_files))).encode()
    ).hexdigest()

    if cache_key in st.session_state and st.session_state.get(cache_key + '_sig') == files_sig:
        return st.session_state[cache_key]

    resultados = {}
    for f in uploaded_files:
        try:
            nombre = f.name
            programa = (
                nombre.replace("FormatoRA_", "").replace("FormatoRA-", "")
                .replace("_PBOG", "").replace("_VNAL", "").replace("_PMED", "")
                .replace("_HBOG", "").replace("_HMED", "").replace("_HVAL", "")
                .replace(".xlsx", "").replace(".xls", "").strip()
            )

            f.seek(0)
            df_perfil = pd.read_excel(f, sheet_name='Paso1 Analisis perfil egreso', header=1, engine='openpyxl')

            f.seek(0)
            try:
                df_ra = pd.read_excel(f, sheet_name='Paso 3 Redacción RA', header=1, engine='openpyxl')
            except Exception:
                df_ra = pd.DataFrame()

            f.seek(0)
            df_micro = pd.read_excel(f, sheet_name='Paso 5 Estrategias micro', header=1, engine='openpyxl')

            f.seek(0)

            if df_perfil.empty:
                continue

            resultado = analizar_cobertura_perfil_completa(df_perfil, df_micro, df_ra)
            resultado['programa'] = programa
            resultado['archivo'] = nombre
            resultados[programa] = resultado
        except Exception as e:
            continue

    st.session_state[cache_key] = resultados
    st.session_state[cache_key + '_sig'] = files_sig
    return resultados


def _color_fila_score_cobertura(score: float) -> str:
    """Color de fondo según zona de riesgo del score (0-1)."""
    if score < 0.25:
        return 'background-color: #ffcccc'
    if score < 0.35:
        return 'background-color: #ffe0cc'
    if score < 0.50:
        return 'background-color: #fff8cc'
    if score < 0.70:
        return 'background-color: #cce5ff'
    return 'background-color: #ccffcc'


def _styler_tabla_cobertura_perfil(df: pd.DataFrame) -> "pd.io.formats.style.Styler":
    """Aplica gradiente de riesgo por fila según score numérico."""
    scores = df['_score_num'].tolist()

    def _fila(row):
        idx = row.name
        s = scores[idx] if idx < len(scores) else 0
        return [_color_fila_score_cobertura(s)] * len(row)

    cols_show = [c for c in df.columns if c != '_score_num']
    return df[cols_show].style.apply(_fila, axis=1)


def _distribucion_scores_cobertura(elementos: list) -> dict:
    """Cuenta elementos por rango de score."""
    rangos = {
        '< 0.25': 0,
        '0.25–0.35': 0,
        '0.35–0.50': 0,
        '0.50–0.70': 0,
        '> 0.70': 0,
    }
    for e in elementos:
        s = e.get('score', 0)
        if s < 0.25:
            rangos['< 0.25'] += 1
        elif s < 0.35:
            rangos['0.25–0.35'] += 1
        elif s < 0.50:
            rangos['0.35–0.50'] += 1
        elif s < 0.70:
            rangos['0.50–0.70'] += 1
        else:
            rangos['> 0.70'] += 1
    return rangos


def pagina_cobertura_perfil(df_micro: pd.DataFrame):
    """Evalua si las asignaturas cubren las actitudes/aptitudes del perfil de egreso."""
    st.title("Cobertura de Perfil de Egreso")
    st.markdown("---")
    st.info(
        "**¿Qué mide esta sección?** Compara cada elemento del perfil de egreso "
        "(Saber, SaberHacer, SaberSer, Áreas profesionales, Tareas profesionales, "
        "Valor agregado) contra el contenido curricular con **TF-IDF + BM25** (score híbrido). "
        "Un elemento se clasifica **CUBIERTO** si su score supera el **umbral por campo** "
        "(p. ej. 0.35 en Saber, 0.38 en Áreas profesionales). "
        "Las filas coloreadas indican nivel de riesgo; la columna **Asignatura trazable** "
        "muestra dónde se detectó la mayor similitud en el currículo."
    )

    uploaded_files = st.session_state.get('archivos_subidos', [])
    if not uploaded_files:
        st.warning("No hay archivos cargados. Sube archivos desde la página de Inicio.")
        return

    with st.spinner("Analizando cobertura del perfil de egreso por programa..."):
        resultados = _cargar_cobertura_perfil_por_programa(uploaded_files)

    if not resultados:
        st.warning("No se pudieron analizar perfiles. Verifica que los archivos tengan la hoja 'Paso1 Analisis perfil egreso'.")
        return

    progs_con_perfil = [r for r in resultados.values() if r['total_elementos'] > 0]
    progs_sin_perfil = [r for r in resultados.values() if r['total_elementos'] == 0]

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Programas con perfil", len(progs_con_perfil))
    col2.metric("Programas sin perfil", len(progs_sin_perfil))
    if progs_con_perfil:
        avg_cob = np.mean([r['cobertura_global'] for r in progs_con_perfil])
        total_brechas = sum(r['num_brechas'] for r in progs_con_perfil)
        en_riesgo = sum(
            1 for r in progs_con_perfil
            for e in r.get('elementos', [])
            if e.get('clasificacion') == 'CUBIERTO' and 0.35 <= e.get('score', 0) < 0.50
        )
        col3.metric("Cobertura global promedio", f"{avg_cob:.1f}%")
        col4.metric("Brechas totales", total_brechas)
        col5.metric("Cobertura superficial", en_riesgo)

    st.markdown("---")

    # ── Heatmap campo × programa ──
    if progs_con_perfil and any(r.get('cobertura_por_campo') for r in progs_con_perfil):
        st.subheader("Mapa de cobertura por campo")
        from config import COLUMNAS_PERFIL
        filas_heat = {}
        for r in progs_con_perfil:
            filas_heat[r['programa']] = {
                c: r.get('cobertura_por_campo', {}).get(c, 0.0)
                for c in COLUMNAS_PERFIL
            }
        df_heat = pd.DataFrame(filas_heat).T
        df_heat = df_heat[[c for c in COLUMNAS_PERFIL if c in df_heat.columns]]
        fig_heat = px.imshow(
            df_heat,
            labels=dict(x='Campo del perfil', y='Programa', color='Cobertura (%)'),
            color_continuous_scale='RdYlGn',
            zmin=0, zmax=100,
            aspect='auto',
            text_auto='.0f',
        )
        fig_heat.update_layout(height=max(280, len(df_heat) * 40 + 120))
        st.plotly_chart(fig_heat, use_container_width=True)

        st.subheader("Comparativa de cobertura por campo")
        df_campo_long = df_heat.reset_index(names='Programa').melt(
            id_vars='Programa',
            var_name='Campo',
            value_name='Cobertura (%)',
        )
        fig_campo = px.bar(
            df_campo_long,
            y='Campo',
            x='Cobertura (%)',
            color='Programa',
            orientation='h',
            barmode='group',
            title='Cobertura por campo y programa',
        )
        fig_campo.update_layout(height=max(400, len(COLUMNAS_PERFIL) * 36))
        st.plotly_chart(fig_campo, use_container_width=True)

    st.markdown("---")

    # ── Resultados por programa ──
    st.subheader("Resultados por Programa")
    ordenados = sorted(progs_con_perfil, key=lambda r: r['cobertura_global'])

    for r in ordenados:
        color = "🟢" if r['cobertura_global'] >= 60 else ("🟡" if r['cobertura_global'] >= 35 else "🔴")
        with st.expander(
            f"{color} **{r['programa']}** — "
            f"Cobertura: {r['cobertura_global']}% | "
            f"{r['total_elementos'] - r['num_brechas']}/{r['total_elementos']} cubiertos | "
            f"{r['num_brechas']} brecha(s)"
        ):
            col_tab, col_rec = st.columns([2, 1])
            with col_tab:
                elems = sorted(r.get('elementos', []), key=lambda e: e.get('score', 0))
                rows = []
                for e in elems:
                    traz = e.get('asignatura_trazable') or '— sin cobertura detectada'
                    if e.get('doc_trazable') and e.get('asignatura_trazable'):
                        traz = f"{traz} · {e['doc_trazable'][:60]}"
                    rows.append({
                        'Campo': e['campo'],
                        'Elemento': e['elemento'][:80] + ('…' if len(e['elemento']) > 80 else ''),
                        'Score': f"{e['score']:.2%}",
                        'Umbral': f"{e.get('umbral', 0.35):.0%}",
                        'Estado': e['clasificacion'],
                        'Asignatura trazable': traz[:100],
                        '_score_num': e['score'],
                    })
                df_elem = pd.DataFrame(rows)
                st.caption(
                    "Colores: rojo <25% · naranja 25–35% · amarillo 35–50% (superficial) · "
                    "azul 50–70% · verde >70%"
                )
                st.dataframe(
                    _styler_tabla_cobertura_perfil(df_elem),
                    use_container_width=True,
                    hide_index=True,
                )

                dist = _distribucion_scores_cobertura(r.get('elementos', []))
                fig_dist = px.bar(
                    x=list(dist.keys()),
                    y=list(dist.values()),
                    labels={'x': 'Rango de score', 'y': 'Cantidad de elementos'},
                    title=f"Distribución de scores — {r['programa']}",
                    text=list(dist.values()),
                )
                fig_dist.update_traces(textposition='outside')
                fig_dist.update_layout(height=280)
                st.plotly_chart(fig_dist, use_container_width=True)

            with col_rec:
                if r.get('cobertura_por_campo'):
                    st.markdown("**Cobertura por campo:**")
                    for campo, pct in sorted(
                        r['cobertura_por_campo'].items(), key=lambda x: x[1]
                    ):
                        st.markdown(f"- {campo}: **{pct}%**")
                if r['recomendaciones']:
                    st.markdown("**Recomendaciones:**")
                    for rec in r['recomendaciones']:
                        st.markdown(f"- {rec}")
                st.markdown(f"**Corpus:** {r.get('corpus_size', 0)} documentos")

    st.markdown("---")

    if len(ordenados) >= 2:
        st.subheader("Comparación de cobertura entre programas")
        df_chart = pd.DataFrame([
            {
                'Programa': r['programa'],
                'Cobertura Global (%)': r['cobertura_global'],
                'Cubiertos': r['total_elementos'] - r['num_brechas'],
                'Brechas': r['num_brechas'],
            }
            for r in ordenados
        ])
        fig = px.bar(
            df_chart, x='Programa', y='Cobertura Global (%)',
            color='Cobertura Global (%)', color_continuous_scale='RdYlGn',
            text='Cobertura Global (%)', range_color=[0, 100],
        )
        fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
        fig.update_layout(height=400, xaxis_tickangle=45)
        st.plotly_chart(fig, use_container_width=True)

        fig2 = px.bar(
            df_chart, x='Programa', y=['Cubiertos', 'Brechas'],
            barmode='group', title='Elementos cubiertos vs. brechas por programa',
            color_discrete_map={'Cubiertos': '#1fb2de', 'Brechas': '#ec0677'},
        )
        fig2.update_layout(height=400, xaxis_tickangle=45)
        st.plotly_chart(fig2, use_container_width=True)


def pagina_familias_curriculares(df: pd.DataFrame, resultados_nlp: Dict):
    """Muestra asignaturas compartidas entre programas (familias curriculares)."""
    st.title("Familias Curriculares")
    st.markdown("---")
    st.info(
        "**¿Qué son las Familias Curriculares?** Conjuntos de programas que "
        "comparten asignaturas. La matriz muestra qué asignaturas están presentes "
        "en múltiples programas y en cuáles específicamente."
    )

    asig_col = _find_column(df, 'Nombre asignatura o módulo')
    if asig_col is None:
        st.warning("No se encontró la columna de asignaturas en los datos")
        return

    agrupado = (
        df.groupby(asig_col)
        .agg(
            programas=('Programa', lambda x: sorted(set(x))),
            sedes=('Sede', lambda x: sorted(set(x))),
            conteo_programas=('Programa', 'nunique')
        )
        .reset_index()
    )

    compartidas = agrupado[agrupado['conteo_programas'] > 1].copy()
    compartidas = compartidas.sort_values('conteo_programas', ascending=False)
    compartidas.rename(columns={asig_col: 'Asignatura'}, inplace=True)

    if compartidas.empty:
        st.success("No se encontraron asignaturas compartidas entre programas")
        return

    st.metric("Asignaturas compartidas", len(compartidas))

    programas_unicos = df['Programa'].nunique()
    st.caption(
        f"De {programas_unicos} programas cargados, se encontraron "
        f"{len(compartidas)} asignaturas que se repiten en 2 o más programas."
    )

    total_programas = df['Programa'].nunique()
    pivot_rows = []
    for _, row in compartidas.iterrows():
        prog_list = row['programas']
        pivot_row = {'Asignatura': row['Asignatura'], 'Total Programas': row['conteo_programas']}
        for prog in sorted(df['Programa'].unique()):
            pivot_row[prog] = 'X' if prog in prog_list else ''
        pivot_rows.append(pivot_row)

    pivot_df = pd.DataFrame(pivot_rows)
    st.dataframe(pivot_df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.subheader("Mapa de calor: Asignaturas × Programas")
    heat_pivot = compartidas.copy()
    for prog in sorted(df['Programa'].unique()):
        heat_pivot[prog] = heat_pivot['programas'].apply(lambda x: 1 if prog in x else 0)

    mat_heat = heat_pivot.drop(
        columns=['programas', 'sedes', 'conteo_programas', 'Asignatura']
    )
    mat_heat.index = heat_pivot['Asignatura'].str[:60]

    if not mat_heat.empty:
        fig = px.imshow(
            mat_heat,
            color_continuous_scale='Blues',
            aspect='auto',
            labels=dict(x='Programa', y='Asignatura', color='Presente')
        )
        fig.update_layout(
            height=max(300, len(mat_heat) * 20),
            margin=dict(l=200, r=20, t=20, b=120)
        )
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("Asignaturas por cantidad de programas")
    hist = compartidas['conteo_programas'].value_counts().sort_index()
    fig_bar = px.bar(
        x=hist.index, y=hist.values,
        labels={'x': 'Programas que comparten la asignatura', 'y': 'Cantidad de asignaturas'},
        text=hist.values
    )
    fig_bar.update_traces(textposition='outside')
    st.plotly_chart(fig_bar, use_container_width=True)


def pagina_config_tendencias():
    """Pagina para configurar tendencias personalizables."""
    st.title("Configurar Tendencias Globales")
    st.markdown("---")

    tendencias = obtener_tendencias()

    # --- Importar tendencias desde JSON ---
    st.subheader("Importar / Exportar Configuracion")
    col_imp, col_exp = st.columns(2)

    with col_imp:
        uploaded_json = st.file_uploader(
            "Importar tendencias desde JSON:",
            type=['json'],
            key='json_upload'
        )
        if uploaded_json is not None:
            try:
                config = json.load(uploaded_json)
                nuevas = config.get('TENDENCIAS_GLOBALES', config)
                if isinstance(nuevas, dict) and len(nuevas) > 0:
                    st.session_state['tendencias'] = nuevas
                    st.success(f"{len(nuevas)} tendencias importadas.")
                    st.rerun()
            except Exception as e:
                st.error(f"Error al importar: {e}")

    with col_exp:
        config_export = {
            "TENDENCIAS_GLOBALES": tendencias,
            "VERSION": "1.0"
        }
        json_str = json.dumps(config_export, ensure_ascii=False, indent=2)
        st.download_button(
            label="Descargar configuracion actual (JSON)",
            data=json_str,
            file_name="config_tendencias.json",
            mime="application/json"
        )

    st.markdown("---")

    # --- Tendencias actuales ---
    st.subheader(f"Tendencias Actuales ({len(tendencias)})")
    for tid, tinfo in tendencias.items():
        with st.expander(f"{tinfo['descripcion']} ({len(tinfo['keywords'])} keywords)"):
            st.markdown(f"**ID:** `{tid}`")
            st.markdown(f"**Color:** {tinfo['color']}")
            st.markdown(f"**Keywords:** {', '.join(tinfo['keywords'])}")

    st.markdown("---")

    # --- Agregar nueva tendencia ---
    st.subheader("Agregar Nueva Tendencia")
    with st.form("nueva_tendencia"):
        col1, col2 = st.columns(2)
        with col1:
            nuevo_id = st.text_input("ID (sin espacios):", placeholder="MI_TENDENCIA")
            nueva_desc = st.text_input("Descripcion:", placeholder="Mi nueva tendencia")
        with col2:
            nuevo_color = st.color_picker("Color:", "#FF5733")
            nuevas_keywords = st.text_area(
                "Keywords (una por linea):",
                placeholder="keyword1\nkeyword2\nfrase clave"
            )

        submitted = st.form_submit_button("Agregar Tendencia")

        if submitted and nuevo_id and nueva_desc and nuevas_keywords:
            keywords_list = [k.strip() for k in nuevas_keywords.split('\n') if k.strip()]
            if keywords_list:
                st.session_state['tendencias'][nuevo_id] = {
                    "keywords": keywords_list,
                    "color": nuevo_color,
                    "descripcion": nueva_desc
                }
                st.success(f"Tendencia '{nueva_desc}' agregada con {len(keywords_list)} keywords.")
                st.rerun()
            else:
                st.error("Debes ingresar al menos una keyword.")

    st.markdown("---")

    # --- Eliminar tendencia ---
    st.subheader("Eliminar Tendencia")
    if tendencias:
        tid_eliminar = st.selectbox(
            "Seleccionar tendencia a eliminar:",
            list(tendencias.keys()),
            format_func=lambda x: tendencias[x]['descripcion']
        )
        if st.button("Eliminar Tendencia Seleccionada"):
            del st.session_state['tendencias'][tid_eliminar]
            st.success("Tendencia eliminada.")
            st.rerun()

    st.markdown("---")

    # --- Restaurar por defecto ---
    if st.button("Restaurar Tendencias por Defecto"):
        st.session_state['tendencias'] = json.loads(json.dumps(TENDENCIAS_DEFAULT))
        st.success("Tendencias restauradas a valores por defecto.")
        st.rerun()


def pagina_datos(df: pd.DataFrame):
    """Pagina de exploracion de datos crudos."""
    st.title("Explorar Datos")
    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        prog_filter = st.multiselect(
            "Programa:", df['Programa'].unique(),
            default=list(df['Programa'].unique())
        )
    with col2:
        tipo_filter = st.multiselect(
            "Tipo de Saber:", df['Tipo de Saber'].unique(),
            default=list(df['Tipo de Saber'].unique())
        )
    with col3:
        semestres_disponibles = sorted(df['Semestre'].dropna().unique(), key=lambda x: str(x))
        sem_filter = st.multiselect(
            "Semestre:", semestres_disponibles,
            default=list(semestres_disponibles)
        )

    df_filtered = df[
        (df['Programa'].isin(prog_filter)) &
        (df['Tipo de Saber'].isin(tipo_filter)) &
        (df['Semestre'].isin(sem_filter))
    ]

    st.markdown(f"**Registros filtrados:** {len(df_filtered)}")

    columnas_mostrar = [
        'Programa', 'Tipo de Saber', 'Semestre',
        'Nombre asignatura o modulo', 'Resultado de aprendizaje',
        'Indicadores de logro asignatura o modulo', 'Nucleos tematicos'
    ]
    cols_disponibles = [c for c in columnas_mostrar if c in df_filtered.columns]
    st.dataframe(df_filtered[cols_disponibles], use_container_width=True, hide_index=True)

    csv = df_filtered.to_csv(index=False, encoding='utf-8-sig')
    st.download_button(
        label="Descargar datos filtrados (CSV)",
        data=csv,
        file_name="datos_filtrados.csv",
        mime="text/csv"
    )


# ============================================================================
# MAIN
# ============================================================================

def main():
    # --- SIDEBAR SIEMPRE VISIBLE (menu + filtros) ---
    # Sidebar styling consistente oscuro
    st.markdown("""
    <style>
    section[data-testid="stSidebar"] {
        background-color: #0F385A !important;
    }
    section[data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }
    section[data-testid="stSidebar"] .stMarkdown h3 {
        color: #4DA6FF !important;
    }
    section[data-testid="stSidebar"] .stExpander summary {
        background-color: #0F385A !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Navegacion con option_menu (SIEMPRE se renderiza)
    PAGINAS = {
        "Inicio":               ("house",          "Resumen general y métricas clave del currículo"),
        "Tipo de Saber":        ("bar-chart",       "Saber, SaberHacer y SaberSer por semestre y asignatura"),
        "Cobertura de Perfil":  ("person-check",    "Cobertura del perfil de egreso vs. currículo"),
        "Cobertura Temática":   ("map",             "Núcleos temáticos: diversidad y densidad por programa"),
        "Tendencias Globales":  ("graph-up-arrow",  "Alineación con IA, Sostenibilidad, Innovación, etc."),
        "Minería de Texto":     ("search",          "Términos clave, similitud y frases frecuentes"),
        "Bloom & Integración":  ("diagram-3",       "Taxonomía de Bloom y mapa de integración temática"),
        "Familias Curriculares":("diagram-2",       "Clustering jerárquico y familias de programas"),
        "Configurar Tendencias":("sliders",         "Personalizar las tendencias globales a detectar"),
        "Explorar Datos":       ("table",           "Explorar y filtrar los registros cargados"),
    }

    with st.sidebar:
        st.title("Configuracion")
        st.markdown("---")
        
        pagina = option_menu(
            menu_title=None,
            options=list(PAGINAS.keys()),
            icons=[v[0] for v in PAGINAS.values()],
            default_index=0,
            key="nav_menu",
            styles={
                "container": {
                    "padding": "4px 0",
                    "background-color": "#0F385A",
                    "border-radius": "0px",
                },
                "icon": {
                    "display": "inline-block",
                    "color": "#42F2F2",
                    "font-size": "15px",
                    "margin-right": "8px",
                },
                "nav-link": {
                    "font-size": "13px",
                    "color": "#FFFFFF",
                    "background-color": "transparent",
                    "padding": "8px 14px",
                    "border-radius": "6px",
                    "margin": "2px 4px",
                    "display": "flex",
                    "align-items": "center",
                    "--hover-color": "rgba(31,178,222,0.25)",
                },
                "nav-link-selected": {
                    "background-color": "rgba(31,178,222,0.35)",
                    "border-left": "3px solid #FBAF17",
                    "font-weight": "600",
                    "color": "#FFFFFF",
                },
            }
        )

        st.caption(f"_{PAGINAS[pagina][1]}_")
        st.markdown("---")
        st.markdown(
            """
            <div style='text-align:center;padding:10px 0;opacity:0.7'>
            <div style='font-size:0.7em;color:rgba(255,255,255,0.6)'>
            Politécnico Grancolombiano<br>
            Análisis Microcurricular v2.0
            </div></div>
            """,
            unsafe_allow_html=True
        )
    
    # Carga de archivos en session state
    if 'archivos_subidos' not in st.session_state:
        st.session_state['archivos_subidos'] = None
    
    uploaded_files = st.session_state['archivos_subidos']
    
    # Scroll al top al cambiar de página
    st_components.html(
        "<script>window.parent.document.querySelector('.main').scrollTo(0, 0);</script>",
        height=0
    )
    
    # ── Sin archivos: mostrar banner + uploader ──────────────────────────
    if not uploaded_files:
        # CSS global para la landing page
        st.markdown("""
        <style>
        /* ── Encabezado de página ── */
        .page-header { margin-bottom: 20px; }
        .page-header h1 { margin: 0; font-size: 2.25rem; color: #0f3460; line-height: 1.05; }
        .page-header p  { margin: 10px 0 0; color: #475569; font-size: 0.98rem; line-height: 1.7; max-width: 800px; }

        /* ── Panel izquierdo (hero) via st.container(key="hero_left_panel") ── */
        /* ── Hero global unificado ── */
        .st-key-hero_section {
            background: linear-gradient(135deg, #1a3a52 0%, #1e5080 60%, #2a6494 100%);
            border-radius: 22px;
            padding: 24px;
            box-shadow: 0 20px 60px rgba(15,23,42,.22);
            border: 1px solid rgba(31,178,222,.16);
            overflow: hidden;
        }
        .st-key-hero_section [data-testid="stHorizontalBlock"] {
            display: flex !important;
            align-items: stretch !important;
            gap: 20px !important;
        }
        .st-key-hero_section [data-testid="stColumn"] {
            display: flex !important;
            flex-direction: column !important;
            gap: 0 !important;
        }
        .st-key-hero_section [data-testid="stVerticalBlock"] {
            display: flex !important;
            flex-direction: column !important;
            flex: 1 1 auto !important;
        }
        .hero-left-panel {
            color: #ffffff !important;
            padding: 30px 32px 28px;
            display: flex;
            flex-direction: column;
            justify-content: center;
            min-height: 245px;
            gap: 18px;
            position: relative;
            z-index: 1;
        }
        .hero-left-panel:before {
            content: '';
            position: absolute;
            inset: 0;
            background: radial-gradient(circle at top left, rgba(31,178,222,.14), transparent 32%);
            pointer-events: none;
        }
        .hero-left-panel > * {
            position: relative;
            z-index: 2;
        }
        .hero-left-panel .hero-bar {
            width: 5px; height: 44px; background:#1fb2de;
            border-radius: 3px;
        }
        .hero-left-panel .hero-label {
            display:block; font-size:.78rem; font-weight:800;
            text-transform:uppercase; letter-spacing:.18em;
            color:#b6e5ff !important;
        }
        .hero-left-panel h2 {
            margin:0; font-size:1.9rem; line-height:1.12;
            font-weight:800; color:#ffffff !important;
            text-shadow: 0 2px 18px rgba(0,0,0,.18);
        }
        .hero-left-panel p {
            margin:0; color: rgba(255,255,255,.92) !important;
            font-size:1.02rem; line-height:1.75;
            max-width: 520px;
        }
        .hero-left-panel .hero-support {
            color: rgba(255,255,255,.85) !important;
            font-size: 0.95rem;
            line-height: 1.7;
            max-width: 520px;
        }
        .st-key-hero_right_card {
            background: rgba(255,255,255,.12);
            border: 1px solid rgba(255,255,255,.18);
            border-radius: 22px;
            padding: 22px 20px;
            min-height: 260px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            gap: 18px;
            backdrop-filter: blur(18px);
            color: #ffffff !important;
        }
        .st-key-hero_right_card * {
            color: #ffffff !important;
        }
        .upload-panel-deco {
            text-align: center;
            padding: 18px 0 16px;
            border-bottom: 1px solid rgba(255,255,255,.16);
            margin-bottom: 10px;
        }
        .upload-icon-box {
            width: 56px; height: 56px; border-radius: 16px;
            background: rgba(31,178,222,.22);
            display: inline-flex; align-items: center; justify-content: center;
            margin: 0 auto 12px;
        }
        .upload-panel-deco h4 { margin: 0 0 6px; font-size: 1.02rem; font-weight: 700; color: #ffffff !important; }
        .upload-panel-deco span { color: #e6f7ff !important; font-size: .84rem; }
        .upload-help {
            color: rgba(255,255,255,.86) !important;
            font-size: .84rem;
            margin-top: 4px;
        }
        .st-key-hero_right_card [data-testid="stFileUploaderDropzone"] button {
            box-shadow: 0 14px 30px rgba(31,178,222,.18) !important;
            transition: transform .18s ease, box-shadow .18s ease !important;
        }
        .st-key-hero_right_card [data-testid="stFileUploaderDropzone"] button:hover {
            transform: translateY(-1px) !important;
            box-shadow: 0 18px 34px rgba(31,178,222,.26) !important;
        }

        /* ── File uploader dentro del panel derecho ── */
        .st-key-hero_section [data-testid="stFileUploaderDropzone"] {
            background: transparent !important;
            border: none !important;
            padding: 0 !important;
        }
        .st-key-hero_section [data-testid="stFileUploader"] section {
            background: transparent !important;
            border: none !important;
            padding: 0 !important;
        }
        .st-key-hero_section [data-testid="stFileUploaderDropzoneInstructions"] {
            display: none !important;
        }
        .st-key-hero_section [data-testid="stFileUploaderDropzone"] button {
            background: #1fb2de !important;
            color: #0f2d44 !important;
            font-weight: 700 !important;
            border-radius: 10px !important;
            border: none !important;
            width: 100% !important;
            min-height: 44px !important;
            font-size: .95rem !important;
            margin-bottom: 10px !important;
        }
        .st-key-hero_section [data-testid="stFileUploaderDropzone"] button:hover {
            background: #19a0cc !important;
        }
        .st-key-hero_section [data-testid="stFileUploaderFileName"] { color: #e0f2fe !important; }
        .st-key-hero_section [data-testid="stFileUploader"] small,
        .st-key-hero_section [data-testid="stFileUploader"] > div > small {
            display: none !important;
        }
        /* Nombres de archivos en la lista del hero */
        .st-key-hero_section [data-testid="stFileUploader"] span,
        .st-key-hero_section [data-testid="stFileUploader"] p,
        .st-key-hero_section [data-testid="stFileUploader"] div,
        .st-key-hero_section [data-testid="stFileUploader"] label {
            color: #e0f2fe !important;
        }
        .st-key-hero_section [data-testid="stBaseButton-secondary"] {
            background: rgba(31,178,222,.08) !important;
            border: 1.5px solid rgba(125,211,240,.5) !important;
            color: #e6f5ff !important;
            border-radius: 10px !important;
            font-weight: 600 !important;
            min-height: 40px !important;
        }
        .st-key-hero_section [data-testid="stBaseButton-secondary"]:hover {
            background: rgba(31,178,222,.2) !important;
            border-color: #b8ecff !important;
        }

        /* ── Tarjeta de estructura ── */
        .estructura-card {
            margin-top: 8px;
            border: 1px solid rgba(125,211,240,.3);
            border-radius: 10px;
            background: rgba(11,45,74,.9);
            color: #eaf6ff;
            padding: 10px 12px;
            font-size: .82rem;
            line-height: 1.5;
        }
        .estructura-card h5 { margin: 0 0 6px; font-size: .9rem; color: #fff; }
        .estructura-card ul { margin: 0 0 8px 16px; padding: 0; }
        .estructura-mini-table { width: 100%; border-collapse: collapse; font-size: .76rem; }
        .estructura-mini-table th, .estructura-mini-table td {
            border: 1px solid rgba(125,211,240,.25);
            padding: 5px 7px; color: #eaf6ff; text-align: left;
        }
        .estructura-mini-table th { background: rgba(31,178,222,.18); font-weight: 700; color: #fff; }

        /* ── Tarjetas de características ── */
        .feature-grid { display:grid; grid-template-columns:repeat(3, minmax(240px, 1fr)); gap:16px; margin-bottom:18px; }
        .feature-grid.row-2 { grid-template-columns:repeat(3, minmax(240px, 1fr)); }
        .feature-card {
            background:#fff; border-radius:18px; padding:20px;
            box-shadow:0 8px 24px rgba(15,23,42,.07);
            min-height:165px; display:flex; flex-direction:column; justify-content:space-between;
            border:1px solid #e8f1ff; transition:all .25s ease;
            border-top:4px solid var(--card-color,#1fb2de);
        }
        .feature-card:hover { transform:translateY(-4px); box-shadow:0 16px 40px rgba(15,23,42,.12); }
        .feature-icon-box { width:42px; height:42px; border-radius:10px; display:flex; align-items:center; justify-content:center; margin-bottom:12px; background:var(--card-bg,#e8f8ff); }
        .feature-card h3 { margin:0 0 9px; font-size:.96rem; font-weight:700; color:#0d314e; }
        .feature-card p  { margin:0; color:#6c7a93; font-size:.86rem; line-height:1.5; flex:1; }
        .feature-meta { display:flex; justify-content:space-between; align-items:center; margin-top:14px; font-size:.76rem; font-weight:700; }
        .feature-meta .status { color:var(--card-color,#0f3460); }
        .feature-meta .status.alert { color:#d93025; }
        .feature-meta .arrow { color:#b0c4de; }
        </style>
        """, unsafe_allow_html=True)

        # ── Encabezado ──
        st.markdown("""
        <div class="page-header">
            <h1>Análisis Microcurricular Basado en Datos</h1>
        </div>
        """, unsafe_allow_html=True)

        # ── Hero: un solo bloque unificado con columnas internas ──
        icon_upload = render_icon_svg('upload', '#1fb2de', 38)
        with st.container(key="hero_section"):
            _c1, _c2 = st.columns([1.6, 1], gap="large")

            with _c1:
                st.markdown("""
                <div class="hero-left-panel">
                    <div class="hero-bar"></div>
                    <span class="hero-label">Sistema de Análisis Microcurricular</span>
                    <p>Optimice los procesos de diseño, actualización y evaluación curricular mediante el uso de analítica avanzada e Inteligencia Artificial.</p>
                    <p class="hero-support">Permite identificar brechas, tendencias, niveles de alineación y oportunidades de mejora en tiempo real para fortalecer la toma de decisiones académicas y la calidad educativa.</p>
                </div>
                """, unsafe_allow_html=True)

            with _c2:
                with st.container(key="hero_right_card"):
                    st.markdown(f"""
                    <div class="upload-panel-deco">
                        <div class="upload-icon-box">{icon_upload}</div>
                        <h4>Selecciona tus archivos</h4>
                        <span>200MB por archivo &middot; XLSX &middot; M&#250;ltiples archivos</span>
                    </div>
                    """, unsafe_allow_html=True)

                    nuevo_upload = st.file_uploader(
                        "Seleccionar archivos",
                        type=['xlsx'],
                        accept_multiple_files=True,
                        key="uploader_main",
                        label_visibility="collapsed"
                    )

                    st.markdown('<p class="upload-help">Máx 200MB por archivo · XLSX · Multiples archivos</p>', unsafe_allow_html=True)

                    if st.button("Estructura del archivo", key="btn_estructura_archivo", use_container_width=True):
                        st.session_state['mostrar_estructura_archivo'] = not st.session_state.get('mostrar_estructura_archivo', False)

                    if st.session_state.get('mostrar_estructura_archivo', False):
                        st.markdown("""
                        <div class="estructura-card">
                        <h5>Estructura del archivo Excel</h5>
                        <ul>
                            <li><strong>Hoja:</strong> Paso 5 Estrategias micro</li>
                            <li><strong>Encabezados:</strong> fila 2</li>
                            <li><strong>Formato:</strong> .xlsx</li>
                        </ul>
                        <table class="estructura-mini-table">
                            <thead><tr><th>Campo</th><th>Ejemplo</th></tr></thead>
                            <tbody>
                                <tr><td>Tipo de Saber</td><td>Saber / SaberHacer / SaberSer</td></tr>
                                <tr><td>Resultado de aprendizaje</td><td>Analiza los fundamentos...</td></tr>
                                <tr><td>Nucleos tematicos</td><td>Derivadas, Integrales</td></tr>
                                <tr><td>Semestre</td><td>1, 2, 3...</td></tr>
                            </tbody>
                        </table>
                    </div>
                    """, unsafe_allow_html=True)

        icon_document = render_icon_svg('document', '#0077C8', 22)
        icon_trend    = render_icon_svg('trend',    '#1fb2de', 22)
        icon_bloom    = render_icon_svg('bloom',    '#7c3aed', 22)
        icon_search   = render_icon_svg('search',   '#059669', 22)
        icon_grid     = render_icon_svg('grid',     '#d97706', 22)
        icon_settings = render_icon_svg('settings', '#dc2626', 22)

        st.markdown(f"""
        <div class="feature-grid">
            <div class="feature-card" style="--card-color:#0077C8;--card-bg:#e8f3ff;">
                <div><div class="feature-icon-box">{icon_document}</div>
                    <h3>Resumen Ejecutivo</h3>
                    <p>Visi&#243;n general del estado de la Matriz de RA y cumplimiento de metas institucionales.</p></div>
                <div class="feature-meta"><span class="status">85% COMPLETADO</span><span class="arrow">&#8250;</span></div>
            </div>
            <div class="feature-card" style="--card-color:#1fb2de;--card-bg:#e0f6ff;">
                <div><div class="feature-icon-box">{icon_trend}</div>
                    <h3>Tendencias Globales</h3>
                    <p>Comparativa con est&#225;ndares internacionales y tem&#225;ticas emergentes en la industria.</p></div>
                <div class="feature-meta"><span class="status">12 NUEVAS</span><span class="arrow">&#8250;</span></div>
            </div>
            <div class="feature-card" style="--card-color:#7c3aed;--card-bg:#f3e8ff;">
                <div><div class="feature-icon-box">{icon_bloom}</div>
                    <h3>Bloom &amp; Integraci&#243;n</h3>
                    <p>Niveles taxon&#243;micos detectados en las competencias y resultados de aprendizaje.</p></div>
                <div class="feature-meta"><span class="status">NIVEL: CREAR</span><span class="arrow">&#8250;</span></div>
            </div>
            <div class="feature-card" style="--card-color:#059669;--card-bg:#d1fae5;">
                <div><div class="feature-icon-box">{icon_search}</div>
                    <h3>Miner&#237;a de Texto</h3>
                    <p>Descubrimiento de patrones y nubes de palabras clave en el contenido curricular.</p></div>
                <div class="feature-meta"><span class="status">PROCESADO</span><span class="arrow">&#8250;</span></div>
            </div>
            <div class="feature-card" style="--card-color:#d97706;--card-bg:#fef3c7;">
                <div><div class="feature-icon-box">{icon_grid}</div>
                    <h3>Tipo de Saber</h3>
                    <p>An&#225;lisis de la balanza entre saber, saber hacer y saber ser en los programas.</p></div>
                <div class="feature-meta"><span class="status">&#8212;</span><span class="arrow">&#8250;</span></div>
            </div>
            <div class="feature-card" style="--card-color:#dc2626;--card-bg:#fee2e2;">
                <div><div class="feature-icon-box">{icon_settings}</div>
                    <h3>Cobertura Tem&#225;tica</h3>
                    <p>Detecci&#243;n de redundancias o vac&#237;os tem&#225;ticos en la malla curricular institucional.</p></div>
                <div class="feature-meta"><span class="status alert">ALERTA BRECHAS</span><span class="arrow">&#8250;</span></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if nuevo_upload is not None and len(nuevo_upload) > 0:
            st.session_state['archivos_subidos'] = nuevo_upload
            st.rerun()

        st.stop()


    # Procesar archivos (usar caché de session_state para evitar reprocesar en cada rerun)
    cache_key = tuple(f.name for f in uploaded_files)
    
    if 'proc_cache_key' in st.session_state and st.session_state['proc_cache_key'] == cache_key:
        # Usar datos cacheados
        df = st.session_state['cached_df']
        failed_list = st.session_state['cached_failed_list']
        totales_oficiales = st.session_state['cached_totales_oficiales']
    else:
        # Procesar archivos y cachear
        df, failed_list = procesar_archivos(uploaded_files)
        totales_oficiales = leer_totales_programa(uploaded_files)
        
        st.session_state['proc_cache_key'] = cache_key
        st.session_state['cached_df'] = df
        st.session_state['cached_failed_list'] = failed_list
        st.session_state['cached_totales_oficiales'] = totales_oficiales
    
    for f in uploaded_files:
        f.seek(0)

    # Barra compacta de archivos cargados + boton cambiar
    col_files, col_btn = st.columns([4, 1])
    with col_files:
        num_cargados = len(uploaded_files) - len(failed_list)
        num_total = len(uploaded_files)
        num_fallidos = len(failed_list)
        if num_fallidos > 0:
            st.success(f"{num_cargados} archivos cargados correctamente, {num_fallidos} archivo(s) no se pudieron cargar de {num_total} archivos seleccionados")
        else:
            st.success(f"{num_cargados} archivos cargados correctamente")
    with col_btn:
        if st.button("Cambiar archivos", use_container_width=True):
            st.session_state['archivos_subidos'] = None
            st.session_state['proc_cache_key'] = None
            st.rerun()
    
    # Mostrar errores SIEMPRE visibles (después de procesar)
    if failed_list:
        st.sidebar.error(f"{len(failed_list)} archivo(s) con error:")
        for f_err in failed_list:
            st.sidebar.warning(f"{f_err['nombre']}: {f_err['causa']}")

    if df.empty:
        st.error(
            "No se pudieron cargar datos. Verifica que los archivos tengan "
            "la hoja 'Paso 5 Estrategias micro' con encabezados en la fila 2."
        )
        st.stop()

    # Info sidebar - simplificada (archivos en area principal)
    st.sidebar.markdown(
        f"""
        <div style='background:rgba(247,148,29,0.15);border:1px solid rgba(247,148,29,0.4);
        border-radius:8px;padding:10px 14px;margin-bottom:8px'>
        <div style='font-size:0.82em;color:#fff'>
        <b>{df['Nombre asignatura o modulo'].nunique()}</b> asignaturas<br>
        <b>{len(df):,}</b> registros procesados
        </div></div>
        """,
        unsafe_allow_html=True
    )

    # MOSTRAR ERRORES EN AREA PRINCIPAL (más visible)
    if failed_list:
        st.error(f"❌ **{len(failed_list)}** archivo(s) no se cargaron:")
        for f_err in failed_list:
            st.warning(f"• {f_err['nombre']}: {f_err['causa']}")

    st.sidebar.markdown("---")
    prog_sel_sidebar = sorted([str(p) for p in df['Programa'].unique() if pd.notna(p)])
    prog_sel = st.sidebar.selectbox("Programa", ["Todos"] + prog_sel_sidebar, key="sel_prog")

    # ── FILTROS GLOBALES en área principal ───────────────────────────────────
    st.markdown("---")
    st.markdown("**🔍 Filtros**")

    # Obtener valores únicos
    modalidades = sorted([str(m) for m in df['Modalidad'].unique() if pd.notna(m)])
    sedes = sorted([str(s) for s in df['Sede'].unique() if pd.notna(s)])
    programas = sorted([str(p) for p in df['Programa'].unique() if pd.notna(p)])
    niveles = sorted([str(n) for n in df['Nivel'].unique() if pd.notna(n)]) if 'Nivel' in df.columns else []

    col_mod, col_sed, col_niv = st.columns([1, 1, 1])
    with col_mod:
        modalidad_sel = st.selectbox("Modalidad", ["Todos"] + modalidades, key="sel_modalidad")
    with col_sed:
        sede_sel = st.selectbox("Sede", ["Todos"] + sedes, key="sel_sede")
    with col_niv:
        nivel_options = niveles if niveles else []
        nivel_detectado_lista = sorted(set(_detectar_nivel(df[df['Programa'] == p].iloc[:1] if p != 'Todos' else df) for p in programas)) if not nivel_options else nivel_options
        nivel_sel = st.selectbox("Nivel", ["Todos"] + nivel_options, key="sel_nivel")

    st.markdown("---")
    
    # Aplicar filtros
    df_filtered = df.copy()
    if modalidad_sel != "Todos":
        df_filtered = df_filtered[df_filtered['Modalidad'] == modalidad_sel]
    if sede_sel != "Todos":
        df_filtered = df_filtered[df_filtered['Sede'] == sede_sel]
    if prog_sel != "Todos":
        df_filtered = df_filtered[df_filtered['Programa'] == prog_sel]
    if nivel_sel != "Todos" and 'Nivel' in df_filtered.columns:
        df_filtered = df_filtered[df_filtered['Nivel'] == nivel_sel]
    
    # Guardar en session_state
    st.session_state['df'] = df
    st.session_state['df_filtered'] = df_filtered
    st.session_state['failed_list'] = failed_list

    # Obtener tendencias
    tendencias = obtener_tendencias()

    # Scroll al top
    st_components.html(
        "<script>window.parent.document.querySelector('.main').scrollTo(0, 0);</script>",
        height=0
    )

    # Renderizar pagina usando datos filtrados para que los filtros afecten todos los artefactos
    if pagina == "Inicio":
        pagina_inicio(df_filtered, totales_oficiales)

    elif pagina == "Tipo de Saber":
        pagina_tipo_saber(df_filtered)

    elif pagina == "Cobertura de Perfil":
        pagina_cobertura_perfil(df_filtered)

    elif pagina == "Cobertura Temática":
        with st.spinner("Analizando cobertura temática..."):
            resultados_cob = analizar_cobertura(df_filtered)
        pagina_cobertura(df_filtered, resultados_cob)

    elif pagina == "Tendencias Globales":
        with st.spinner("Detectando tendencias globales..."):
            resultados_tend = analizar_tendencias(df_filtered, tendencias)
        pagina_tendencias(df_filtered, tendencias, resultados_tend)

    elif pagina == "Minería de Texto":
        with st.spinner("Ejecutando análisis de texto..."):
            resultados_nlp = analizar_nlp(df_filtered)
        pagina_nlp(df_filtered, resultados_nlp)

    elif pagina == "Bloom & Integración":
        taxonomias_bloom = leer_taxonomias_bloom(uploaded_files)
        pagina_bloom_integracion(df_filtered, taxonomias_bloom)

    elif pagina == "Familias Curriculares":
        with st.spinner("Agrupando programas en familias curriculares..."):
            resultados_nlp = analizar_nlp(df_filtered)
        pagina_familias_curriculares(df_filtered, resultados_nlp)

    elif pagina == "Configurar Tendencias":
        pagina_config_tendencias()

    elif pagina == "Explorar Datos":
        pagina_datos(df_filtered)


if __name__ == '__main__':
    main()
